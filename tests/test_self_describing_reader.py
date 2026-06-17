"""Tests for the self-describing Excel reader using the user's example sheets."""

import os

import pytest
import openpyxl

from flextool.process_inputs.read_self_describing_excel import (
    find_crossing_point,
    parse_sheet_metadata,
    detect_and_parse_sheet,
    extract_sheet_data,
    _convert_value,
    _parse_entity_def,
    _parse_filter_def,
    _parse_strict_bool,
    ENTITY_EXISTENCE,
)

EXAMPLE_FILE = "projects/africa2/converted/example_input_template.xlsx"


@pytest.fixture(scope="module")
def workbook():
    if not os.path.exists(EXAMPLE_FILE):
        pytest.skip(
            f"External example workbook {EXAMPLE_FILE!r} is not checked in; "
            f"set the file at this path or convert to a committed synthetic "
            f"fixture to exercise these tests."
        )
    wb = openpyxl.load_workbook(EXAMPLE_FILE, data_only=True)
    yield wb
    wb.close()


# ---------------------------------------------------------------------------
# Unit tests for parsing helpers
# ---------------------------------------------------------------------------


class TestParseEntityDef:
    def test_single_dim(self):
        result = _parse_entity_def("entity: node")
        assert len(result) == 1
        assert result[0].class_name == "node"
        assert result[0].dimensions == ["node"]

    def test_multi_dim(self):
        result = _parse_entity_def("entity: commodity, node")
        assert len(result) == 1
        assert result[0].class_name == "commodity__node"
        assert result[0].dimensions == ["commodity", "node"]

    def test_multi_class(self):
        text = "entity: (unit__inputNode: (unit, node), unit__outputNode: (unit, node))"
        result = _parse_entity_def(text)
        assert len(result) == 2
        assert result[0].class_name == "unit__inputNode"
        assert result[0].dimensions == ["unit", "node"]
        assert result[1].class_name == "unit__outputNode"
        assert result[1].dimensions == ["unit", "node"]


class TestParseFilterDef:
    def test_basic(self):
        text = "filter: {unit__inputNode: ^input$, unit__outputNode: ^output$}"
        result = _parse_filter_def(text)
        assert result["unit__inputNode"] == "^input$"
        assert result["unit__outputNode"] == "^output$"


# ---------------------------------------------------------------------------
# Crossing point detection
# ---------------------------------------------------------------------------


class TestCrossingPoint:
    def test_node_c(self, workbook):
        ws = workbook["node_c"]
        def_row, def_col = find_crossing_point(ws)
        # node_c: row 1 has "description" at col C (index 2)
        # Scanning col C down: C1=description, C2=data type, C3=parameter, C4=empty
        # → def_row = 2 (0-based, which is Excel row 3)
        assert def_col == 2, f"Expected def_col=2, got {def_col}"
        assert def_row == 2, f"Expected def_row=2, got {def_row}"

    def test_unit_node_constraint_c(self, workbook):
        ws = workbook["unit_node_constraint_c"]
        def_row, def_col = find_crossing_point(ws)
        # Row 1 has "description" at col G (index 6)
        # Col G: G1=description, G2=data type, G3=parameter, G4=empty
        assert def_col == 6, f"Expected def_col=6, got {def_col}"
        assert def_row == 2, f"Expected def_row=2, got {def_row}"


# ---------------------------------------------------------------------------
# Sheet metadata parsing
# ---------------------------------------------------------------------------


class TestParseMetadata:
    def test_node_c_metadata(self, workbook):
        ws = workbook["node_c"]
        meta = parse_sheet_metadata(ws)
        assert meta is not None
        assert meta.sheet_name == "node_c"

        # Column definitions
        assert meta.alt_col == 0  # "alternative" at col A
        assert meta.entity_classes[0].class_name == "node"
        assert meta.entity_classes[0].dimensions == ["node"]

        # Entity existence column (recognised from "Entity Alternative" alias)
        assert meta.entity_existence_col is not None

        # Parameter columns should include node_type, inflow, etc.
        param_names = set(meta.param_cols.values())
        assert "node_type" in param_names
        assert "inflow" in param_names
        assert "existing" in param_names
        # entity existence should be in params (mapped from Entity Alternative)
        assert "entity existence" in param_names

        # Data types
        assert any(dt == "string" for dt in meta.data_types.values())
        assert any(dt == "float" for dt in meta.data_types.values())

        # Data starts at row 3 (0-based)
        assert meta.data_start_row == 3

    def test_unit_node_constraint_c_metadata(self, workbook):
        ws = workbook["unit_node_constraint_c"]
        meta = parse_sheet_metadata(ws)
        assert meta is not None

        # Should have two entity classes
        assert len(meta.entity_classes) == 2
        class_names = {ec.class_name for ec in meta.entity_classes}
        assert "unit__inputNode" in class_names
        assert "unit__outputNode" in class_names

        # Should have filter
        assert meta.filter_col is not None
        assert "unit__inputNode" in meta.filter_map
        assert meta.filter_map["unit__inputNode"] == "^input$"

        # Should have index column
        assert meta.index_col is not None
        assert meta.index_name == "constraint"


# ---------------------------------------------------------------------------
# Link sheet
# ---------------------------------------------------------------------------


class TestLinkSheet:
    def test_commodity_node(self, workbook):
        ws = workbook["commodity_node"]
        meta, _is_scenario = detect_and_parse_sheet(ws)
        assert meta is not None
        assert meta.entity_classes[0].class_name == "commodity__node"
        assert meta.entity_classes[0].dimensions == ["commodity", "node"]

    def test_commodity_node_data(self, workbook):
        ws = workbook["commodity_node"]
        meta, _is_scenario = detect_and_parse_sheet(ws)
        data = extract_sheet_data(ws, meta)
        assert len(data.link_entities) >= 2
        # Should contain Coal→Coal_node and Gas→Gas_node
        entity_set = set(data.link_entities)
        assert ("Coal", "Coal_node") in entity_set
        assert ("Gas", "Gas_node") in entity_set


# ---------------------------------------------------------------------------
# Transposed (timeseries) sheet
# ---------------------------------------------------------------------------


class TestTransposedSheet:
    def test_profile_t(self, workbook):
        ws = workbook["profile_t"]
        meta, _is_scenario = detect_and_parse_sheet(ws)
        assert meta is not None
        assert meta.is_transposed is True

        # Entity class should be profile
        assert meta.entity_classes[0].class_name == "profile"

        # Index should be time
        assert meta.index_name == "time"

        # Alternative and entity rows should be found
        assert meta.alt_row is not None
        assert meta.entity_row is not None

    def test_profile_t_data(self, workbook):
        ws = workbook["profile_t"]
        meta, _is_scenario = detect_and_parse_sheet(ws)
        data = extract_sheet_data(ws, meta)
        assert len(data.records) > 0

        # Check that we got records for different profiles
        entities = {r["entity_byname"] for r in data.records}
        assert ("Wind1",) in entities
        assert ("Battery_profile",) in entities

        # All records should have time index
        for r in data.records:
            assert r["index_value"] is not None
            assert r["index_name"] == "time"


# ---------------------------------------------------------------------------
# Data extraction
# ---------------------------------------------------------------------------


class TestDataExtraction:
    def test_node_c_data(self, workbook):
        ws = workbook["node_c"]
        meta, _is_scenario = detect_and_parse_sheet(ws)
        data = extract_sheet_data(ws, meta)
        assert len(data.records) > 0

        # Check we got some expected data
        alternatives = {r["alternative"] for r in data.records}
        assert "Base" in alternatives

        entities = {r["entity_byname"] for r in data.records}
        assert ("Coal_node",) in entities or ("NodeA",) in entities

        # Check entity existence records
        ee_records = [r for r in data.records if r["param_name"] == ENTITY_EXISTENCE]
        assert len(ee_records) > 0

    def test_unit_node_constraint_c_data(self, workbook):
        ws = workbook["unit_node_constraint_c"]
        meta, _is_scenario = detect_and_parse_sheet(ws)
        data = extract_sheet_data(ws, meta)
        assert len(data.records) > 0

        # Should have records from both entity classes
        classes = {r["entity_class"] for r in data.records}
        assert "unit__inputNode" in classes or "unit__outputNode" in classes

        # Check index values (constraint names)
        index_vals = {r["index_value"] for r in data.records if r["index_value"]}
        assert "c01" in index_vals or "gas_export" in index_vals


# ---------------------------------------------------------------------------
# Strict-mode invariants — the v2 reader is a faithful inverse of the
# exporter.  No fuzzy boolean coercion, no header aliases, no force-floats
# in transposed sheets.  These tests pin those invariants without needing
# the external example workbook.
# ---------------------------------------------------------------------------


def _make_ws(wb_data: list[list[str | int | float | None]]):
    """Build a one-shot in-memory worksheet from a 2D Python list."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "probe"
    for r_idx, row in enumerate(wb_data, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    return wb, ws


class TestStrictBooleanParse:
    """``_parse_strict_bool`` accepts only TRUE / FALSE (case-insensitive)."""

    def test_true_accepted(self):
        assert _parse_strict_bool(
            "TRUE", sheet_name="s", row_1based=1, col_1based=1,
            field_name="entity existence",
        ) is True

    def test_false_accepted(self):
        assert _parse_strict_bool(
            "FALSE", sheet_name="s", row_1based=1, col_1based=1,
            field_name="entity existence",
        ) is False

    def test_case_insensitive(self):
        assert _parse_strict_bool(
            "True", sheet_name="s", row_1based=1, col_1based=1,
            field_name="entity existence",
        ) is True
        assert _parse_strict_bool(
            "false", sheet_name="s", row_1based=2, col_1based=2,
            field_name="entity existence",
        ) is False

    def test_yes_rejected(self):
        with pytest.raises(ValueError, match="expected 'TRUE' or 'FALSE'"):
            _parse_strict_bool(
                "yes", sheet_name="node_c", row_1based=5, col_1based=4,
                field_name="entity existence",
            )

    def test_one_rejected(self):
        # Previously coerced silently to True; now an error.
        with pytest.raises(ValueError, match="expected 'TRUE' or 'FALSE'"):
            _parse_strict_bool(
                "1", sheet_name="node_c", row_1based=5, col_1based=4,
                field_name="entity existence",
            )

    def test_error_message_locates_cell(self):
        with pytest.raises(ValueError) as exc:
            _parse_strict_bool(
                "Y", sheet_name="profile_t", row_1based=12, col_1based=7,
                field_name="entity existence",
            )
        msg = str(exc.value)
        assert "profile_t" in msg
        assert "row 12" in msg
        assert "col 7" in msg
        assert "'Y'" in msg


class TestConvertValueBoolean:
    """``_convert_value`` with dtype 'boolean' is strict."""

    def test_true_false(self):
        assert _convert_value("TRUE", "boolean") is True
        assert _convert_value("FALSE", "boolean") is False
        assert _convert_value("true", "boolean") is True

    def test_yes_no_rejected(self):
        # The pre-fix code accepted "yes" → True and silently mapped
        # everything else to False; now both raise.
        with pytest.raises(ValueError, match="expected 'TRUE' or 'FALSE'"):
            _convert_value("yes", "boolean")
        with pytest.raises(ValueError, match="expected 'TRUE' or 'FALSE'"):
            _convert_value("no", "boolean")


class TestConvertValueMapLeafDtype:
    """``_convert_value`` derives the leaf scalar type from the part of the
    dtype before any ``(...)`` container suffix, so Map/Array leaves keep
    their float type instead of round-tripping as strings.
    """

    def test_float_map_and_array_leaves_convert(self):
        for dt in ("float (1d-map)", "float (2d-map)", "float (3d-map)", "float (array)"):
            assert _convert_value("10", dt) == 10.0
            assert isinstance(_convert_value("10", dt), float)

    def test_compound_string_float_prefers_float(self):
        # A param typed string/float that stored a number → float;
        # a genuinely non-numeric cell falls back to the raw string.
        assert _convert_value("3000000", "string/float (1d-map)") == 3000000.0
        assert _convert_value("auto", "string/float (1d-map)") == "auto"

    def test_float_map_keeps_non_finite(self):
        assert _convert_value("inf", "float (1d-map)") == float("inf")

    def test_string_map_leaves_stay_raw(self):
        # No float token → keep verbatim (numeric-looking strings preserved).
        assert _convert_value("01", "string (1d-map)") == "01"

    def test_boolean_array_token_stays_raw(self):
        # 'boolean (array)' cells carry period tokens (round-trip form),
        # NOT TRUE/FALSE — they must not strict-parse and must not raise.
        assert _convert_value("p2025", "boolean (array)") == "p2025"


class TestEntityExistenceHeaderStrict:
    """The legacy ``"entity alternative"`` alias is no longer rewritten.

    The exporter emits exactly ``"entity existence"`` (excel_writer.py:2526).
    A hand-edited sheet carrying the legacy header must not be silently
    rebranded — the column is treated as an unknown parameter so the user
    sees a mismatch instead of a phantom rename.
    """

    def test_legacy_header_not_rewritten(self):
        # Minimal v2 layout with the legacy "Entity Alternative" header
        # in what would be the entity-existence column.
        sheet = [
            ["navigate", None, "description", None, None],
            [None, None, "data type", "string", "string"],
            ["alternative", "entity: node", "parameter",
             "Entity Alternative", "node_type"],
            ["Base", "n1", None, "TRUE", "balance"],
        ]
        _wb, ws = _make_ws(sheet)
        meta = parse_sheet_metadata(ws)
        assert meta is not None
        # The legacy header is NOT promoted to entity_existence_col.
        assert meta.entity_existence_col is None
        # It is kept verbatim as an (unknown) parameter so the rest of
        # the pipeline either ignores it or surfaces a missing-parameter
        # warning — never silently rewritten.
        assert "Entity Alternative" in meta.param_cols.values()


class TestMissingDataTypeDefaultsToString:
    """When a column has no dtype declared, fall back to 'string'."""

    def test_no_dtype_emits_warning_and_returns_string(self, caplog):
        # Layout: declared data-type row covers col D ('declared'),
        # but the unknown-dtype column E gets no dtype at all and no
        # default_data_type either.  Cell should pass through verbatim.
        sheet = [
            ["navigate", None, "description", None, None],
            [None, None, "data type", "string", None],
            ["alternative", "entity: node", "parameter",
             "declared", "no_dtype"],
            ["Base", "n1", None, "abc", "01"],
        ]
        _wb, ws = _make_ws(sheet)
        meta = parse_sheet_metadata(ws)
        assert meta is not None
        with caplog.at_level("WARNING"):
            data = extract_sheet_data(ws, meta)
        # Find the no_dtype record
        no_dtype_recs = [r for r in data.records if r["param_name"] == "no_dtype"]
        assert no_dtype_recs, "expected a no_dtype record"
        # Value is preserved as a string ("01" — NOT coerced to 1.0).
        assert no_dtype_recs[0]["value"] == "01"
        # Warning was emitted about the missing data type.
        assert any(
            "no data type declared" in rec.message
            for rec in caplog.records
        )


class TestEmptyAlternativePropagatesNone:
    """Entity-only records preserve a real ``None`` alt (no '' synthesis)."""

    def test_entity_only_row_has_none_alt(self):
        # Layout: alternative column is empty, no parameter values → only
        # an entity-only record should be emitted.
        sheet = [
            ["navigate", None, "description", None],
            [None, None, "data type", "float"],
            ["alternative", "entity: node", "parameter", "inflow"],
            [None, "n1", None, None],  # no alt, no param value
        ]
        _wb, ws = _make_ws(sheet)
        meta = parse_sheet_metadata(ws)
        assert meta is not None
        data = extract_sheet_data(ws, meta)
        # We expect exactly one entity-only record for n1.
        ent_recs = [r for r in data.records if r["entity_byname"] == ("n1",)]
        assert ent_recs, "expected an entity-only record for n1"
        assert ent_recs[0]["param_name"] == ""
        assert ent_recs[0]["alternative"] is None


class TestTransposedTripletDtypeRespected:
    """Transposed extractor uses the triplet's data-type, not 'float'."""

    def test_string_array_keeps_strings(self):
        # Minimal transposed sheet with a string (array) triplet — values
        # "01" and "02" must survive as strings, NOT coerce to 1.0 / 2.0.
        sheet = [
            ["index: period", "parameter: tier_label | data type: string (array)",
             None],
            [None, "entity: solve", "s1"],
            ["index:", "alternative", "base"],
            ["p1", None, "01"],
            ["p2", None, "02"],
        ]
        _wb, ws = _make_ws(sheet)
        meta, _ = detect_and_parse_sheet(ws)
        assert meta is not None and meta.is_transposed
        data = extract_sheet_data(ws, meta)
        vals = {r["value"] for r in data.records}
        # Strings preserved verbatim — would have been 1.0 / 2.0 before.
        assert "01" in vals and "02" in vals
        assert 1.0 not in vals and 2.0 not in vals
