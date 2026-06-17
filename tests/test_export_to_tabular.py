"""Tests for the FlexTool Spine DB to Excel exporter.

Verifies database reading, sheet specification building, and Excel output
against the real examples.sqlite database.
"""

from __future__ import annotations

import openpyxl
import pytest

from flextool.export_to_tabular.db_reader import DatabaseContents, read_database
from flextool.export_to_tabular.sheet_config import build_sheet_specs, SheetSpec, load_settings
from flextool.export_to_tabular.export_to_excel import export_to_excel

from flextool._resources import package_data_path
from flextool.update_flextool.initialize_database import initialize_database

MASTER_TEMPLATE = package_data_path("schemas/spinedb_schema.json")

# Parameters on split_params classes that are intentionally not surfaced in the
# Excel whitelist (e.g. handled by a dedicated writer or never user-editable).
# Keys are entity-class names; values are the param names to exclude.
WHITELIST_EXEMPT_PARAMS: dict[str, set[str]] = {
    "model": {"version"},  # written by write_version_sheet, not a user param
}


def _find_def_row(ws: "openpyxl.worksheet.worksheet.Worksheet") -> int:
    """Return the row number of the v2 parameter-name header row.

    The header sits at row 3 by default but shifts to row 4 when any
    parameter on the sheet has a default value (an optional 'default'
    metadata row is inserted between data type and the header).
    """
    for r in (3, 4):
        if ws.cell(row=r, column=1).value == "alternative":
            return r
    raise AssertionError(
        f"Could not locate parameter-name header row in sheet '{ws.title}': "
        f"row 3 col A = {ws.cell(row=3, column=1).value!r}, "
        f"row 4 col A = {ws.cell(row=4, column=1).value!r}"
    )


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def example_db_url(tmp_path_factory: pytest.TempPathFactory) -> str:
    """Materialize ``templates_examples.json`` into a fresh tmp SQLite.

    Avoids reading the user-facing ``templates/examples.sqlite``
    directly — that file would otherwise need to be regenerated on
    every schema bump just to keep these tests green, clobbering any
    user edits in the process.
    """
    json_src = package_data_path("schemas/canonical_databases/templates_examples.json")
    db_dir = tmp_path_factory.mktemp("examples_db")
    sqlite_path = db_dir / "examples.sqlite"
    initialize_database(str(json_src), str(sqlite_path))
    return f"sqlite:///{sqlite_path}"


@pytest.fixture(scope="module")
def db_contents(example_db_url: str) -> DatabaseContents:
    """Read the example database once for all tests in this module."""
    return read_database(example_db_url)


@pytest.fixture(scope="module")
def sheet_specs(db_contents: DatabaseContents) -> list[SheetSpec]:
    """Build sheet specs once for all tests in this module."""
    return build_sheet_specs(db_contents)


@pytest.fixture(scope="module")
def exported_workbook(
    tmp_path_factory: pytest.TempPathFactory,
    db_contents: DatabaseContents,
    example_db_url: str,
) -> openpyxl.Workbook:
    """Export to a temporary Excel file and return the opened workbook."""
    out_dir = tmp_path_factory.mktemp("export")
    out_path = out_dir / "test_output.xlsx"
    export_to_excel(example_db_url, str(out_path), include_advanced=True)
    wb = openpyxl.load_workbook(str(out_path))
    yield wb
    wb.close()


# ---------------------------------------------------------------------------
# Test 1: test_read_database
# ---------------------------------------------------------------------------

class TestReadDatabase:
    """Verify that read_database produces a fully populated DatabaseContents."""

    def test_entity_class_count(self, db_contents: DatabaseContents) -> None:
        assert len(db_contents.entity_classes) >= 17, (
            f"Expected >= 17 entity classes, got {len(db_contents.entity_classes)}"
        )

    def test_key_entities_exist(self, db_contents: DatabaseContents) -> None:
        for cls_name in ("node", "unit", "connection"):
            assert cls_name in db_contents.entities, f"Missing entities for class '{cls_name}'"
            assert len(db_contents.entities[cls_name]) > 0, f"No entities for class '{cls_name}'"

    def test_parameter_definitions_exist(self, db_contents: DatabaseContents) -> None:
        assert len(db_contents.parameter_definitions) > 0, "No parameter definitions loaded"
        assert "node" in db_contents.parameter_definitions
        assert len(db_contents.parameter_definitions["node"]) > 0

    def test_parameter_values_populated(self, db_contents: DatabaseContents) -> None:
        assert len(db_contents.parameter_values) > 0, "No parameter values loaded"

    def test_alternatives_non_empty(self, db_contents: DatabaseContents) -> None:
        assert len(db_contents.alternatives) > 0, "No alternatives loaded"

    def test_scenarios_populated(self, db_contents: DatabaseContents) -> None:
        assert len(db_contents.scenarios) > 0, "No scenarios loaded"

    def test_version_is_number(self, db_contents: DatabaseContents) -> None:
        assert db_contents.version is not None, "Version is None"
        assert isinstance(db_contents.version, (int, float)), (
            f"Version should be a number, got {type(db_contents.version)}"
        )


# ---------------------------------------------------------------------------
# Test 2: test_build_sheet_specs
# ---------------------------------------------------------------------------

class TestBuildSheetSpecs:
    """Verify that build_sheet_specs produces correct sheet specifications."""

    def test_total_sheet_count(self, sheet_specs: list[SheetSpec]) -> None:
        count = len(sheet_specs)
        assert 55 <= count <= 65, f"Expected 55-65 sheets, got {count}"

    def test_specific_sheets_exist(self, sheet_specs: list[SheetSpec]) -> None:
        names = {s.sheet_name for s in sheet_specs}
        for expected in ("node_c", "node_p", "node_t", "unit_c", "scenario", "navigate"):
            assert expected in names, f"Sheet '{expected}' not found. Available: {sorted(names)}"

    def test_unit_node_c_has_direction_column(self, sheet_specs: list[SheetSpec]) -> None:
        spec = next(s for s in sheet_specs if s.sheet_name == "unit_node_c")
        assert spec.direction_column == "input_output", (
            f"unit_node_c direction_column should be 'input_output', got '{spec.direction_column}'"
        )

    def test_connection_c_has_extra_entity_columns(self, sheet_specs: list[SheetSpec]) -> None:
        spec = next(s for s in sheet_specs if s.sheet_name == "connection_c")
        assert spec.extra_entity_columns == ["left_node", "right_node"], (
            f"connection_c extra_entity_columns should be ['left_node', 'right_node'], "
            f"got {spec.extra_entity_columns}"
        )

    def test_unit_node_constraint_c_has_unpack_index_column(self, sheet_specs: list[SheetSpec]) -> None:
        spec = next(s for s in sheet_specs if s.sheet_name == "unit_node_constraint_c")
        assert spec.unpack_index_column == "constraint", (
            f"unit_node_constraint_c unpack_index_column should be 'constraint', "
            f"got '{spec.unpack_index_column}'"
        )

    def test_commodity_node_has_link_layout(self, sheet_specs: list[SheetSpec]) -> None:
        spec = next(s for s in sheet_specs if s.sheet_name == "commodity_node")
        assert spec.layout == "link", (
            f"commodity_node layout should be 'link', got '{spec.layout}'"
        )

    def test_parameter_ordering_on_node_c(self, sheet_specs: list[SheetSpec], db_contents: DatabaseContents) -> None:
        spec = next(s for s in sheet_specs if s.sheet_name == "node_c")
        params = spec.parameter_names
        assert len(params) > 0, "node_c has no parameter_names"


# ---------------------------------------------------------------------------
# Regression: a param eligible for BOTH constant and periodic layouts (e.g.
# node.existing — float for some entities, 1d-map for others) is split across
# a ``*_c`` constant sheet and a ``*_p`` periodic sheet.  Its scalar value
# belongs ONLY on the constant sheet; the periodic writer must NOT also emit a
# scalar-only row (empty period cell), or the value duplicates and collides on
# import round-trip.  This is generic (keyed off the spec/param type-lists),
# not hard-coded to ``node.existing``.
# ---------------------------------------------------------------------------


def _periodic_index_and_param_cols(
    ws: "openpyxl.worksheet.worksheet.Worksheet", def_row: int
) -> "tuple[int, dict[str, int]]":
    """Return (index_col, {param_name: col}) for a v2 periodic sheet."""
    index_col = None
    param_start = None
    for c in range(1, ws.max_column + 1):
        label = str(ws.cell(row=def_row, column=c).value or "").strip()
        if label.lower().startswith("index:"):
            index_col = c
        if label.lower() == "parameter":
            param_start = c
            break
    assert index_col is not None and param_start is not None
    param_cols: dict[str, int] = {}
    for c in range(param_start + 1, ws.max_column + 1):
        label = str(ws.cell(row=def_row, column=c).value or "").strip()
        if not label:
            break
        param_cols[label] = c
    return index_col, param_cols


class TestScalarSiblingNoPeriodicDuplication:
    """Constant-eligible scalars must not leak onto the periodic sibling sheet."""

    def test_periodic_specs_mark_constant_sibling_params(
        self, sheet_specs: list[SheetSpec]
    ) -> None:
        """Every periodic spec flags exactly the params shared with a constant
        sibling over the same entity classes — generically, with at least one
        real overlap present in the example DB (e.g. node.existing)."""
        const_params: dict[frozenset, set[str]] = {}
        for s in sheet_specs:
            if s.layout == "constant":
                key = frozenset(s.entity_classes)
                const_params.setdefault(key, set()).update(s.parameter_names)

        total_overlap = 0
        for s in sheet_specs:
            if s.layout != "periodic":
                continue
            siblings = const_params.get(frozenset(s.entity_classes), set())
            expected = set(s.parameter_names) & siblings
            assert s.scalar_params_on_constant_sibling == expected, (
                f"{s.sheet_name}: scalar_params_on_constant_sibling="
                f"{s.scalar_params_on_constant_sibling}, expected {expected}"
            )
            total_overlap += len(expected)
        assert total_overlap > 0, (
            "Example DB exercises no constant/periodic-shared param; the "
            "regression no longer covers the original bug."
        )

    def test_node_existing_is_a_constant_sibling_param(
        self, sheet_specs: list[SheetSpec]
    ) -> None:
        spec = next(s for s in sheet_specs if s.sheet_name == "node_p")
        assert "existing" in spec.scalar_params_on_constant_sibling

    def test_no_scalar_only_rows_for_constant_sibling_params(
        self, exported_workbook: openpyxl.Workbook, sheet_specs: list[SheetSpec]
    ) -> None:
        """In every exported periodic sheet, no data row carries an empty
        period cell together with a filled value for a constant-sibling param
        (that filled cell would be a duplicated scalar)."""
        periodic = {s.sheet_name: s for s in sheet_specs if s.layout == "periodic"}
        checked = 0
        for sheet_name, spec in periodic.items():
            if not spec.scalar_params_on_constant_sibling:
                continue
            if sheet_name not in exported_workbook.sheetnames:
                continue
            ws = exported_workbook[sheet_name]
            def_row = _find_def_row(ws)
            index_col, param_cols = _periodic_index_and_param_cols(ws, def_row)
            for pname in spec.scalar_params_on_constant_sibling:
                col = param_cols.get(pname)
                if col is None:
                    continue
                for r in range(def_row + 1, ws.max_row + 1):
                    idx = ws.cell(row=r, column=index_col).value
                    val = ws.cell(row=r, column=col).value
                    if (idx is None or str(idx).strip() == "") and val not in (None, ""):
                        raise AssertionError(
                            f"{sheet_name} row {r}: scalar-only row leaked for "
                            f"constant-sibling param '{pname}' (value={val!r})"
                        )
                checked += 1
        assert checked > 0, "No constant-sibling periodic param columns checked"


# ---------------------------------------------------------------------------
# Test 3: test_export_produces_valid_excel
# ---------------------------------------------------------------------------

class TestExportProducesValidExcel:
    """Verify that export_to_excel produces a valid Excel file with all sheets."""

    def test_all_expected_sheets_exist(self, exported_workbook: openpyxl.Workbook, sheet_specs: list[SheetSpec]) -> None:
        wb_sheets = set(exported_workbook.sheetnames)
        spec_sheets = {s.sheet_name for s in sheet_specs}
        # 'version' is generated as a spec but written separately — exclude from check
        spec_sheets.discard("version")
        missing = spec_sheets - wb_sheets
        assert not missing, f"Missing sheets in exported workbook: {missing}"

    def test_each_sheet_has_header_row(self, exported_workbook: openpyxl.Workbook) -> None:
        for name in exported_workbook.sheetnames:
            ws = exported_workbook[name]
            row1_values = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            has_content = any(v is not None for v in row1_values)
            assert has_content, f"Sheet '{name}' has no content in row 1"


# ---------------------------------------------------------------------------
# Test 4: test_node_c_content (v2 self-describing format)
# ---------------------------------------------------------------------------

class TestNodeCContent:
    """Verify detailed content of the node_c sheet in v2 format.

    V2 format:
      Row 1: navigate | (blank) | description texts...
      Row 2: (blank)  | (blank) | data type | type values...
      Row 3: alternative | entity: node | parameter | param names...
      Row 4+: data rows
    """

    def test_header_row_structure(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["node_c"]
        # Definition row is 3 by default, 4 when an optional 'default' row is inserted.
        def_row = _find_def_row(ws)
        assert ws.cell(row=def_row, column=1).value == "alternative", (
            f"Row {def_row}, Col 1 should be 'alternative', "
            f"got '{ws.cell(row=def_row, column=1).value}'"
        )
        assert ws.cell(row=def_row, column=2).value == "entity: node", (
            f"Row {def_row}, Col 2 should be 'entity: node', "
            f"got '{ws.cell(row=def_row, column=2).value}'"
        )

    def test_navigate_in_row_1(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["node_c"]
        assert ws.cell(row=1, column=1).value == "navigate"

    def test_expected_node_entities(self, exported_workbook: openpyxl.Workbook, db_contents: DatabaseContents) -> None:
        ws = exported_workbook["node_c"]
        # In v2, entity names are in column 2, starting from row 4
        node_names_in_sheet: set[str] = set()
        for row in range(4, ws.max_row + 1):
            val = ws.cell(row=row, column=2).value
            if val is not None:
                node_names_in_sheet.add(str(val))

        db_node_entities = db_contents.entities.get("node", [])
        db_node_names = {e["entity_byname"][0] for e in db_node_entities}
        present = db_node_names & node_names_in_sheet
        assert len(present) > 0, (
            f"No DB node entities found in node_c sheet. "
            f"DB nodes: {db_node_names}, sheet nodes: {node_names_in_sheet}"
        )

    def test_description_row_has_content(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["node_c"]
        # Row 1 has descriptions starting from column 3
        desc_values = [ws.cell(row=1, column=c).value for c in range(3, ws.max_column + 1)]
        descriptions = [v for v in desc_values if v is not None and v != "navigate"]
        assert len(descriptions) > 0, "Row 1 (description row) has no parameter descriptions"

    def test_data_type_row(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["node_c"]
        # Row 2 should have 'data type' label
        assert ws.cell(row=2, column=3).value == "data type", (
            f"Row 2, Col 3 should be 'data type', got '{ws.cell(row=2, column=3).value}'"
        )


# ---------------------------------------------------------------------------
# Test 5: test_node_t_content (v2 transposed timeseries)
# ---------------------------------------------------------------------------

class TestNodeTContent:
    """Verify that node_t has the correct v2 transposed timeseries layout.

    V2 format (multi-param sheets carry a per-column 'data type' row so the
    float leaf type survives the round-trip):
      Row 1: navigate    | info text...
      Row 2: (blank)      | data type    | dtypes...
      Row 3: (blank)      | entity: node | entity names...
      Row 4: (blank)      | alternative  | alt names...
      Row 5: index: time  | parameter    | param names...
      Row 6+: time values | data...

    Row positions shift with the presence of the data-type/default rows, so
    the tests locate header rows by their column-B label rather than a fixed
    index.
    """

    @staticmethod
    def _label_row(ws: "openpyxl.worksheet.worksheet.Worksheet", label: str) -> int | None:
        for r in range(1, min(10, ws.max_row) + 1):
            if str(ws.cell(row=r, column=2).value or "").strip().lower() == label:
                return r
        return None

    def test_transposed_layout_structure(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["node_t"]
        ent_row = self._label_row(ws, "entity: node")
        alt_row = self._label_row(ws, "alternative")
        assert ent_row is not None, "node_t: no 'entity: node' label row found"
        assert alt_row is not None, "node_t: no 'alternative' label row found"

    def test_data_type_row_present(self, exported_workbook: openpyxl.Workbook) -> None:
        """Multi-param _t sheets must declare a per-column data type so float
        time-series do not round-trip as string-valued Maps."""
        ws = exported_workbook["node_t"]
        dt_row = self._label_row(ws, "data type")
        assert dt_row is not None, "node_t: missing 'data type' header row"
        dtypes = [
            str(ws.cell(row=dt_row, column=c).value or "")
            for c in range(3, ws.max_column + 1)
        ]
        assert any("float" in d for d in dtypes), (
            f"node_t: expected at least one float column data type, got {dtypes}"
        )

    def test_time_index_label(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["node_t"]
        # 'index: time' sits on column A of the last header row (the
        # parameter row); locate it rather than assuming a fixed position.
        found = any(
            str(ws.cell(row=r, column=1).value or "").strip() == "index: time"
            for r in range(1, min(10, ws.max_row) + 1)
        )
        assert found, "node_t: 'index: time' label not found in column A"

    def test_entity_names_in_headers(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["node_t"]
        # Entity names in row 2, columns 3+
        entity_names: list[str] = []
        for col in range(3, ws.max_column + 1):
            val = ws.cell(row=2, column=col).value
            if val is not None:
                entity_names.append(str(val))
        assert len(entity_names) > 0, "node_t: no entity names found in row 2"

    def test_time_values_exist(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["node_t"]
        # Row 5+ should have time index values in column 1
        time_values: list[str] = []
        for row in range(5, min(15, ws.max_row + 1)):
            val = ws.cell(row=row, column=1).value
            if val is not None:
                time_values.append(str(val))
        assert len(time_values) > 0, "node_t: no time index values found"


# ---------------------------------------------------------------------------
# Test 6: test_scenario_sheet
# ---------------------------------------------------------------------------

class TestScenarioSheet:
    """Verify the scenario sheet structure."""

    def test_scenario_names_in_row_2(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["scenario"]
        scenario_names: list[str] = []
        for col in range(2, ws.max_column + 1):
            val = ws.cell(row=2, column=col).value
            if val is not None:
                scenario_names.append(str(val))
        assert len(scenario_names) > 0, "scenario sheet: no scenario names in row 2"

    def test_base_alternative_in_row_3(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["scenario"]
        label = ws.cell(row=3, column=1).value
        assert label == "base_alternative", (
            f"scenario sheet: row 3, col 1 should be 'base_alternative', got '{label}'"
        )

    def test_alternatives_in_subsequent_rows(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["scenario"]
        if ws.max_row >= 4:
            label = ws.cell(row=4, column=1).value
            assert label is not None and "alternative" in str(label).lower(), (
                f"scenario sheet: row 4, col 1 should be an alternative label, got '{label}'"
            )


# ---------------------------------------------------------------------------
# Test 7: test_link_sheet (v2 format)
# ---------------------------------------------------------------------------

class TestLinkSheet:
    """Verify the commodity_node link sheet in v2 format.

    V2 format:
      Row 1: navigate | ...
      Row 2: entity: commodity, node | commodity | node
      Row 3+: data rows
    """

    def test_commodity_node_headers(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["commodity_node"]
        # Row 2 has the column headers in v2
        assert ws.cell(row=2, column=2).value == "commodity", (
            f"commodity_node row 2 col 2 should be 'commodity', got '{ws.cell(row=2, column=2).value}'"
        )
        assert ws.cell(row=2, column=3).value == "node", (
            f"commodity_node row 2 col 3 should be 'node', got '{ws.cell(row=2, column=3).value}'"
        )

    def test_expected_link_entries(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["commodity_node"]
        pairs: set[tuple[str, str]] = set()
        for row in range(3, ws.max_row + 1):
            c = ws.cell(row=row, column=2).value
            n = ws.cell(row=row, column=3).value
            if c is not None and n is not None:
                pairs.add((str(c), str(n)))
        assert len(pairs) > 0, "No commodity-node pairs found. Sheet is empty."


# ---------------------------------------------------------------------------
# Test 8: test_constraint_sheet (v2 format)
# ---------------------------------------------------------------------------

class TestConstraintSheet:
    """Verify the unit_node_constraint_c sheet in v2 format.

    V2 format:
      Row 1: navigate | ... | description | desc texts...
      Row 2: ... | data type | types...
      Row 3: alternative | entity: unit, node | unit | node | input_output | index: constraint | parameter | param names...
      Row 4+: data rows
    """

    def test_constraint_index_column(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["unit_node_constraint_c"]
        # Find the 'index: constraint' column in row 3
        headers = [ws.cell(row=3, column=c).value for c in range(1, ws.max_column + 1)]
        assert "index: constraint" in headers, (
            f"unit_node_constraint_c: 'index: constraint' not found in row 3. Headers: {headers}"
        )

    def test_constraint_flow_coeff_values(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["unit_node_constraint_c"]
        # Find parameter names in row 3
        {ws.cell(row=3, column=c).value: c for c in range(1, ws.max_column + 1)}
        assert "constraint_flow_coeff" in [ws.cell(row=3, column=c).value for c in range(1, ws.max_column + 1)], (
            "unit_node_constraint_c: 'constraint_flow_coeff' not found in row 3 headers"
        )

    def test_has_data_rows(self, exported_workbook: openpyxl.Workbook) -> None:
        ws = exported_workbook["unit_node_constraint_c"]
        # Should have data starting from row 4
        has_data = False
        for row in range(4, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val is not None:
                has_data = True
                break
        assert has_data, "unit_node_constraint_c: no data rows found"


# ---------------------------------------------------------------------------
# Test 9: schema ⇄ export_settings.yaml sync for split_params classes
# ---------------------------------------------------------------------------

class TestSplitParamsSchemaSync:
    """Guard against schema drift on classes that use explicit param whitelists.

    Classes listed in ``split_params`` use hardcoded param lists rather than
    schema discovery, so a parameter added via migration is silently dropped
    from the Excel export until someone updates ``export_settings.yaml``.
    This test catches that class of gap in both directions.
    """

    @pytest.fixture(scope="class")
    def schema_params_by_class(self) -> dict[str, set[str]]:
        import json
        assert MASTER_TEMPLATE.exists(), f"Master template not found: {MASTER_TEMPLATE}"
        data = json.loads(MASTER_TEMPLATE.read_text())
        result: dict[str, set[str]] = {}
        for row in data.get("parameter_definitions", []):
            cls, pname = row[0], row[1]
            result.setdefault(cls, set()).add(pname)
        return result

    @pytest.fixture(scope="class")
    def listed_params_by_class(self) -> dict[str, set[str]]:
        settings = load_settings()
        result: dict[str, set[str]] = {}
        for cls, sub_groups in settings.get("split_params", {}).items():
            for _sub_name, rule in sub_groups.items():
                result.setdefault(cls, set()).update(rule.get("params", []))
        return result

    def test_no_schema_params_missing_from_yaml(
        self,
        schema_params_by_class: dict[str, set[str]],
        listed_params_by_class: dict[str, set[str]],
    ) -> None:
        gaps: dict[str, set[str]] = {}
        for cls, listed in listed_params_by_class.items():
            schema = schema_params_by_class.get(cls, set())
            exempt = WHITELIST_EXEMPT_PARAMS.get(cls, set())
            missing = schema - listed - exempt
            if missing:
                gaps[cls] = missing
        assert not gaps, (
            "Schema parameters missing from export_settings.yaml split_params "
            f"whitelists: {gaps}. Add each to the appropriate sub-group or to "
            "WHITELIST_EXEMPT_PARAMS if intentionally not surfaced."
        )

    def test_no_yaml_params_missing_from_schema(
        self,
        schema_params_by_class: dict[str, set[str]],
        listed_params_by_class: dict[str, set[str]],
    ) -> None:
        stale: dict[str, set[str]] = {}
        for cls, listed in listed_params_by_class.items():
            schema = schema_params_by_class.get(cls, set())
            extra = listed - schema
            if extra:
                stale[cls] = extra
        assert not stale, (
            "export_settings.yaml split_params lists parameters not in the "
            f"master template schema: {stale}. Remove or rename."
        )


# ---------------------------------------------------------------------------
# Phase B: ladder routing for facet-leaf parameters (price/quantity)
# ---------------------------------------------------------------------------


class TestLadderRouting:
    """``commodity.price_ladder_*`` must route via the dedicated ``ladder``
    layout, NOT via ``stochastic``.

    The two ``{price, quantity}`` facets surface as two real Excel
    parameter columns (``parameter: price``, ``parameter: quantity``)
    sharing the entity + tier ([+ period]) index columns.  The DB-side
    encoding is a nested Map whose innermost level is the facet axis;
    that combination happens at the import boundary.
    """

    def test_classify_returns_ladder_for_price_ladder_cumulative(self) -> None:
        from flextool.export_to_tabular.sheet_config import classify_param_types

        types = classify_param_types(
            ("2d_map",),
            entity_class="commodity",
            param_name="price_ladder_cumulative",
        )
        assert types == {"ladder"}, (
            f"Expected {{'ladder'}} for commodity.price_ladder_cumulative; "
            f"got {types}"
        )

    def test_classify_returns_ladder_for_price_ladder_annual(self) -> None:
        from flextool.export_to_tabular.sheet_config import classify_param_types

        # price_ladder_annual admits both 2d_map and 3d_map per the
        # registry; both must collapse to {ladder} (no stochastic /
        # timeseries leakage from the schema-declared 3d_map type).
        types = classify_param_types(
            ("2d_map", "3d_map"),
            entity_class="commodity",
            param_name="price_ladder_annual",
        )
        assert types == {"ladder"}
        assert "stochastic" not in types
        assert "timeseries" not in types

    def test_classify_without_registry_keeps_stochastic(self) -> None:
        """A genuinely stochastic 3d_map param (no registry entry)
        still routes to ``stochastic`` — confirms the registry override
        is scoped to (entity_class, param_name) pairs we tagged."""
        from flextool.export_to_tabular.sheet_config import classify_param_types

        types = classify_param_types(
            ("3d_map",),
            entity_class="unit__inputNode",
            param_name="profile",
        )
        assert "stochastic" in types
        assert "ladder" not in types

    def test_sheet_specs_include_ladder_sheets_for_commodity(
        self, sheet_specs: list[SheetSpec],
    ) -> None:
        labels = {(s.sheet_name, s.layout) for s in sheet_specs}
        assert ("price_ladder_cumulative", "ladder") in labels, (
            f"Expected price_ladder_cumulative sheet with layout=ladder; "
            f"got: {sorted(labels)}"
        )
        assert ("price_ladder_annual", "ladder") in labels

    def test_no_commodity_stochastic_sheet_for_price_ladder(
        self, sheet_specs: list[SheetSpec],
    ) -> None:
        """The facet-leaf params must NOT also surface on a duplicated
        ``commodity_s`` / ``commodity_s_*`` sheet — that would let two
        writers claim the same (entity, param, alt) keys."""
        for spec in sheet_specs:
            if spec.sheet_name.startswith("commodity_s") and spec.layout == "stochastic":
                assert not any(
                    p in ("price_ladder_annual", "price_ladder_cumulative")
                    for p in spec.parameter_names
                ), (
                    f"Stochastic sheet {spec.sheet_name} carries facet-leaf "
                    f"params {spec.parameter_names}"
                )

    def test_exported_workbook_has_ladder_sheets(
        self, exported_workbook: openpyxl.Workbook,
    ) -> None:
        assert "price_ladder_cumulative" in exported_workbook.sheetnames
        assert "price_ladder_annual" in exported_workbook.sheetnames

    def test_price_ladder_cumulative_sheet_shape(
        self, exported_workbook: openpyxl.Workbook,
    ) -> None:
        """Depth-2 sheet: alternative | entity: commodity | index: tier
        | parameter | price | quantity."""
        ws = exported_workbook["price_ladder_cumulative"]
        assert ws.cell(row=3, column=1).value == "alternative"
        assert ws.cell(row=3, column=2).value == "entity: commodity"
        assert ws.cell(row=3, column=3).value == "index: tier"
        assert ws.cell(row=3, column=4).value == "parameter"
        assert ws.cell(row=3, column=5).value == "price"
        assert ws.cell(row=3, column=6).value == "quantity"
        # Data type row sets per-facet types.
        assert ws.cell(row=2, column=5).value == "float"
        assert ws.cell(row=2, column=6).value == "float"

    def test_price_ladder_annual_sheet_has_period_col(
        self, exported_workbook: openpyxl.Workbook,
    ) -> None:
        """Depth-3 sheet adds an ``index: period`` column between entity
        and tier."""
        ws = exported_workbook["price_ladder_annual"]
        assert ws.cell(row=3, column=3).value == "index: period"
        assert ws.cell(row=3, column=4).value == "index: tier"
        assert ws.cell(row=3, column=5).value == "parameter"
        assert ws.cell(row=3, column=6).value == "price"
        assert ws.cell(row=3, column=7).value == "quantity"

    def test_price_ladder_inf_is_string_sentinel(
        self, exported_workbook: openpyxl.Workbook,
    ) -> None:
        """``quantity = inf`` must be written as the string ``"inf"`` —
        openpyxl drops actual non-finite floats."""
        ws = exported_workbook["price_ladder_cumulative"]
        # Scan data rows for the unbounded tail tier.
        found_inf = False
        for r in range(4, ws.max_row + 1):
            qty = ws.cell(row=r, column=6).value
            if qty == "inf":
                found_inf = True
                break
        assert found_inf, (
            "Expected the canonical coal/tier-2 row to carry quantity='inf'; "
            "no such row found."
        )

    def test_price_ladder_round_trip_byte_identical(
        self, tmp_path_factory: pytest.TempPathFactory, example_db_url: str,
    ) -> None:
        """End-to-end: canonical → export → import → identical Maps."""
        import json
        import subprocess
        import sys

        from spinedb_api import DatabaseMapping, from_database, to_database

        out_dir = tmp_path_factory.mktemp("ladder_rt")
        xlsx = out_dir / "rt.xlsx"
        export_to_excel(example_db_url, str(xlsx), include_advanced=True)
        rt_db = out_dir / "rt.sqlite"
        initialize_database(str(MASTER_TEMPLATE), str(rt_db))
        subprocess.run(
            [
                sys.executable,
                "-m",
                "flextool.cli.cmd_read_self_describing_tabular_input",
                str(xlsx),
                f"sqlite:///{rt_db}",
            ],
            check=True,
        )

        def _ladder_values(url: str) -> dict:
            out = {}
            with DatabaseMapping(url) as db:
                for p in db.get_parameter_value_items():
                    if p["parameter_definition_name"] not in (
                        "price_ladder_annual",
                        "price_ladder_cumulative",
                    ):
                        continue
                    k = (
                        p["entity_class_name"],
                        p["parameter_definition_name"],
                        tuple(p["entity_byname"]),
                        p["alternative_name"],
                    )
                    out[k] = (p["value"], p["type"])
            return out

        def _to_json(rec):
            val = from_database(rec[0], rec[1])
            v, _ = to_database(val)
            return json.loads(v) if isinstance(v, (bytes, str)) else v

        src_vals = _ladder_values(example_db_url)
        rt_vals = _ladder_values(f"sqlite:///{rt_db}")
        assert src_vals, "Canonical DB has no price_ladder data — fixture drift."
        assert set(src_vals) == set(rt_vals), (
            f"price_ladder key set differs across round-trip: "
            f"src-only={set(src_vals) - set(rt_vals)}, "
            f"rt-only={set(rt_vals) - set(src_vals)}"
        )
        for k in src_vals:
            assert _to_json(src_vals[k]) == _to_json(rt_vals[k]), (
                f"Map JSON differs for {k}: src={_to_json(src_vals[k])} "
                f"rt={_to_json(rt_vals[k])}"
            )


class TestLadderReaderUnit:
    """Reader-side unit tests for the ladder layout."""

    def test_reader_emits_price_quantity_records(
        self, tmp_path_factory: pytest.TempPathFactory,
    ) -> None:
        """A synthesised minimal price_ladder_cumulative sheet must
        produce one record per (commodity, alt, tier, facet)."""
        from flextool.process_inputs.read_self_describing_excel import (
            read_self_describing_excel,
        )

        out_dir = tmp_path_factory.mktemp("ladder_reader")
        xlsx_path = out_dir / "mini.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "price_ladder_cumulative"
        ws.cell(row=1, column=1, value="navigate")
        ws.cell(row=1, column=4, value="description")
        ws.cell(row=2, column=4, value="data type")
        ws.cell(row=2, column=5, value="float")
        ws.cell(row=2, column=6, value="float")
        ws.cell(row=3, column=1, value="alternative")
        ws.cell(row=3, column=2, value="entity: commodity")
        ws.cell(row=3, column=3, value="index: tier")
        ws.cell(row=3, column=4, value="parameter")
        ws.cell(row=3, column=5, value="price")
        ws.cell(row=3, column=6, value="quantity")
        ws.cell(row=4, column=1, value="base")
        ws.cell(row=4, column=2, value="coal")
        ws.cell(row=4, column=3, value="1")
        ws.cell(row=4, column=5, value=20.0)
        ws.cell(row=4, column=6, value=1.0)
        ws.cell(row=5, column=1, value="base")
        ws.cell(row=5, column=2, value="coal")
        ws.cell(row=5, column=3, value="2")
        ws.cell(row=5, column=5, value=30.0)
        ws.cell(row=5, column=6, value="inf")
        wb.save(str(xlsx_path))
        wb.close()

        sheets = read_self_describing_excel(str(xlsx_path))
        assert len(sheets) == 1
        recs = sheets[0].records
        keyed = {
            (r["entity_byname"], r["param_name"], r["index_value"]): r["value"]
            for r in recs
        }
        assert keyed[(("coal",), "price", "1")] == 20.0
        assert keyed[(("coal",), "quantity", "1")] == 1.0
        assert keyed[(("coal",), "price", "2")] == 30.0
        # 'inf' sentinel must round-trip to float infinity.
        assert keyed[(("coal",), "quantity", "2")] == float("inf")

    def test_writer_to_db_combines_facets(
        self, tmp_path_factory: pytest.TempPathFactory,
    ) -> None:
        """Feed price/quantity records to write_sheet_data_to_db; expect
        a single ``commodity.price_ladder_cumulative`` nested Map."""
        from spinedb_api import DatabaseMapping, from_database

        from flextool.process_inputs.read_self_describing_excel import (
            EntityClassDef,
            SheetData,
            SheetMetadata,
        )
        from flextool.process_inputs.write_self_describing_to_db import (
            write_sheet_data_to_db,
        )

        out_dir = tmp_path_factory.mktemp("ladder_writer")
        db_path = out_dir / "rt.sqlite"
        initialize_database(str(MASTER_TEMPLATE), str(db_path))
        db_url = f"sqlite:///{db_path}"

        meta = SheetMetadata(sheet_name="price_ladder_cumulative")
        meta.entity_classes = [EntityClassDef("commodity", ["commodity"])]
        sheet = SheetData(sheet_name=meta.sheet_name, metadata=meta)
        for tier, price, quantity in [("1", 20.0, 1.0), ("2", 30.0, float("inf"))]:
            for facet, val in (("price", price), ("quantity", quantity)):
                sheet.records.append({
                    "alternative": "base",
                    "entity_class": "commodity",
                    "entity_byname": ("coal",),
                    "param_name": facet,
                    "value": val,
                    "index_value": tier,
                    "index_name": "tier",
                    "data_type": "float",
                })

        write_sheet_data_to_db(
            [sheet], db_url, purge_first=False, keep_entities=True,
        )
        with DatabaseMapping(db_url) as db:
            ladder = [
                p for p in db.get_parameter_value_items()
                if p["parameter_definition_name"] == "price_ladder_cumulative"
                and p["entity_byname"] == ("coal",)
            ]
        assert len(ladder) == 1
        val = from_database(ladder[0]["value"], ladder[0]["type"])
        # Outer Map indexed by tier
        assert list(val.indexes) == ["1", "2"]
        assert val.index_name == "tier"
        leaf1 = val.values[0]
        assert list(leaf1.indexes) == ["price", "quantity"]
        assert list(leaf1.values) == [20.0, 1.0]
        leaf2 = val.values[1]
        assert leaf2.values[1] == float("inf")


class TestRoundTripGenericRegressions:
    """End-to-end export -> import regressions for the generic round-trip
    corruptions found via the H2_trade model.

    The pre-existing ``test_price_ladder_round_trip_byte_identical`` only
    exercised a single cumulative ladder, so it never surfaced the
    annual/cumulative collision, the lexical tier ordering, the dropped
    decomposition knobs, or the vanishing empty Array.  This builds a DB
    from the live schema (never a checked-in .sqlite) carrying all four
    shapes at once and asserts they survive a real workbook round-trip.
    """

    @pytest.fixture(scope="class")
    def round_tripped(self, tmp_path_factory: pytest.TempPathFactory) -> dict:
        from spinedb_api import (
            Array, DatabaseMapping, Map, from_database, to_database,
        )

        out = tmp_path_factory.mktemp("generic_rt")
        src = out / "src.sqlite"
        initialize_database(str(MASTER_TEMPLATE), str(src))
        src_url = f"sqlite:///{src}"

        def _ladder(tiers: list[tuple[str, float, float]]) -> Map:
            # Map(tier -> Map(price/quantity)).  ``tiers`` intentionally
            # unsorted so the import-side ordering is tested.
            return Map(
                indexes=[t for t, _, _ in tiers],
                values=[
                    Map(indexes=["price", "quantity"], values=[p, q])
                    for _, p, q in tiers
                ],
                index_name="tier",
            )

        with DatabaseMapping(src_url) as db:
            db.add_alternative(name="override")
            for ent in ("h2_a", "co2_a"):
                db.add_entity(entity_class_name="commodity", entity_byname=(ent,))
            for solve in ("s_main", "s_child"):
                db.add_entity(entity_class_name="solve", entity_byname=(solve,))

            def put(ec, bn, pn, alt, value):
                v, t = to_database(value)
                db.add_parameter_value(
                    entity_class_name=ec, entity_byname=bn,
                    parameter_definition_name=pn, alternative_name=alt,
                    value=v, type=t,
                )

            # Annual + cumulative ladders coexisting, both 2d (tier,) — the
            # shape that collided.  Tiers given out of order to test sorting.
            put("commodity", ("h2_a",), "price_method", "base", "price_ladder_annual")
            put("commodity", ("h2_a",), "price_ladder_annual", "base",
                _ladder([("2", 2.0, 20.0), ("10", 10.0, 100.0), ("1", 1.0, 10.0)]))
            put("commodity", ("co2_a",), "price_method", "base", "price_ladder_cumulative")
            put("commodity", ("co2_a",), "price_ladder_cumulative", "base",
                _ladder([("1", 5.0, 0.5), ("2", 9.0, float("inf"))]))

            # solve: decomposition knob + an empty contains_solves override.
            put("solve", ("s_main",), "decomposition", "base", "lagrangian")
            put("solve", ("s_main",), "contains_solves", "base", Array(["s_child"]))
            put("solve", ("s_main",), "contains_solves", "override", Array([]))
            db.commit_session("setup")

        xlsx = out / "rt.xlsx"
        export_to_excel(src_url, str(xlsx), include_advanced=True)
        rt = out / "rt.sqlite"
        initialize_database(str(MASTER_TEMPLATE), str(rt))
        from flextool.process_inputs.read_self_describing_excel import (
            read_self_describing_excel,
        )
        from flextool.process_inputs.write_self_describing_to_db import (
            write_sheet_data_to_db,
        )
        sheets = read_self_describing_excel(str(xlsx), skip_sheets={"navigate", "version"})
        write_sheet_data_to_db(sheets, f"sqlite:///{rt}", purge_first=True, keep_entities=True)

        result: dict = {}
        with DatabaseMapping(f"sqlite:///{rt}") as db:
            for p in db.get_parameter_value_items():
                key = (
                    p["parameter_definition_name"],
                    tuple(p["entity_byname"]),
                    p["alternative_name"],
                )
                result[key] = from_database(p["value"], p["type"])
        return result

    def test_annual_ladder_keeps_its_identity(self, round_tripped: dict) -> None:
        # Was relabelled to price_ladder_cumulative before the fix.
        assert ("price_ladder_annual", ("h2_a",), "base") in round_tripped
        assert ("price_ladder_cumulative", ("h2_a",), "base") not in round_tripped

    def test_cumulative_ladder_keeps_its_identity(self, round_tripped: dict) -> None:
        assert ("price_ladder_cumulative", ("co2_a",), "base") in round_tripped
        assert ("price_ladder_annual", ("co2_a",), "base") not in round_tripped

    def test_ladder_tiers_in_numeric_order(self, round_tripped: dict) -> None:
        val = round_tripped[("price_ladder_annual", ("h2_a",), "base")]
        # Numeric, not the lexical "1","10","2".
        assert list(val.indexes) == ["1", "2", "10"]

    def test_decomposition_survives(self, round_tripped: dict) -> None:
        assert round_tripped.get(("decomposition", ("s_main",), "base")) == "lagrangian"

    def test_empty_contains_solves_override_survives(self, round_tripped: dict) -> None:
        from spinedb_api import Array
        key = ("contains_solves", ("s_main",), "override")
        assert key in round_tripped, (
            "empty contains_solves override was dropped on round-trip"
        )
        val = round_tripped[key]
        assert isinstance(val, Array)
        assert list(val.values) == []

    def test_nonempty_contains_solves_unaffected(self, round_tripped: dict) -> None:
        val = round_tripped[("contains_solves", ("s_main",), "base")]
        assert list(val.values) == ["s_child"]


class TestGenericNestedMapRoundTrip:
    """Round-trip for GENERIC (non-facet) multi-index nested float Maps.

    The self-describing exporter emits a depth-N nested-float Map (e.g. a 3d
    ``profile``) onto a stochastic ``_s`` sheet with one left-side ``index:``
    column per Map level plus a value column.  The reader previously rebuilt
    only a single index level, silently dropping the value on re-import.
    These tests cover the reconstruction of the full
    ``Map(idx0 -> Map(idx1 -> ... -> leaf))`` for both depth-3 (end-to-end
    export -> import) and depth-2 (synthetic SheetData -> DB writer).
    """

    @staticmethod
    def _profile_3d() -> "object":
        from spinedb_api import Map

        # branch -> analysis_time -> time -> float.  Top-level branch keys
        # are intentionally NOT alphabetical so the importer's order
        # preservation (string axis = insertion order, not sorted) is tested;
        # the ragged second branch (one analysis_time) tests sparse levels.
        return Map(
            indexes=["realized", "f1"],
            values=[
                Map(
                    indexes=["t0001", "t0002"],
                    values=[
                        Map(indexes=["t01", "t02"], values=[0.5, 0.7],
                            index_name="time"),
                        Map(indexes=["t01", "t02"], values=[0.1, 0.2],
                            index_name="time"),
                    ],
                    index_name="analysis_time",
                ),
                Map(
                    indexes=["t0001"],
                    values=[
                        Map(indexes=["t01", "t02"], values=[0.9, 0.8],
                            index_name="time"),
                    ],
                    index_name="analysis_time",
                ),
            ],
            index_name="branch",
        )

    def test_depth3_profile_round_trip_byte_identical(
        self, tmp_path_factory: pytest.TempPathFactory,
    ) -> None:
        """A 3d nested-float ``profile`` Map survives export -> import with
        its nested structure, every index key, per-level index names, and
        float leaves intact (byte-identical Spine JSON)."""
        import json

        from spinedb_api import DatabaseMapping, from_database, to_database

        from flextool.export_to_tabular.export_to_excel import export_to_excel
        from flextool.process_inputs.read_self_describing_excel import (
            read_self_describing_excel,
        )
        from flextool.process_inputs.write_self_describing_to_db import (
            write_sheet_data_to_db,
        )

        out = tmp_path_factory.mktemp("nested3d_rt")
        src = out / "src.sqlite"
        initialize_database(str(MASTER_TEMPLATE), str(src))
        src_url = f"sqlite:///{src}"
        with DatabaseMapping(src_url) as db:
            db.add_alternative(name="Base")
            db.add_entity(entity_class_name="profile", entity_byname=("wind1",))
            value, type_ = to_database(self._profile_3d())
            db.add_parameter_value(
                entity_class_name="profile", entity_byname=("wind1",),
                parameter_definition_name="profile", alternative_name="Base",
                value=value, type=type_,
            )
            db.commit_session("setup")

        xlsx = out / "out.xlsx"
        export_to_excel(src_url, str(xlsx))

        # The stochastic _s sheet must yield one record per leaf.  The
        # fixture is ragged: realized has 2 analysis_times x 2 times (4
        # leaves), f1 has 1 analysis_time x 2 times (2 leaves) = 6 total.
        sheets = read_self_describing_excel(str(xlsx))
        s_sheet = next(s for s in sheets if s.sheet_name == "profile_s")
        assert len(s_sheet.records) == 6
        # Every record carries the two outer index levels in extra_index_values.
        for rec in s_sheet.records:
            extra = rec.get("extra_index_values")
            assert extra and [a for a, _ in extra] == ["branch", "analysis_time"]
            assert rec["index_name"] == "time"
            assert isinstance(rec["value"], float)

        rt = out / "rt.sqlite"
        initialize_database(str(MASTER_TEMPLATE), str(rt))
        rt_url = f"sqlite:///{rt}"
        write_sheet_data_to_db(
            sheets, rt_url, purge_first=True, keep_entities=True,
        )

        def _profile(url: str):
            with DatabaseMapping(url) as db:
                for p in db.get_parameter_value_items():
                    if (p["parameter_definition_name"] == "profile"
                            and p["entity_byname"] == ("wind1",)):
                        return from_database(p["value"], p["type"])
            return None

        src_val = _profile(src_url)
        rt_val = _profile(rt_url)
        assert rt_val is not None, "3d profile vanished on round-trip"
        # Structural spot check: a deep leaf survives with its float value.
        assert rt_val.index_name == "branch"
        assert list(rt_val.indexes) == ["realized", "f1"]
        f1_at = rt_val.values[1]
        assert f1_at.index_name == "analysis_time"
        assert f1_at.values[0].values[1] == 0.8  # f1 / t0001 / t02
        # Byte-identical Spine JSON (float diffs at ~1e-15 OK, structure exact).
        src_json = json.loads(to_database(src_val)[0])
        rt_json = json.loads(to_database(rt_val)[0])
        assert src_json == rt_json, (
            f"3d profile Map JSON differs on round-trip:\n"
            f"src={json.dumps(src_json)}\nrt ={json.dumps(rt_json)}"
        )

    def test_depth2_generic_nested_map_reconstruction(
        self, tmp_path_factory: pytest.TempPathFactory,
    ) -> None:
        """A synthetic multi-``index:`` SheetData (two index levels per leaf)
        rebuilds ``Map(idx0 -> Map(idx1 -> float))`` with the correct
        per-level index names, float leaves, and authoring (insertion) order
        for the string outer axis."""
        from spinedb_api import DatabaseMapping, from_database

        from flextool.process_inputs.read_self_describing_excel import (
            EntityClassDef,
            SheetData,
            SheetMetadata,
        )
        from flextool.process_inputs.write_self_describing_to_db import (
            write_sheet_data_to_db,
        )

        out = tmp_path_factory.mktemp("nested2d_rt")
        db_path = out / "rt.sqlite"
        initialize_database(str(MASTER_TEMPLATE), str(db_path))
        db_url = f"sqlite:///{db_path}"
        with DatabaseMapping(db_url) as db:
            db.add_alternative(name="Base")
            db.add_entity(entity_class_name="profile", entity_byname=("w1",))
            db.commit_session("setup")

        meta = SheetMetadata(
            sheet_name="profile_s", is_transposed=True, is_stochastic=True,
        )
        meta.entity_classes = [EntityClassDef("profile", ["profile"])]
        sheet = SheetData(sheet_name="profile_s", metadata=meta)
        # branch (string, NOT alphabetical) -> time -> float.
        leaves = {
            ("realized", "t01"): 0.5, ("realized", "t02"): 0.7,
            ("f9", "t01"): 0.9, ("f9", "t02"): 0.8,
        }
        for (branch, tval), leaf in leaves.items():
            sheet.records.append({
                "alternative": "Base", "entity_class": "profile",
                "entity_byname": ("w1",), "param_name": "profile",
                "value": leaf, "index_value": tval, "index_name": "time",
                "data_type": "float",
                "extra_index_values": [("branch", branch)],
            })

        write_sheet_data_to_db(
            [sheet], db_url, purge_first=False, keep_entities=True,
        )

        with DatabaseMapping(db_url) as db:
            vals = [
                from_database(p["value"], p["type"])
                for p in db.get_parameter_value_items()
                if p["parameter_definition_name"] == "profile"
            ]
        assert len(vals) == 1
        top = vals[0]
        assert top.index_name == "branch"
        # Insertion order preserved — a string axis is NOT alphabetised.
        assert list(top.indexes) == ["realized", "f9"]
        realized = top.values[0]
        assert realized.index_name == "time"
        assert list(realized.indexes) == ["t01", "t02"]
        assert list(realized.values) == [0.5, 0.7]
        assert all(isinstance(v, float) for v in realized.values)
        f9 = top.values[1]
        assert list(f9.values) == [0.9, 0.8]

    def test_numeric_outer_axis_sorts_numerically(
        self, tmp_path_factory: pytest.TempPathFactory,
    ) -> None:
        """An integer-like outer axis orders 1,2,…,10 (numeric), not the
        lexical 1,10,2 a plain string sort would yield."""
        from spinedb_api import DatabaseMapping, from_database

        from flextool.process_inputs.read_self_describing_excel import (
            EntityClassDef,
            SheetData,
            SheetMetadata,
        )
        from flextool.process_inputs.write_self_describing_to_db import (
            write_sheet_data_to_db,
        )

        out = tmp_path_factory.mktemp("nested_numeric")
        db_path = out / "rt.sqlite"
        initialize_database(str(MASTER_TEMPLATE), str(db_path))
        db_url = f"sqlite:///{db_path}"
        with DatabaseMapping(db_url) as db:
            db.add_alternative(name="Base")
            db.add_entity(entity_class_name="profile", entity_byname=("w1",))
            db.commit_session("setup")

        meta = SheetMetadata(
            sheet_name="profile_s", is_transposed=True, is_stochastic=True,
        )
        meta.entity_classes = [EntityClassDef("profile", ["profile"])]
        sheet = SheetData(sheet_name="profile_s", metadata=meta)
        # Outer axis given out of order: 2, 10, 1.
        for branch in ("2", "10", "1"):
            sheet.records.append({
                "alternative": "Base", "entity_class": "profile",
                "entity_byname": ("w1",), "param_name": "profile",
                "value": float(branch), "index_value": "t01",
                "index_name": "time", "data_type": "float",
                "extra_index_values": [("branch", branch)],
            })

        write_sheet_data_to_db(
            [sheet], db_url, purge_first=False, keep_entities=True,
        )
        with DatabaseMapping(db_url) as db:
            top = next(
                from_database(p["value"], p["type"])
                for p in db.get_parameter_value_items()
                if p["parameter_definition_name"] == "profile"
            )
        assert list(top.indexes) == ["1", "2", "10"]


class TestRepresentativePeriodWeightsRoundTrip:
    """``timeset.representative_period_weights`` is a rank-2 nested Map
    (base period -> representative period -> weight).  The schema declares
    it as ``2d_map``, which type inference would route to the periodic
    writer (where the Map value is silently dropped).  ``export_settings``
    pins its ``timeset_s`` sub-group to the stochastic layout so it lands
    on the multi-index writer and survives a real workbook round-trip.
    """

    @staticmethod
    def _weights_map() -> "object":
        """Mirror the shape produced by
        ``representative_periods.preprocess._build_weights_map``:
        ``Map(base_start -> Map(rep_start -> weight))``, sparse, with the
        base axis given out of timestep order to exercise key/order
        preservation."""
        from spinedb_api import Map

        return Map(
            indexes=["t0001", "t0169", "t0337"],
            values=[
                Map(indexes=["t0001"], values=[2.5],
                    index_name="representative_period"),
                Map(indexes=["t0001", "t0337"], values=[1.25, 3.75],
                    index_name="representative_period"),
                Map(indexes=["t0337"], values=[4.0],
                    index_name="representative_period"),
            ],
            index_name="base_period",
        )

    @pytest.fixture(scope="class")
    def round_tripped(self, tmp_path_factory: pytest.TempPathFactory) -> dict:
        from spinedb_api import DatabaseMapping, from_database, to_database

        from flextool.process_inputs.read_self_describing_excel import (
            read_self_describing_excel,
        )
        from flextool.process_inputs.write_self_describing_to_db import (
            write_sheet_data_to_db,
        )

        out = tmp_path_factory.mktemp("rpw_rt")
        src = out / "src.sqlite"
        initialize_database(str(MASTER_TEMPLATE), str(src))
        src_url = f"sqlite:///{src}"
        with DatabaseMapping(src_url) as db:
            db.add_alternative(name="Base")
            db.add_entity(entity_class_name="timeset", entity_byname=("rp_set",))
            value, type_ = to_database(self._weights_map())
            db.add_parameter_value(
                entity_class_name="timeset", entity_byname=("rp_set",),
                parameter_definition_name="representative_period_weights",
                alternative_name="Base", value=value, type=type_,
            )
            db.commit_session("setup")

        xlsx = out / "out.xlsx"
        export_to_excel(src_url, str(xlsx), include_advanced=True)

        rt = out / "rt.sqlite"
        initialize_database(str(MASTER_TEMPLATE), str(rt))
        rt_url = f"sqlite:///{rt}"
        sheets = read_self_describing_excel(
            str(xlsx), skip_sheets={"navigate", "version"}
        )
        write_sheet_data_to_db(sheets, rt_url, purge_first=True, keep_entities=True)

        got = None
        with DatabaseMapping(rt_url) as db:
            for p in db.get_parameter_value_items():
                if p["parameter_definition_name"] == "representative_period_weights":
                    got = from_database(p["value"], p["type"])
        return {"xlsx": xlsx, "value": got}

    def test_routes_to_stochastic_sheet(
        self, sheet_specs: list[SheetSpec],
    ) -> None:
        labels = {(s.sheet_name, s.layout) for s in sheet_specs}
        assert ("timeset_s", "stochastic") in labels, (
            f"Expected timeset_s sheet with layout=stochastic; got: "
            f"{sorted(labels)}"
        )

    def test_exported_sheet_has_two_index_columns(
        self, round_tripped: dict,
    ) -> None:
        ws = openpyxl.load_workbook(round_tripped["xlsx"])["timeset_s"]
        idx_labels = [
            c.value
            for row in ws.iter_rows()
            for c in row
            if isinstance(c.value, str) and c.value.lower().startswith("index:")
        ]
        assert idx_labels == [
            "index: base_period",
            "index: representative_period",
        ], idx_labels

    def test_nested_map_survives_byte_identical(
        self, round_tripped: dict,
    ) -> None:
        import json

        from spinedb_api import to_database

        got = round_tripped["value"]
        assert got is not None, (
            "representative_period_weights vanished on round-trip"
        )
        src_json = json.loads(to_database(self._weights_map())[0])
        rt_json = json.loads(to_database(got)[0])
        assert src_json == rt_json, (
            f"representative_period_weights Map JSON differs on round-trip:\n"
            f"src={json.dumps(src_json)}\nrt ={json.dumps(rt_json)}"
        )
