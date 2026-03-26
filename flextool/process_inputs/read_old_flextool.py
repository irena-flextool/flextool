"""Read old-format FlexTool Excel files (.xlsm) and return structured data.

This module parses the 18-sheet .xlsm format used by earlier versions of FlexTool
and returns an OldFlexToolData dataclass containing all parsed information, ready
for database import.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)

# Sheets that should be skipped entirely
_SKIP_SHEETS = frozenset({
    "info", "calc", "dropdown values", "Getting Started",
    "Sensitivity scenarios", "Sensitivity definitions",
    "Settings and filters", "Tool Information",
})

# Mapping from normalized unit-sheet column headers to UnitInstance field names.
# Headers not listed here go into extra_params.
_UNIT_COLUMN_MAP: dict[str, str] = {
    "unitgroup": "unit_group",
    "unit type": "unit_type",
    "fuel": "fuel",
    "cf profile": "cf_profile",
    "inflow": "inflow_profile",
    # Grid/node columns — may appear as separate columns or combined with "/"
    "input grid": "input_grid",
    "input node": "input_node",
    "output grid": "output_grid",
    "output node": "output_node",
    "output2 grid": "output2_grid",
    "output2 node": "output2_node",
    # Capacity and investment
    "capacity (mw)": "capacity_mw",
    "invested capacity (mw)": "invested_capacity_mw",
    "max invest (mw)": "max_invest_mw",
    "storage (mwh)": "storage_mwh",
    "invested storage (mwh)": "invested_storage_mwh",
    "max invest (mwh)": "max_invest_mwh",
    "storage start": "storage_start",
    "storage start/finish": "storage_start",
    "storage finish": "storage_finish",
    "reserve increase ratio": "reserve_increase_ratio",
    "use efficiency time series": "use_efficiency_ts",
    # Output2 / CHP constraint coefficients
    "output2 eq coeff": "output2_eq_coeff",
    "output2 eq constant": "output2_eq_constant",
    "output2 gt coeff": "output2_gt_coeff",
    "output2 gt constant": "output2_gt_constant",
    "output2 lt coeff": "output2_lt_coeff",
    "output2 lt constant": "output2_lt_constant",
    "output2 max capacity": "output2_max_capacity",
    "output2 max capacity (mw)": "output2_max_capacity",
    "output2 max capacity ratio": "output2_max_capacity",
    "fueluse increase eq x output2": "fueluse_increase_eq",
    "inflow multiplier": "inflow_multiplier",
}

# Columns that need special grid/node splitting
_UNIT_GRID_NODE_COLS: dict[str, tuple[str, str]] = {
    "input grid/node": ("input_grid", "input_node"),
    "output grid/node": ("output_grid", "output_node"),
    "output2 grid/node": ("output2_grid", "output2_node"),
}

# String-typed fields in UnitInstance (everything else is float | None)
_UNIT_STRING_FIELDS = frozenset({
    "unit_group", "unit_type", "fuel", "cf_profile", "inflow_profile",
    "input_grid", "input_node", "output_grid", "output_node",
    "output2_grid", "output2_node",
})


# ---------------------------------------------------------------------------
# Dataclasses
# ---------------------------------------------------------------------------

@dataclass
class SensitivityOverride:
    """A single parameter override from the Sensitivity definitions sheet."""
    section: str          # "master", "nodeGroup", "gridNode", "unit_type", "fuel", "unitGroup", "units", "nodeNode"
    entity_ids: dict[str, str]  # entity identifier columns, e.g. {"nodeGroup": "reserve1"}
    param_name: str       # column header name, e.g. "co2_cost", "invested capacity (MW)"
    value: float | str    # the override value


@dataclass
class MasterParams:
    """Global model parameters from 'master' sheet."""
    params: dict[str, float]


@dataclass
class GridNode:
    """A node definition from 'gridNode' sheet."""
    grid: str
    node: str
    node_groups: list[str]
    demand_mwh: float | None
    import_mwh: float | None
    capacity_margin_mw: float | None
    non_synchronous_share: float | None
    use_ts_reserve: float | None
    use_dynamic_reserve: float | None
    print_results: float | None


@dataclass
class NodeGroup:
    """A node group from 'nodeGroup' sheet."""
    name: str
    capacity_margin_mw: float | None
    non_synchronous_share: float | None
    inertia_limit_mws: float | None
    use_ts_reserve: float | None
    use_dynamic_reserve: float | None


@dataclass
class UnitType:
    """A technology template from 'unit_type' sheet."""
    name: str
    params: dict[str, float | None]


@dataclass
class Fuel:
    """A fuel definition from 'fuel' sheet."""
    name: str
    price_per_mwh: float
    co2_content: float


@dataclass
class UnitGroup:
    """A unit group from 'unitGroup' sheet."""
    name: str
    max_invest_mw: float | None
    min_invest_mw: float | None
    max_invest_mwh: float | None
    min_invest_mwh: float | None


@dataclass
class UnitInstance:
    """A unit instance from 'units' sheet."""
    unit_group: str | None
    unit_type: str
    fuel: str | None
    cf_profile: str | None
    inflow_profile: str | None
    input_grid: str | None
    input_node: str | None
    output_grid: str | None
    output_node: str | None
    capacity_mw: float | None
    invested_capacity_mw: float | None
    max_invest_mw: float | None
    storage_mwh: float | None
    invested_storage_mwh: float | None
    max_invest_mwh: float | None
    storage_start: float | None
    storage_finish: float | None
    reserve_increase_ratio: float | None
    use_efficiency_ts: float | None
    output2_grid: str | None
    output2_node: str | None
    output2_eq_coeff: float | None
    output2_eq_constant: float | None
    output2_gt_coeff: float | None
    output2_gt_constant: float | None
    output2_lt_coeff: float | None
    output2_lt_constant: float | None
    output2_max_capacity: float | None
    fueluse_increase_eq: float | None
    inflow_multiplier: float | None
    extra_params: dict[str, float | None] = field(default_factory=dict)


@dataclass
class NodeNodeConnection:
    """A transmission connection from 'nodeNode' sheet."""
    grid: str
    node1: str
    node2: str
    cap_rightward_mw: float | None
    cap_leftward_mw: float | None
    invested_capacity_mw: float | None
    max_invest_mw: float | None
    loss: float | None
    invest_cost_per_kw: float | None
    lifetime: float | None
    interest: float | None
    is_hvdc: bool


@dataclass
class TimeSeriesData:
    """A time series: dict mapping time_id (str) to value (float)."""
    name: str
    data: dict[str, float]


@dataclass
class UnitTimeSeries:
    """Time series for a specific unit parameter."""
    grid: str
    node: str
    unit: str
    param_name: str
    data: dict[str, float]


@dataclass
class DemandTimeSeries:
    """Demand time series for a specific grid/node."""
    grid: str
    node: str
    data: dict[str, float]


@dataclass
class TimeStep:
    """A time step definition from 'ts_time' sheet."""
    time_id: str
    in_use: bool
    time_jump: float
    in_use_invest: bool
    time_jump_invest: float


@dataclass
class OldFlexToolData:
    """Complete parsed data from one old FlexTool .xlsm file."""
    master: MasterParams
    grid_nodes: list[GridNode]
    node_groups: list[NodeGroup]
    unit_types: list[UnitType]
    fuels: list[Fuel]
    unit_groups: list[UnitGroup]
    units: list[UnitInstance]
    connections: list[NodeNodeConnection]
    cf_profiles: list[TimeSeriesData]
    inflow_profiles: list[TimeSeriesData]
    demand_ts: list[DemandTimeSeries]
    import_ts: list[DemandTimeSeries]
    reserve_node_ts: list[TimeSeriesData]
    reserve_group_ts: list[TimeSeriesData]
    unit_ts: list[UnitTimeSeries]
    time_steps: list[TimeStep]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe_float(value: Any) -> float | None:
    """Convert a cell value to float, returning None for empty/non-numeric cells."""
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _safe_str(value: Any) -> str | None:
    """Convert a cell value to a stripped string, returning None for empty cells."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _row_is_empty(row: tuple[Any, ...]) -> bool:
    """Return True if every value in the row is None."""
    return all(v is None for v in row)


def _split_grid_node(value: Any) -> tuple[str | None, str | None]:
    """Split a 'grid/node' string into (grid, node). Returns (None, None) for empty."""
    if value is None:
        return None, None
    s = str(value).strip()
    if not s:
        return None, None
    if "/" in s:
        parts = s.split("/", 1)
        return parts[0].strip(), parts[1].strip()
    return s, None


def _get_sheet(wb: openpyxl.Workbook, name: str) -> Any | None:
    """Get a worksheet by name, returning None if it doesn't exist."""
    if name in wb.sheetnames:
        return wb[name]
    logger.warning("Sheet '%s' not found in workbook; skipping.", name)
    return None


# ---------------------------------------------------------------------------
# Sheet readers
# ---------------------------------------------------------------------------

def _read_master(wb: openpyxl.Workbook) -> MasterParams:
    """Read the 'master' sheet into MasterParams."""
    ws = _get_sheet(wb, "master")
    params: dict[str, float] = {}
    if ws is None:
        return MasterParams(params=params)

    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue  # skip header
        if _row_is_empty(row):
            continue
        param_name = _safe_str(row[0]) if len(row) > 0 else None
        raw_value = row[1] if len(row) > 1 else None
        if param_name is not None:
            val = _safe_float(raw_value)
            if val is not None:
                params[param_name] = val
            else:
                logger.warning("master: non-numeric value for '%s': %r", param_name, raw_value)
    return MasterParams(params=params)


def _read_grid_nodes(wb: openpyxl.Workbook) -> list[GridNode]:
    """Read the 'gridNode' sheet."""
    ws = _get_sheet(wb, "gridNode")
    if ws is None:
        return []

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Header row
    headers = [_safe_str(h) for h in rows[0]]
    result: list[GridNode] = []

    def _col(name: str) -> int | None:
        """Find column index by header name (case-insensitive)."""
        name_lower = name.lower()
        for i, h in enumerate(headers):
            if h is not None and h.lower() == name_lower:
                return i
        return None

    def _get(row: tuple[Any, ...], col_idx: int | None) -> Any:
        if col_idx is None or col_idx >= len(row):
            return None
        return row[col_idx]

    c_grid = _col("grid")
    c_node = _col("node")
    c_ng1 = _col("nodeGroup")
    c_ng2 = _col("nodeGroup2")
    c_ng3 = _col("nodeGroup3")
    c_demand = _col("demand (MWh)")
    c_import = _col("import (MWh)")
    c_cap_margin = _col("capacity margin (MW)")
    c_nonsync = _col("non synchronous share")
    c_ts_reserve = _col("use ts_reserve")
    c_dyn_reserve = _col("use dynamic reserve")
    c_print = _col("print results")

    for row in rows[1:]:
        if _row_is_empty(row):
            continue
        grid = _safe_str(_get(row, c_grid))
        node = _safe_str(_get(row, c_node))
        if grid is None or node is None:
            continue

        groups: list[str] = []
        for c in (c_ng1, c_ng2, c_ng3):
            g = _safe_str(_get(row, c))
            if g:
                groups.append(g)

        result.append(GridNode(
            grid=grid,
            node=node,
            node_groups=groups,
            demand_mwh=_safe_float(_get(row, c_demand)),
            import_mwh=_safe_float(_get(row, c_import)),
            capacity_margin_mw=_safe_float(_get(row, c_cap_margin)),
            non_synchronous_share=_safe_float(_get(row, c_nonsync)),
            use_ts_reserve=_safe_float(_get(row, c_ts_reserve)),
            use_dynamic_reserve=_safe_float(_get(row, c_dyn_reserve)),
            print_results=_safe_float(_get(row, c_print)),
        ))
    return result


def _read_node_groups(wb: openpyxl.Workbook) -> list[NodeGroup]:
    """Read the 'nodeGroup' sheet."""
    ws = _get_sheet(wb, "nodeGroup")
    if ws is None:
        return []

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [_safe_str(h) for h in rows[0]]
    result: list[NodeGroup] = []

    def _col(name: str) -> int | None:
        name_lower = name.lower()
        for i, h in enumerate(headers):
            if h is not None and h.lower() == name_lower:
                return i
        return None

    def _get(row: tuple[Any, ...], col_idx: int | None) -> Any:
        if col_idx is None or col_idx >= len(row):
            return None
        return row[col_idx]

    c_name = _col("nodeGroup")
    c_cap = _col("capacity margin (MW)")
    c_nonsync = _col("non synchronous share")
    c_inertia = _col("inertia limit (MWs)")
    c_ts_reserve = _col("use ts_reserve")
    c_dyn_reserve = _col("use dynamic reserve")

    for row in rows[1:]:
        if _row_is_empty(row):
            continue
        name = _safe_str(_get(row, c_name))
        if name is None:
            continue
        result.append(NodeGroup(
            name=name,
            capacity_margin_mw=_safe_float(_get(row, c_cap)),
            non_synchronous_share=_safe_float(_get(row, c_nonsync)),
            inertia_limit_mws=_safe_float(_get(row, c_inertia)),
            use_ts_reserve=_safe_float(_get(row, c_ts_reserve)),
            use_dynamic_reserve=_safe_float(_get(row, c_dyn_reserve)),
        ))
    return result


def _read_unit_types(wb: openpyxl.Workbook) -> list[UnitType]:
    """Read the 'unit_type' sheet."""
    ws = _get_sheet(wb, "unit_type")
    if ws is None:
        return []

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [_safe_str(h) for h in rows[0]]
    result: list[UnitType] = []

    # First column is the unit type name; remaining columns are parameters
    for row in rows[1:]:
        if _row_is_empty(row):
            continue
        name = _safe_str(row[0]) if len(row) > 0 else None
        if name is None:
            continue
        params: dict[str, float | None] = {}
        for i in range(1, min(len(headers), len(row))):
            header = headers[i]
            if header is not None:
                key = header.strip().lower()
                params[key] = _safe_float(row[i])
        result.append(UnitType(name=name, params=params))
    return result


def _read_fuels(wb: openpyxl.Workbook) -> list[Fuel]:
    """Read the 'fuel' sheet."""
    ws = _get_sheet(wb, "fuel")
    if ws is None:
        return []

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    result: list[Fuel] = []
    for row in rows[1:]:
        if _row_is_empty(row):
            continue
        name = _safe_str(row[0]) if len(row) > 0 else None
        if name is None:
            continue
        price = _safe_float(row[1] if len(row) > 1 else None)
        co2 = _safe_float(row[2] if len(row) > 2 else None)
        result.append(Fuel(
            name=name,
            price_per_mwh=price if price is not None else 0.0,
            co2_content=co2 if co2 is not None else 0.0,
        ))
    return result


def _read_unit_groups(wb: openpyxl.Workbook) -> list[UnitGroup]:
    """Read the 'unitGroup' sheet."""
    ws = _get_sheet(wb, "unitGroup")
    if ws is None:
        return []

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [_safe_str(h) for h in rows[0]]
    result: list[UnitGroup] = []

    def _col(name: str) -> int | None:
        name_lower = name.lower()
        for i, h in enumerate(headers):
            if h is not None and h.lower() == name_lower:
                return i
        return None

    def _get(row: tuple[Any, ...], col_idx: int | None) -> Any:
        if col_idx is None or col_idx >= len(row):
            return None
        return row[col_idx]

    c_name = _col("unitGroup")
    c_max_mw = _col("max invest MW")
    c_min_mw = _col("min invest MW")
    c_max_mwh = _col("max invest MWh")
    c_min_mwh = _col("min invest MWh")

    for row in rows[1:]:
        if _row_is_empty(row):
            continue
        name = _safe_str(_get(row, c_name))
        if name is None:
            continue
        result.append(UnitGroup(
            name=name,
            max_invest_mw=_safe_float(_get(row, c_max_mw)),
            min_invest_mw=_safe_float(_get(row, c_min_mw)),
            max_invest_mwh=_safe_float(_get(row, c_max_mwh)),
            min_invest_mwh=_safe_float(_get(row, c_min_mwh)),
        ))
    return result


def _read_units(wb: openpyxl.Workbook) -> list[UnitInstance]:
    """Read the 'units' sheet."""
    ws = _get_sheet(wb, "units")
    if ws is None:
        return []

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Build header-to-index mapping
    raw_headers = rows[0]
    headers: list[str | None] = [_safe_str(h) for h in raw_headers]
    normalized_headers: list[str | None] = [
        h.strip().lower() if h is not None else None for h in headers
    ]

    result: list[UnitInstance] = []

    for row in rows[1:]:
        # Stop reading when all values in a row are None
        if _row_is_empty(row):
            continue

        # Build a dict of field_name -> value for this row
        field_values: dict[str, Any] = {
            "unit_group": None,
            "unit_type": "",
            "fuel": None,
            "cf_profile": None,
            "inflow_profile": None,
            "input_grid": None,
            "input_node": None,
            "output_grid": None,
            "output_node": None,
            "capacity_mw": None,
            "invested_capacity_mw": None,
            "max_invest_mw": None,
            "storage_mwh": None,
            "invested_storage_mwh": None,
            "max_invest_mwh": None,
            "storage_start": None,
            "storage_finish": None,
            "reserve_increase_ratio": None,
            "use_efficiency_ts": None,
            "output2_grid": None,
            "output2_node": None,
            "output2_eq_coeff": None,
            "output2_eq_constant": None,
            "output2_gt_coeff": None,
            "output2_gt_constant": None,
            "output2_lt_coeff": None,
            "output2_lt_constant": None,
            "output2_max_capacity": None,
            "fueluse_increase_eq": None,
            "inflow_multiplier": None,
        }
        extra: dict[str, float | None] = {}

        for col_idx, norm_h in enumerate(normalized_headers):
            if norm_h is None or col_idx >= len(row):
                continue
            cell_value = row[col_idx]

            # Check for combined grid/node columns (e.g. "input grid/node")
            if norm_h in _UNIT_GRID_NODE_COLS:
                grid_field, node_field = _UNIT_GRID_NODE_COLS[norm_h]
                g, n = _split_grid_node(cell_value)
                field_values[grid_field] = g
                field_values[node_field] = n
            elif norm_h in _UNIT_COLUMN_MAP:
                field_name = _UNIT_COLUMN_MAP[norm_h]
                if field_name in _UNIT_STRING_FIELDS:
                    field_values[field_name] = _safe_str(cell_value)
                else:
                    field_values[field_name] = _safe_float(cell_value)
            else:
                # Extra / unknown columns
                original_header = headers[col_idx]
                if original_header is not None:
                    extra[original_header] = _safe_float(cell_value)

        # Skip rows where unit_type is missing
        if not field_values["unit_type"]:
            continue

        result.append(UnitInstance(
            unit_group=field_values["unit_group"],
            unit_type=field_values["unit_type"],
            fuel=field_values["fuel"],
            cf_profile=field_values["cf_profile"],
            inflow_profile=field_values["inflow_profile"],
            input_grid=field_values["input_grid"],
            input_node=field_values["input_node"],
            output_grid=field_values["output_grid"],
            output_node=field_values["output_node"],
            capacity_mw=field_values["capacity_mw"],
            invested_capacity_mw=field_values["invested_capacity_mw"],
            max_invest_mw=field_values["max_invest_mw"],
            storage_mwh=field_values["storage_mwh"],
            invested_storage_mwh=field_values["invested_storage_mwh"],
            max_invest_mwh=field_values["max_invest_mwh"],
            storage_start=field_values["storage_start"],
            storage_finish=field_values["storage_finish"],
            reserve_increase_ratio=field_values["reserve_increase_ratio"],
            use_efficiency_ts=field_values["use_efficiency_ts"],
            output2_grid=field_values["output2_grid"],
            output2_node=field_values["output2_node"],
            output2_eq_coeff=field_values["output2_eq_coeff"],
            output2_eq_constant=field_values["output2_eq_constant"],
            output2_gt_coeff=field_values["output2_gt_coeff"],
            output2_gt_constant=field_values["output2_gt_constant"],
            output2_lt_coeff=field_values["output2_lt_coeff"],
            output2_lt_constant=field_values["output2_lt_constant"],
            output2_max_capacity=field_values["output2_max_capacity"],
            fueluse_increase_eq=field_values["fueluse_increase_eq"],
            inflow_multiplier=field_values["inflow_multiplier"],
            extra_params=extra,
        ))
    return result


def _read_connections(wb: openpyxl.Workbook) -> list[NodeNodeConnection]:
    """Read the 'nodeNode' sheet."""
    ws = _get_sheet(wb, "nodeNode")
    if ws is None:
        return []

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [_safe_str(h) for h in rows[0]]
    result: list[NodeNodeConnection] = []

    def _col(name: str) -> int | None:
        name_lower = name.lower()
        for i, h in enumerate(headers):
            if h is not None and h.lower() == name_lower:
                return i
        return None

    def _get(row: tuple[Any, ...], col_idx: int | None) -> Any:
        if col_idx is None or col_idx >= len(row):
            return None
        return row[col_idx]

    c_grid = _col("grid")
    c_n1 = _col("node1")
    c_n2 = _col("node2")
    c_cap_r = _col("cap.rightward (MW)")
    c_cap_l = _col("cap.leftward (MW)")
    c_inv_cap = _col("invested capacity (MW)")
    c_max_inv = _col("max invest (MW)")
    c_loss = _col("loss")
    c_inv_cost = _col("inv.cost/kW")
    c_life = _col("lifetime")
    c_interest = _col("interest")
    c_annuity = _col("annuity")
    c_hvdc = _col("HVDC")

    for row in rows[1:]:
        if _row_is_empty(row):
            continue
        grid = _safe_str(_get(row, c_grid))
        n1 = _safe_str(_get(row, c_n1))
        n2 = _safe_str(_get(row, c_n2))
        if grid is None or n1 is None or n2 is None:
            continue

        hvdc_val = _safe_float(_get(row, c_hvdc))
        is_hvdc = bool(hvdc_val) if hvdc_val is not None else False

        result.append(NodeNodeConnection(
            grid=grid,
            node1=n1,
            node2=n2,
            cap_rightward_mw=_safe_float(_get(row, c_cap_r)),
            cap_leftward_mw=_safe_float(_get(row, c_cap_l)),
            invested_capacity_mw=_safe_float(_get(row, c_inv_cap)),
            max_invest_mw=_safe_float(_get(row, c_max_inv)),
            loss=_safe_float(_get(row, c_loss)),
            invest_cost_per_kw=_safe_float(_get(row, c_inv_cost)),
            lifetime=_safe_float(_get(row, c_life)),
            interest=_safe_float(_get(row, c_interest)),
            is_hvdc=is_hvdc,
        ))
    return result


def _read_profile_ts(
    wb: openpyxl.Workbook,
    sheet_name: str,
) -> list[TimeSeriesData]:
    """Read a profile time-series sheet (ts_cf, ts_inflow).

    Layout:
      Row 0: col 1 has a label, cols 2+ have profile names.
      Row 1: col 0 has "time" / "Time".
      Rows 2+: col 0 has time ids, cols 2+ have values.
    """
    ws = _get_sheet(wb, sheet_name)
    if ws is None:
        return []

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        logger.warning("Sheet '%s' has fewer than 3 rows; skipping.", sheet_name)
        return []

    # Row 0: profile names starting at col index that has data
    header_row = rows[0]
    # Determine the first data column (skip cols 0 and possibly 1 if it's a label)
    profile_names: dict[int, str] = {}
    for col_idx in range(1, len(header_row)):
        name = _safe_str(header_row[col_idx])
        if name is not None:
            profile_names[col_idx] = name

    if not profile_names:
        logger.warning("Sheet '%s': no profile names found in header row.", sheet_name)
        return []

    # Determine the starting data column: smallest col_idx that has a profile name
    # But the actual profile names may be labels in col 1 -- find the first numeric col
    first_data_col = min(profile_names.keys())

    # Build time series, skipping header rows (rows 0 and 1)
    series_data: dict[int, dict[str, float]] = {ci: {} for ci in profile_names}

    for row in rows[2:]:
        if _row_is_empty(row):
            continue
        time_id = _safe_str(row[0]) if len(row) > 0 else None
        if time_id is None:
            continue
        for col_idx in profile_names:
            if col_idx < len(row):
                val = _safe_float(row[col_idx])
                if val is not None:
                    series_data[col_idx][time_id] = val

    result: list[TimeSeriesData] = []
    for col_idx, name in sorted(profile_names.items()):
        # Skip if name looks like a label for the header row itself
        # (e.g. "cf_profile", "inflow") -- these are column 1 labels, not profile names
        if col_idx == first_data_col and first_data_col == 1:
            # Check if there are profiles at higher indices; if so, col 1 is likely a label
            if len(profile_names) > 1:
                continue
        data = series_data[col_idx]
        if data:
            result.append(TimeSeriesData(name=name, data=data))
    return result


def _read_demand_ts(
    wb: openpyxl.Workbook,
    sheet_name: str,
) -> list[DemandTimeSeries]:
    """Read a demand-style time-series sheet (ts_energy, ts_import).

    Layout:
      Row 0: col 1 has "grid", cols 2+ have grid names.
      Row 1: col 1 has "node", cols 2+ have node names.
      Row 2: col 0 has "time".
      Rows 3+: col 0 has time ids, cols 2+ have values.
    """
    ws = _get_sheet(wb, sheet_name)
    if ws is None:
        return []

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 4:
        logger.warning("Sheet '%s' has fewer than 4 rows; skipping.", sheet_name)
        return []

    grid_row = rows[0]
    node_row = rows[1]

    # Build column -> (grid, node) mapping
    col_map: dict[int, tuple[str, str]] = {}
    for col_idx in range(2, max(len(grid_row), len(node_row))):
        grid = _safe_str(grid_row[col_idx]) if col_idx < len(grid_row) else None
        node = _safe_str(node_row[col_idx]) if col_idx < len(node_row) else None
        if grid is not None and node is not None:
            col_map[col_idx] = (grid, node)

    if not col_map:
        logger.warning("Sheet '%s': no grid/node pairs found.", sheet_name)
        return []

    # Read data rows (row 3+)
    series_data: dict[int, dict[str, float]] = {ci: {} for ci in col_map}

    for row in rows[3:]:
        if _row_is_empty(row):
            continue
        time_id = _safe_str(row[0]) if len(row) > 0 else None
        if time_id is None:
            continue
        for col_idx in col_map:
            if col_idx < len(row):
                val = _safe_float(row[col_idx])
                if val is not None:
                    series_data[col_idx][time_id] = val

    result: list[DemandTimeSeries] = []
    for col_idx, (grid, node) in sorted(col_map.items()):
        data = series_data[col_idx]
        if data:
            result.append(DemandTimeSeries(grid=grid, node=node, data=data))
    return result


def _read_reserve_ts(
    wb: openpyxl.Workbook,
    sheet_name: str,
) -> list[TimeSeriesData]:
    """Read a reserve time-series sheet (ts_reserve_node, ts_reserve_nodeGroup).

    Layout:
      Row 0: col 1 has "node"/"nodeGroup", cols 2+ have names.
      Row 1: col 0 has "Time"/"time".
      Rows 2+: col 0 has time ids, cols 2+ have values.
    """
    ws = _get_sheet(wb, sheet_name)
    if ws is None:
        return []

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        logger.warning("Sheet '%s' has fewer than 3 rows; skipping.", sheet_name)
        return []

    header_row = rows[0]
    # Profile names at col 2+
    name_map: dict[int, str] = {}
    for col_idx in range(2, len(header_row)):
        name = _safe_str(header_row[col_idx])
        if name is not None:
            name_map[col_idx] = name

    if not name_map:
        # Try starting from col 1 in case there's no label col
        for col_idx in range(1, len(header_row)):
            name = _safe_str(header_row[col_idx])
            if name is not None:
                name_map[col_idx] = name

    if not name_map:
        logger.warning("Sheet '%s': no names found in header row.", sheet_name)
        return []

    series_data: dict[int, dict[str, float]] = {ci: {} for ci in name_map}

    for row in rows[2:]:
        if _row_is_empty(row):
            continue
        time_id = _safe_str(row[0]) if len(row) > 0 else None
        if time_id is None:
            continue
        for col_idx in name_map:
            if col_idx < len(row):
                val = _safe_float(row[col_idx])
                if val is not None:
                    series_data[col_idx][time_id] = val

    result: list[TimeSeriesData] = []
    for col_idx, name in sorted(name_map.items()):
        data = series_data[col_idx]
        if data:
            result.append(TimeSeriesData(name=name, data=data))
    return result


def _read_unit_ts(wb: openpyxl.Workbook) -> list[UnitTimeSeries]:
    """Read the 'ts_unit' sheet.

    Layout:
      Row 0: col 1 has "grid", cols 2+ have grid names.
      Row 1: col 1 has "node", cols 2+ have node names.
      Row 2: col 1 has "unit", cols 2+ have unit names.
      Row 3: col 1 has "unit_ts_param", cols 2+ have parameter names.
      Row 4: col 0 has "Time".
      Rows 5+: col 0 has time ids, cols 2+ have values.
    """
    ws = _get_sheet(wb, "ts_unit")
    if ws is None:
        return []

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 6:
        logger.warning("Sheet 'ts_unit' has fewer than 6 rows; skipping.")
        return []

    grid_row = rows[0]
    node_row = rows[1]
    unit_row = rows[2]
    param_row = rows[3]

    # Build column -> (grid, node, unit, param) mapping
    col_map: dict[int, tuple[str, str, str, str]] = {}
    max_len = max(len(grid_row), len(node_row), len(unit_row), len(param_row))
    for col_idx in range(2, max_len):
        grid = _safe_str(grid_row[col_idx]) if col_idx < len(grid_row) else None
        node = _safe_str(node_row[col_idx]) if col_idx < len(node_row) else None
        unit = _safe_str(unit_row[col_idx]) if col_idx < len(unit_row) else None
        param = _safe_str(param_row[col_idx]) if col_idx < len(param_row) else None
        if grid is not None and node is not None and unit is not None and param is not None:
            col_map[col_idx] = (grid, node, unit, param)

    if not col_map:
        logger.warning("Sheet 'ts_unit': no unit time series columns found.")
        return []

    series_data: dict[int, dict[str, float]] = {ci: {} for ci in col_map}

    for row in rows[5:]:
        if _row_is_empty(row):
            continue
        time_id = _safe_str(row[0]) if len(row) > 0 else None
        if time_id is None:
            continue
        for col_idx in col_map:
            if col_idx < len(row):
                val = _safe_float(row[col_idx])
                if val is not None:
                    series_data[col_idx][time_id] = val

    result: list[UnitTimeSeries] = []
    for col_idx, (grid, node, unit, param) in sorted(col_map.items()):
        data = series_data[col_idx]
        if data:
            result.append(UnitTimeSeries(
                grid=grid, node=node, unit=unit, param_name=param, data=data,
            ))
    return result


def _read_time_steps(wb: openpyxl.Workbook) -> list[TimeStep]:
    """Read the 'ts_time' sheet."""
    ws = _get_sheet(wb, "ts_time")
    if ws is None:
        return []

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [_safe_str(h) for h in rows[0]]
    result: list[TimeStep] = []

    def _col(name: str) -> int | None:
        name_lower = name.lower()
        for i, h in enumerate(headers):
            if h is not None and h.lower() == name_lower:
                return i
        return None

    def _get(row: tuple[Any, ...], col_idx: int | None) -> Any:
        if col_idx is None or col_idx >= len(row):
            return None
        return row[col_idx]

    c_time = _col("time")
    c_in_use = _col("in_use")
    c_jump = _col("time_jump")
    c_in_use_inv = _col("in_use_invest")
    c_jump_inv = _col("time_jump_invest")

    for row in rows[1:]:
        if _row_is_empty(row):
            continue
        time_id = _safe_str(_get(row, c_time))
        if time_id is None:
            continue

        in_use_val = _safe_float(_get(row, c_in_use))
        in_use_inv_val = _safe_float(_get(row, c_in_use_inv))
        jump_val = _safe_float(_get(row, c_jump))
        jump_inv_val = _safe_float(_get(row, c_jump_inv))

        result.append(TimeStep(
            time_id=time_id,
            in_use=bool(in_use_val) if in_use_val is not None else False,
            time_jump=jump_val if jump_val is not None else 0.0,
            in_use_invest=bool(in_use_inv_val) if in_use_inv_val is not None else False,
            time_jump_invest=jump_inv_val if jump_inv_val is not None else 0.0,
        ))
    return result


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def read_old_flextool(file_path: str) -> OldFlexToolData:
    """Read an old-format FlexTool Excel file and return structured data.

    Args:
        file_path: Path to the .xlsm file.

    Returns:
        An OldFlexToolData instance with all parsed sheets.
    """
    logger.info("Reading old FlexTool file: %s", file_path)
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    try:
        # Log which sheets are present and which will be skipped
        for name in wb.sheetnames:
            if name in _SKIP_SHEETS:
                logger.debug("Skipping sheet: %s", name)

        master = _read_master(wb)
        grid_nodes = _read_grid_nodes(wb)
        node_groups = _read_node_groups(wb)
        unit_types = _read_unit_types(wb)
        fuels = _read_fuels(wb)
        unit_groups = _read_unit_groups(wb)
        units = _read_units(wb)
        connections = _read_connections(wb)
        cf_profiles = _read_profile_ts(wb, "ts_cf")
        inflow_profiles = _read_profile_ts(wb, "ts_inflow")
        demand_ts = _read_demand_ts(wb, "ts_energy")
        import_ts = _read_demand_ts(wb, "ts_import")
        reserve_node_ts = _read_reserve_ts(wb, "ts_reserve_node")
        reserve_group_ts = _read_reserve_ts(wb, "ts_reserve_nodeGroup")
        unit_ts = _read_unit_ts(wb)
        time_steps = _read_time_steps(wb)

        logger.info(
            "Parsed: %d nodes, %d unit types, %d units, %d connections, "
            "%d cf profiles, %d time steps",
            len(grid_nodes), len(unit_types), len(units), len(connections),
            len(cf_profiles), len(time_steps),
        )

        return OldFlexToolData(
            master=master,
            grid_nodes=grid_nodes,
            node_groups=node_groups,
            unit_types=unit_types,
            fuels=fuels,
            unit_groups=unit_groups,
            units=units,
            connections=connections,
            cf_profiles=cf_profiles,
            inflow_profiles=inflow_profiles,
            demand_ts=demand_ts,
            import_ts=import_ts,
            reserve_node_ts=reserve_node_ts,
            reserve_group_ts=reserve_group_ts,
            unit_ts=unit_ts,
            time_steps=time_steps,
        )
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Entity identifier columns by section type
# ---------------------------------------------------------------------------

_SENSITIVITY_ENTITY_ID_COLS: dict[str, set[str]] = {
    "master": set(),
    "nodeGroup": {"nodegroup"},
    "gridNode": {"grid", "node", "nodegroup", "nodegroup2", "nodegroup3"},
    "unit_type": {"unit type"},
    "fuel": {"fuel"},
    "unitGroup": {"unitgroup"},
    "units": {
        "unitgroup", "unittype", "fuel", "cf profile", "inflow",
        "input grid", "input node", "output grid", "output node",
    },
    "nodeNode": {"grid", "node1", "node2"},
}

# String-typed entity ID columns (stay as strings, never float-convert)
_SENSITIVITY_STRING_COLS: set[str] = {
    "nodegroup", "nodegroup2", "nodegroup3",
    "grid", "node", "node1", "node2",
    "unit type", "unittype", "unitgroup",
    "fuel", "cf profile", "inflow",
    "input grid", "input node", "output grid", "output node",
}


def read_old_flextool_sensitivities(
    file_path: str,
) -> dict[str, list[SensitivityOverride]]:
    """Read the 'Sensitivity definitions' sheet and return per-scenario overrides.

    The sheet contains multiple sections separated by empty rows. Each section
    starts with a header row like "Scenario definitions master:" followed by
    column headers and data rows.

    Args:
        file_path: Path to the old FlexTool .xlsm file.

    Returns:
        A dict mapping scenario_name -> list of SensitivityOverride objects.
    """
    logger.info("Reading sensitivities from: %s", file_path)
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    try:
        ws = _get_sheet(wb, "Sensitivity definitions")
        if ws is None:
            logger.warning("No 'Sensitivity definitions' sheet found.")
            return {}

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {}

        result: dict[str, list[SensitivityOverride]] = {}
        _parse_sensitivity_sections(rows, result)

        total = sum(len(v) for v in result.values())
        logger.info(
            "Parsed %d sensitivity scenarios with %d total overrides.",
            len(result), total,
        )
        return result

    finally:
        wb.close()


def _parse_sensitivity_sections(
    rows: list[tuple[Any, ...]],
    result: dict[str, list[SensitivityOverride]],
) -> None:
    """Parse all sections from the Sensitivity definitions sheet rows."""
    i = 0
    while i < len(rows):
        row = rows[i]
        # Detect section header: col 0 starts with "Scenario definitions"
        cell0 = _safe_str(row[0]) if len(row) > 0 else None
        if cell0 is not None and cell0.lower().startswith("scenario definitions"):
            # This row is the header row with column names in col 2+
            headers = [_safe_str(c) for c in row]
            # Extract the section type from the header text after ":"
            # e.g. "Scenario definitions master:" -> determine from data rows
            i += 1
            # Parse data rows until next empty row or next section header
            i = _parse_sensitivity_data_rows(rows, i, headers, result)
        else:
            i += 1


def _parse_sensitivity_data_rows(
    rows: list[tuple[Any, ...]],
    start: int,
    headers: list[str | None],
    result: dict[str, list[SensitivityOverride]],
) -> int:
    """Parse data rows for one section. Returns the index after the section."""
    i = start
    while i < len(rows):
        row = rows[i]
        if _row_is_empty(row):
            # Empty row signals end of section
            return i + 1

        # Check if this is the start of a new section header
        cell0 = _safe_str(row[0]) if len(row) > 0 else None
        if cell0 is not None and cell0.lower().startswith("scenario definitions"):
            return i  # Don't advance — let the caller re-process this row

        # Col 0: scenario name, Col 1: sheet/section type
        scenario_name = cell0
        section_type = _safe_str(row[1]) if len(row) > 1 else None

        # Skip rows where col 0 is empty (placeholder rows with just sheet name)
        if scenario_name is None or section_type is None:
            i += 1
            continue

        # Determine entity ID vs parameter columns based on section type
        entity_id_col_names = _SENSITIVITY_ENTITY_ID_COLS.get(section_type, set())

        # Build entity_ids from the ID columns
        entity_ids: dict[str, str] = {}
        for col_idx in range(2, min(len(headers), len(row))):
            header = headers[col_idx]
            if header is None:
                continue
            header_lower = header.strip().lower()
            if header_lower in entity_id_col_names:
                val = _safe_str(row[col_idx])
                if val is not None:
                    entity_ids[header.strip()] = val

        # Create overrides for each non-empty parameter cell
        for col_idx in range(2, min(len(headers), len(row))):
            header = headers[col_idx]
            if header is None:
                continue
            header_lower = header.strip().lower()
            if header_lower in entity_id_col_names:
                continue  # Skip entity ID columns

            raw_value = row[col_idx]
            if raw_value is None:
                continue

            # Try float first, fall back to string
            float_val = _safe_float(raw_value)
            if float_val is not None:
                value: float | str = float_val
            else:
                str_val = _safe_str(raw_value)
                if str_val is None:
                    continue
                value = str_val

            override = SensitivityOverride(
                section=section_type,
                entity_ids=dict(entity_ids),  # copy
                param_name=header.strip(),
                value=value,
            )
            result.setdefault(scenario_name, []).append(override)

        i += 1

    return i
