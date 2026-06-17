"""Read self-describing FlexTool Excel files.

Each sheet embeds its own metadata via a 2D grid of definitions:
- A **definition row** labels columns (alternative, entity, parameters, etc.)
- A **definition column** labels metadata rows (description, data type, etc.)
- Their intersection is the **crossing point**.

The reader auto-discovers the crossing point and interprets the sheet
layout from the embedded keywords, eliminating the need for an external
JSON import specification.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Keywords recognised in the definition column (row labels)
# ---------------------------------------------------------------------------

ROW_KEYWORDS = frozenset({
    "description",
    "data type",
    "parameter",
    "alternative",
    "entity",
    "index",
    "filter",
})

# Keywords that can appear as column definitions in the definition row.
# Many start with a prefix (e.g. "entity: node"), so we check prefixes.
COL_DEF_PREFIXES = (
    "alternative",
    "entity",
    "filter",
    "index",
    "parameter",
)

# Special parameter names recognised by the importer.
#
# The exporter (``excel_writer.py``) emits exactly ``"entity existence"`` as
# the header label and ``"TRUE"`` / ``"FALSE"`` as the cell values.  The v2
# reader is a faithful inverse of that exporter — it does NOT accept legacy
# aliases (the old ``"entity alternative"`` header is rejected so users see
# a clear error rather than silent renames).
ENTITY_EXISTENCE = "entity existence"
# Strings the exporter EVER emits for a boolean-valued cell (case-insensitive
# match).  Anything else is treated as a data error and raised — silent
# coercion of ``"yes"`` / ``"no"`` / ``"1"`` would mask data corruption.
_BOOL_TRUE_LITERAL = "true"
_BOOL_FALSE_LITERAL = "false"


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class EntityClassDef:
    """One entity class mapping parsed from the entity definition."""
    class_name: str
    dimensions: list[str]


@dataclass
class SheetMetadata:
    """Parsed metadata from a single sheet."""

    sheet_name: str

    # Crossing point coordinates (0-based)
    def_row: int = 0       # the definition row index
    def_col: int = 0       # the definition column index

    # Column definitions (from the definition row, left of crossing)
    # Each maps column index -> definition string
    col_defs: dict[int, str] = field(default_factory=dict)

    # Row definitions (from the definition column, above crossing)
    # Each maps row index -> definition string
    row_defs: dict[int, str] = field(default_factory=dict)

    # Parameter names (from the definition row, right of crossing)
    # Maps column index -> parameter name
    param_cols: dict[int, str] = field(default_factory=dict)

    # Default parameter name (from "parameter: xxx" in the label cell)
    default_parameter: str | None = None

    # Default data type (from "data type: xxx" in the label cell)
    default_data_type: str | None = None

    # Data types per parameter column index
    data_types: dict[int, str] = field(default_factory=dict)

    # Descriptions per parameter column index
    descriptions: dict[int, str] = field(default_factory=dict)

    # Parsed entity class definitions
    entity_classes: list[EntityClassDef] = field(default_factory=list)

    # Filter column index and regex map {class_name: pattern}
    filter_col: int | None = None
    filter_map: dict[str, str] = field(default_factory=dict)

    # Index column/row info
    index_col: int | None = None
    index_name: str | None = None

    # Extra index columns (for ladder-style sheets with multiple ``index:`` cols
    # on the left, e.g. ``index: period | index: tier``).  Each entry is
    # ``(column_index, axis_name)``.  When non-empty, ``index_col`` /
    # ``index_name`` carries the LAST entry; ``extra_index_cols`` carries the
    # preceding ones in left-to-right order so the data extractor can recover
    # all index levels.
    extra_index_cols: list[tuple[int, str]] = field(default_factory=list)

    # Alternative column index (or None if alternatives are in a row)
    alt_col: int | None = None

    # Entity existence column index
    entity_existence_col: int | None = None

    # Data start row (0-based, first row after all metadata rows)
    data_start_row: int = 0

    # Whether this is a transposed (timeseries-style) sheet
    is_transposed: bool = False

    # For transposed: row indices for alternative, parameter, entity
    alt_row: int | None = None
    param_row: int | None = None
    entity_row: int | None = None
    index_col_transposed: int | None = None  # column with index values (usually 0)

    # Whether this is a stochastic (_s) sheet — a transposed sheet whose
    # entity / alternative / parameter live in COLUMN headers (like a
    # timeseries sheet) but whose index is a MULTI-LEVEL nested Map spread
    # across N left-side ``index:`` columns (one column per Map level).
    # The definition column (carrying the ``parameter:`` triplet, the
    # ``entity:`` label and the ``alternative`` label) sits at ``def_col``
    # (= number of index columns).  Each data column to its right is one
    # (entity, alternative) series whose cells carry the leaf scalars.
    is_stochastic: bool = False

    # Ordered ``(column_index, axis_name)`` for the stochastic index
    # columns, left-to-right (outermost Map level first).
    stoch_index_cols: list[tuple[int, str]] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Crossing point detection
# ---------------------------------------------------------------------------


def _cell_value(ws: Worksheet, row: int, col: int) -> str:
    """Get cell value as stripped string (0-based row/col)."""
    val = ws.cell(row=row + 1, column=col + 1).value  # openpyxl is 1-based
    if val is None:
        return ""
    return str(val).strip()


def _actual_max_row(ws: Worksheet) -> int:
    """Get the actual last row with data (not the Excel theoretical max)."""
    # ws.max_row in non-read-only mode should be correct, but guard against
    # sheets where it reports 1048576
    max_r = ws.max_row or 0
    if max_r > 100000:
        # Scan backwards from a reasonable limit
        for r in range(min(max_r, 10000), 0, -1):
            for c in range(1, min((ws.max_column or 1) + 1, 50)):
                if ws.cell(row=r, column=c).value is not None:
                    return r
        return 0
    return max_r


def _is_info_row(text: str) -> bool:
    """Check if text is an INFO row (to be skipped by the reader)."""
    return text.upper().startswith("INFO:")


def _is_keyword(text: str) -> bool:
    """Check if text starts with a recognised keyword."""
    t = text.lower()
    for prefix in COL_DEF_PREFIXES:
        if t == prefix or t.startswith(prefix + ":") or t.startswith(prefix + " "):
            return True
    if t in ROW_KEYWORDS:
        return True
    return False


def find_crossing_point(ws: Worksheet) -> tuple[int, int]:
    """Find the crossing point of the definition row and definition column.

    Algorithm:
    1. Scan row 0 (Excel row 1) rightward until a recognised keyword is found.
       That column is the definition column.
    2. Scan that column downward until an empty cell.  Go back one row.
       That row is the definition row.

    Returns:
        (def_row, def_col) — 0-based indices.

    Raises:
        ValueError: if no crossing point can be found.
    """
    max_col = ws.max_column or 1

    # Step 1: scan row 0 rightward for a keyword
    def_col = None
    for c in range(max_col):
        val = _cell_value(ws, 0, c)
        if val and _is_keyword(val):
            def_col = c
            break

    if def_col is None:
        raise ValueError(
            f"Sheet '{ws.title}': could not find definition column "
            f"(no keyword found in row 1)"
        )

    # Step 2: scan definition column downward until empty cell
    max_row = _actual_max_row(ws)
    def_row = 0
    for r in range(max_row):
        val = _cell_value(ws, r, def_col)
        if not val:
            # Empty cell found — go back one
            def_row = max(0, r - 1)
            break
    else:
        # No empty cell found — last row with content
        def_row = max_row - 1

    return (def_row, def_col)


# ---------------------------------------------------------------------------
# Definition parsing
# ---------------------------------------------------------------------------


def _parse_entity_def(text: str) -> list[EntityClassDef]:
    """Parse an entity definition string.

    Examples:
        "entity: node"
            → [EntityClassDef("node", ["node"])]
        "entity: commodity, node"
            → [EntityClassDef("commodity__node", ["commodity", "node"])]
        "entity: (unit__inputNode: (unit, node), unit__outputNode: (unit, node))"
            → [EntityClassDef("unit__inputNode", ["unit", "node"]),
               EntityClassDef("unit__outputNode", ["unit", "node"])]
    """
    # Strip "entity:" prefix
    text = text.strip()
    if text.lower().startswith("entity:"):
        text = text[len("entity:"):].strip()
    elif text.lower().startswith("entity name:"):
        text = text[len("entity name:"):].strip()
    elif text.lower() == "entity":
        return []

    # Check for multi-class syntax: (class1: (dim1, dim2), class2: (dim3, dim4))
    if text.startswith("(") and text.endswith(")"):
        inner = text[1:-1].strip()
        results = []
        # Split on top-level commas (not inside parens)
        parts = _split_top_level(inner)
        for part in parts:
            part = part.strip()
            # Parse "class_name: (dim1, dim2)"
            match = re.match(r'(\w+)\s*:\s*\(([^)]+)\)', part)
            if match:
                class_name = match.group(1).strip()
                dims = [d.strip() for d in match.group(2).split(",")]
                results.append(EntityClassDef(class_name, dims))
        return results

    # Simple case: "node" or "commodity, node"
    dims = [d.strip() for d in text.split(",")]
    if len(dims) == 1:
        return [EntityClassDef(dims[0], dims)]
    else:
        class_name = "__".join(dims)
        return [EntityClassDef(class_name, dims)]


def _split_top_level(text: str) -> list[str]:
    """Split text on commas that are not inside parentheses."""
    parts = []
    depth = 0
    current = []
    for ch in text:
        if ch == "(":
            depth += 1
            current.append(ch)
        elif ch == ")":
            depth -= 1
            current.append(ch)
        elif ch == "," and depth == 0:
            parts.append("".join(current))
            current = []
        else:
            current.append(ch)
    if current:
        parts.append("".join(current))
    return parts


def _parse_filter_def(text: str) -> dict[str, str]:
    """Parse a filter definition string.

    Example:
        "filter: {unit__inputNode: ^input$, unit__outputNode: ^output$}"
        → {"unit__inputNode": "^input$", "unit__outputNode": "^output$"}
    """
    text = text.strip()
    if text.lower().startswith("filter:"):
        text = text[len("filter:"):].strip()

    # Remove outer braces or parens
    if (text.startswith("{") and text.endswith("}")) or \
       (text.startswith("(") and text.endswith(")")):
        text = text[1:-1].strip()

    result = {}
    for part in _split_top_level(text):
        part = part.strip()
        if ":" in part:
            key, val = part.split(":", 1)
            result[key.strip()] = val.strip()
    return result


def _parse_index_def(text: str) -> str:
    """Parse an index definition: "index: time" → "time"."""
    text = text.strip()
    if text.lower().startswith("index:"):
        return text[len("index:"):].strip()
    return text


def _parse_default_value(text: str) -> tuple[str, str | None]:
    """Parse a definition that may have a default value.

    "parameter: profile" → ("parameter", "profile")
    "parameter"          → ("parameter", None)
    "data type: float"   → ("data type", "float")

    Also handles the triplet format with ``|`` separator:
    "parameter: profile | data type: float | description: ..."
    → ("parameter", "profile")
    """
    # Handle | separator: only parse the first segment
    if "|" in text:
        text = text.split("|")[0].strip()

    for keyword in sorted(ROW_KEYWORDS, key=len, reverse=True):
        if text.lower().startswith(keyword):
            rest = text[len(keyword):].strip()
            if rest.startswith(":"):
                return keyword, rest[1:].strip()
            elif rest == "":
                return keyword, None
    return text, None


def _extract_triplet_field(text: str, field: str) -> str | None:
    """Return the value of a labelled segment in a pipe-separated triplet.

    Looks for ``<field>: <value>`` inside a ``"a | b | c"`` style string
    (case-insensitive on the field name).  Returns the trimmed value or
    ``None`` when the field isn't present.
    """
    if not text or "|" not in text:
        return None
    fl = field.lower()
    for seg in text.split("|"):
        s = seg.strip()
        if s.lower().startswith(fl + ":"):
            return s[len(field) + 1:].strip()
    return None


# ---------------------------------------------------------------------------
# Sheet metadata extraction
# ---------------------------------------------------------------------------


def parse_sheet_metadata(ws: Worksheet) -> SheetMetadata | None:
    """Parse a sheet's embedded metadata.

    Returns None if the sheet has no recognisable metadata (e.g. navigate, version).
    """
    try:
        def_row, def_col = find_crossing_point(ws)
    except ValueError:
        return None

    meta = SheetMetadata(sheet_name=ws.title, def_row=def_row, def_col=def_col)

    # --- Parse row definitions (definition column, above crossing point) ---
    for r in range(def_row):
        val = _cell_value(ws, r, def_col)
        if val:
            meta.row_defs[r] = val

    # The crossing point cell itself is also a row def
    crossing_val = _cell_value(ws, def_row, def_col)
    if crossing_val:
        meta.row_defs[def_row] = crossing_val

    # --- Parse column definitions (definition row, left of crossing) ---
    for c in range(def_col):
        val = _cell_value(ws, def_row, c)
        if val:
            meta.col_defs[c] = val.strip()

    # --- Interpret column definitions ---
    # Multiple ``index:`` columns on the left are recorded in
    # ``extra_index_cols`` (in column order); the rightmost wins for
    # ``index_col`` / ``index_name`` to keep backwards compatibility with
    # single-index sheets.
    index_seen: list[tuple[int, str]] = []
    for c, defn in meta.col_defs.items():
        dl = defn.lower()
        if dl == "alternative":
            meta.alt_col = c
        elif dl.startswith("entity"):
            meta.entity_classes = _parse_entity_def(defn)
        elif dl.startswith("filter"):
            meta.filter_col = c
            meta.filter_map = _parse_filter_def(defn)
        elif dl.startswith("index"):
            index_seen.append((c, _parse_index_def(defn)))
    if index_seen:
        index_seen.sort()
        meta.index_col, meta.index_name = index_seen[-1]
        # Preceding (left-of-rightmost) index columns become extras.
        meta.extra_index_cols = list(index_seen[:-1])

    # --- Parse parameter columns (definition row, right of crossing) ---
    # Stop at the first empty cell (separator column before parameter reference)
    max_col = ws.max_column or 1
    for c in range(def_col + 1, max_col):
        val = _cell_value(ws, def_row, c)
        if not val:
            break  # empty separator column = end of parameter area
        vl = val.lower()
        if vl == ENTITY_EXISTENCE:
            meta.entity_existence_col = c
            meta.param_cols[c] = ENTITY_EXISTENCE
        else:
            meta.param_cols[c] = val

    # --- Parse crossing point for defaults ---
    # NOTE: kept defensively — the current exporter ALWAYS writes a
    # parameter name into every right-of-crossing header cell, so the
    # back-fill loop below is dead code on writer-generated files.  It
    # remains in place to keep hand-edited sheets (where the user relies
    # on the crossing's ``parameter: <default>`` to label an empty
    # column) loadable.  If the exporter contract changes, revisit.
    keyword, default = _parse_default_value(crossing_val)
    if keyword.lower() == "parameter" and default:
        meta.default_parameter = default
        # Apply default to all param columns that are empty
        for c in range(def_col + 1, max_col):
            if c not in meta.param_cols:
                val = _cell_value(ws, def_row, c)
                if not val:
                    meta.param_cols[c] = default

    # --- Parse row metadata (description, data type) ---
    for r, label in meta.row_defs.items():
        if r == def_row:
            continue  # skip the definition row itself
        kw, default_val = _parse_default_value(label)
        kw_lower = kw.lower()

        if kw_lower == "description":
            for c in meta.param_cols:
                desc = _cell_value(ws, r, c)
                if desc:
                    meta.descriptions[c] = desc

        elif kw_lower == "data type":
            if default_val:
                meta.default_data_type = default_val
            for c in meta.param_cols:
                dt = _cell_value(ws, r, c)
                # NOTE: ``.lower()`` is defensive — the exporter already
                # emits all data-type labels lowercase ("float", "string",
                # "boolean (array)", "string (array)", "*-1d-map", ...).
                # Kept in case a hand-edited sheet uses mixed case so the
                # downstream dtype dispatch (``_convert_value``) keeps
                # matching its lowercase ``elif`` branches.
                if dt:
                    meta.data_types[c] = dt.lower()
                elif default_val:
                    meta.data_types[c] = default_val.lower()

    # Data starts after the definition row
    meta.data_start_row = def_row + 1

    return meta


def parse_link_sheet_metadata(ws: Worksheet) -> SheetMetadata | None:
    """Parse a link-only sheet (no parameters, just entity dimensions).

    Link sheets have a simpler format:
        Row 1: navigate link
        Row 2: entity definition in col A, dimension names in cols B+
        Row 3+: data
    """
    meta = SheetMetadata(sheet_name=ws.title)

    # Look for entity definition in column A, rows 1-3
    for r in range(min(3, ws.max_row or 1)):
        val = _cell_value(ws, r, 0)
        if val.lower().startswith("entity"):
            meta.entity_classes = _parse_entity_def(val)
            meta.def_row = r
            meta.data_start_row = r + 1
            return meta

    return None


def parse_transposed_sheet_metadata(ws: Worksheet) -> SheetMetadata | None:
    """Parse a transposed (timeseries-style) sheet.

    In transposed sheets, column B serves as the row definition column:
    - Row with "alternative" in col B → alternative values in cols C+
    - Row with "parameter" in col B → parameter names in cols C+
    - Row with "entity: X" in col B → entity names in cols C+
    - Column A has "index: time" and time values below
    """
    meta = SheetMetadata(sheet_name=ws.title, is_transposed=True)
    max_row = _actual_max_row(ws)

    # Scan column A for "index:" definition (skip INFO rows)
    for r in range(min(10, max_row)):
        val = _cell_value(ws, r, 0)
        if _is_info_row(val):
            continue
        if val.lower().startswith("index:"):
            meta.index_name = _parse_index_def(val)
            meta.index_col_transposed = 0
            meta.data_start_row = r + 1

    # Scan column B for row role definitions (skip INFO rows)
    label_col = 1  # column B
    datatype_row: int | None = None
    for r in range(min(10, max_row)):
        val_a = _cell_value(ws, r, 0)
        if _is_info_row(val_a):
            continue
        val = _cell_value(ws, r, label_col)
        if not val:
            continue
        vl = val.lower()
        kw, default_val = _parse_default_value(val)
        kw_lower = kw.lower()

        if vl == "alternative" or vl.startswith("alternative"):
            meta.alt_row = r
        elif kw_lower == "data type":
            # Multi-param _t sheets carry a per-column data-type row so the
            # leaf scalar type survives the round-trip (single-param sheets
            # fold it into the triplet instead).
            datatype_row = r
            if default_val:
                meta.default_data_type = default_val.lower()
        elif kw_lower == "parameter":
            meta.param_row = r
            if default_val:
                meta.default_parameter = default_val
        elif vl.startswith("entity"):
            meta.entity_row = r
            meta.entity_classes = _parse_entity_def(val)

    # Read per-column data types (one per transposed data column, C+).
    if datatype_row is not None:
        max_col = ws.max_column or 1
        for c in range(2, max_col):
            dt = _cell_value(ws, datatype_row, c)
            if dt:
                meta.data_types[c] = dt.lower()

    # Determine data start row (row after last metadata row)
    metadata_rows = [
        r for r in [meta.alt_row, meta.param_row, meta.entity_row, datatype_row]
        if r is not None
    ]
    if metadata_rows:
        meta.data_start_row = max(metadata_rows) + 1

    return meta


# ---------------------------------------------------------------------------
# Scenario sheet parsing
# ---------------------------------------------------------------------------


def _is_scenario_sheet(ws: Worksheet) -> bool:
    """Check if the worksheet is a scenario sheet.

    Scenario sheets have "Scenario names" in row 1, col 2 (0-based: row 0, col 1).
    """
    val = _cell_value(ws, 0, 1)
    return val.lower() == "scenario names" if val else False


def parse_scenario_sheet(ws: Worksheet) -> SheetData:
    """Parse a scenario sheet into SheetData records.

    Scenario sheet layout (Excel rows, 1-based):
        Row 1: navigate | Scenario names | ...
        Row 2: (empty)  | scenario_1     | scenario_2 | ...
        Row 3: base_alternative | alt_for_s1 | alt_for_s2 | ...
        Row 4: alternative_1   | alt_for_s1 | alt_for_s2 | ...
        Row 5: alternative_2   | (empty)    | alt_for_s2 | ...

    Produces records with:
        entity_class = "scenario"
        entity_byname = (scenario_name,)
        param_name = row label (e.g. "base_alternative", "alternative_1")
        alternative = the alternative name in that cell
        value = the rank (row index - 2, i.e. base_alternative=0, alternative_1=1, ...)
    """
    meta = SheetMetadata(sheet_name=ws.title)
    meta.entity_classes = [EntityClassDef("scenario", ["scenario"])]
    data = SheetData(sheet_name=ws.title, metadata=meta)

    max_row = _actual_max_row(ws)
    max_col = ws.max_column or 1

    # Row 2 (0-based row 1): scenario names in cols B+ (0-based col 1+)
    scenario_names: dict[int, str] = {}
    for c in range(1, max_col):
        name = _cell_value(ws, 1, c)
        if name:
            scenario_names[c] = name

    if not scenario_names:
        return data

    # Row 3+ (0-based row 2+): row label in col A, alternative names in cols B+
    for r in range(2, max_row):
        row_label = _cell_value(ws, r, 0)
        if not row_label:
            continue

        rank = r - 2  # base_alternative (row 2, 0-based) = rank 0

        for c, scenario_name in scenario_names.items():
            alt_name = _cell_value(ws, r, c)
            if not alt_name:
                continue

            data.records.append({
                "entity_class": "scenario",
                "entity_byname": (scenario_name,),
                "param_name": row_label,
                "alternative": alt_name,
                "value": rank,
                "index_value": None,
                "index_name": None,
            })

    return data


# ---------------------------------------------------------------------------
# Stochastic (_s) sheet parsing
# ---------------------------------------------------------------------------


def _is_index_label(text: str) -> bool:
    """Return True when *text* is an ``index:`` column definition."""
    return text.lower().startswith("index:") or text.lower() == "index"


def _find_stochastic_def_col(ws: Worksheet) -> int | None:
    """Locate the definition column of a stochastic (_s) sheet, or None.

    A stochastic sheet is a transposed sheet whose nested-Map index is
    spread across ``N`` left-side ``index:`` columns (one per Map level)
    followed by a definition column that carries the ``alternative`` /
    ``parameter`` / ``entity`` row roles (the entity/alt/param VALUES live
    in the column headers to its right, like a timeseries sheet).

    The signal: there are at least TWO ``index:`` columns on the left, and
    immediately after the last of them comes a column whose header rows
    carry the ``alternative`` (and ``parameter`` / ``entity``) role
    labels.  A single-index transposed sheet (timeseries) keeps its lone
    ``index:`` in column A and its def labels in column B — that case is
    handled by the existing transposed path and is explicitly NOT matched
    here (it has fewer than two ``index:`` columns).
    """
    max_row = _actual_max_row(ws)
    max_col = ws.max_column or 1
    scan_rows = min(8, max_row)

    # The role-definition row is the one carrying the ``alternative`` label
    # (in the definition column).  On a stochastic sheet that same row also
    # carries the ``index:`` labels for every Map level in the columns to
    # its left.  Anchor on that row, then verify the left columns are all
    # ``index:``.
    for role_row in range(scan_rows):
        def_col: int | None = None
        for c in range(max_col):
            if _cell_value(ws, role_row, c).lower() == "alternative":
                def_col = c
                break
        if def_col is None or def_col < 2:
            # Need at least two ``index:`` columns to the left to be a
            # genuine nested-Map (multi-index) sheet; a lone ``index:`` in
            # column A is an ordinary single-index transposed sheet.
            continue
        if all(
            _is_index_label(_cell_value(ws, role_row, c)) for c in range(def_col)
        ):
            return def_col
    return None


def parse_stochastic_sheet_metadata(ws: Worksheet, def_col: int) -> SheetMetadata:
    """Parse a stochastic (_s) sheet's embedded metadata.

    ``def_col`` is the 0-based definition column index (= the number of
    left-side ``index:`` columns).  The header rows carry the role labels
    (``parameter:``, ``entity:``, ``alternative``) in that column; the
    entity / alternative / leaf-data values live in the columns to its
    right (transposed).  The ``index:`` columns 0..def_col-1 give the
    nested-Map level axis names, outermost first.
    """
    meta = SheetMetadata(
        sheet_name=ws.title, is_transposed=True, is_stochastic=True,
    )
    meta.def_col = def_col
    max_row = _actual_max_row(ws)
    scan_rows = min(8, max_row)

    # Index columns (left of def_col), in order.  Axis name comes from the
    # role-definition row; collect from whichever header row carries it.
    for c in range(def_col):
        axis = ""
        for r in range(scan_rows):
            val = _cell_value(ws, r, c)
            if val and _is_index_label(val):
                axis = _parse_index_def(val)
                break
        meta.stoch_index_cols.append((c, axis))
    if meta.stoch_index_cols:
        # Keep single-index compatibility fields pointing at the LAST level.
        meta.index_col, meta.index_name = meta.stoch_index_cols[-1]

    # Role rows in the definition column.
    triplet_dtype: str | None = None
    for r in range(scan_rows):
        val = _cell_value(ws, r, def_col)
        if not val:
            continue
        vl = val.lower()
        kw, default_val = _parse_default_value(val)
        kw_lower = kw.lower()
        if vl == "alternative" or vl.startswith("alternative"):
            meta.alt_row = r
        elif kw_lower == "parameter":
            meta.param_row = r
            if default_val:
                meta.default_parameter = default_val
            # The single-param triplet carries the canonical data type.
            triplet_dtype = _extract_triplet_field(val, "data type")
        elif vl.startswith("entity"):
            meta.entity_row = r
            meta.entity_classes = _parse_entity_def(val)

    if triplet_dtype:
        meta.default_data_type = triplet_dtype.lower()

    # Data rows start after the last header (role) row.
    role_rows = [
        r for r in (meta.alt_row, meta.param_row, meta.entity_row)
        if r is not None
    ]
    meta.data_start_row = (max(role_rows) + 1) if role_rows else def_col + 1
    return meta


def _extract_stochastic(ws: Worksheet, meta: SheetMetadata) -> SheetData:
    """Extract per-leaf records from a stochastic (_s) sheet.

    Each data column (def_col+1 onward) is one (entity, alternative)
    series.  Each data row carries one full nested-Map index tuple in the
    left ``index:`` columns.  We emit one record per (column, row) cell,
    carrying the first N-1 index levels in ``extra_index_values`` and the
    last level in ``index_value`` / ``index_name`` — mirroring the ladder
    reader so the DB-side nested-Map reconstruction can rebuild the
    full ``Map(idx0 -> Map(idx1 -> ... -> leaf))``.
    """
    data = SheetData(sheet_name=meta.sheet_name, metadata=meta)
    max_row = _actual_max_row(ws)
    max_col = ws.max_column or 1
    def_col = meta.def_col

    entity_class = meta.entity_classes[0] if meta.entity_classes else None
    if entity_class is None:
        return data

    n_dims = len(entity_class.dimensions)
    leaf_dtype = meta.default_data_type or "string"

    for c in range(def_col + 1, max_col):
        alt = _cell_value(ws, meta.alt_row, c) if meta.alt_row is not None else None
        pname = (
            _cell_value(ws, meta.param_row, c) if meta.param_row is not None else None
        )
        ent_name = (
            _cell_value(ws, meta.entity_row, c) if meta.entity_row is not None else None
        )
        if not pname and meta.default_parameter:
            pname = meta.default_parameter
        if not alt or not ent_name or not pname:
            continue

        # The entity header carries only the FIRST dimension element for a
        # multi-dim class (the writer emits ``byname[0]``).  Stochastic
        # params in the schema are single-dimension (profile, node, …); a
        # multi-dim header would be ambiguous, so we only build a byname we
        # can trust.
        if n_dims == 1:
            entity_byname: tuple[str, ...] = (ent_name,)
        else:
            # Reaching here means a multi-dim class was routed onto an _s
            # sheet whose header can't carry the full byname — surface it
            # rather than silently fabricating a wrong entity.
            logger.warning(
                "Sheet '%s' col %d: stochastic sheet for multi-dim class "
                "%r cannot recover the full entity byname from a single "
                "header value %r; skipping column.",
                meta.sheet_name, c + 1, entity_class.class_name, ent_name,
            )
            continue

        for r in range(meta.data_start_row, max_row):
            val = _cell_value(ws, r, c)
            if not val:
                continue
            # Read all index levels for this row.
            level_vals: list[tuple[str, str | None]] = []
            for col_idx, axis in meta.stoch_index_cols:
                level_vals.append((axis, _cell_value(ws, r, col_idx) or None))
            # A row missing any index level is malformed — skip it.
            if any(v is None for _, v in level_vals):
                continue

            converted = _convert_value(val, leaf_dtype)
            record: dict[str, Any] = {
                "alternative": alt,
                "entity_class": entity_class.class_name,
                "entity_byname": entity_byname,
                "param_name": pname,
                "value": converted,
                "index_value": level_vals[-1][1],
                "index_name": level_vals[-1][0],
                "data_type": leaf_dtype,
            }
            if len(level_vals) > 1:
                record["extra_index_values"] = level_vals[:-1]
            data.records.append(record)

    return data


# ---------------------------------------------------------------------------
# Determine sheet type and parse accordingly
# ---------------------------------------------------------------------------


def detect_and_parse_sheet(ws: Worksheet) -> tuple[SheetMetadata | None, bool]:
    """Auto-detect sheet type and parse its metadata.

    Tries scenario first, then standard (constant/periodic), then link,
    then transposed.

    Returns:
        (metadata, is_scenario) — metadata is None for unrecognised sheets.
        is_scenario is True when the sheet is a scenario sheet (needs special
        extraction via parse_scenario_sheet).
    """
    # Skip very small sheets
    if (ws.max_row or 0) < 2 or (ws.max_column or 0) < 2:
        return None, False

    # Check for scenario sheet: "Scenario names" in row 1, col B
    if _is_scenario_sheet(ws):
        return SheetMetadata(sheet_name=ws.title), True

    # Check for a stochastic (_s) sheet: N>=2 left ``index:`` columns then a
    # definition column carrying the transposed entity/alt/param roles.  Must
    # run BEFORE the standard-sheet probe, which would otherwise misread the
    # transposed ``alternative`` value header as a parameter column.
    stoch_def_col = _find_stochastic_def_col(ws)
    if stoch_def_col is not None:
        return parse_stochastic_sheet_metadata(ws, stoch_def_col), False

    # Check for transposed sheet first: "alternative" or "parameter" in column B (row 0-4)
    for r in range(min(5, _actual_max_row(ws))):
        val_a = _cell_value(ws, r, 0)
        if val_a and _is_info_row(val_a):
            continue  # skip INFO rows
        val_b = _cell_value(ws, r, 1)  # column B
        if val_b:
            vl = val_b.lower()
            if vl in ("alternative", "parameter") or vl.startswith("parameter:"):
                return parse_transposed_sheet_metadata(ws), False

    # Check for standard sheet (has keywords like "description" in row 1)
    for c in range(2, min(20, ws.max_column or 1)):  # start at col C (skip A=navigate, B)
        val = _cell_value(ws, 0, c)
        if val and _is_keyword(val) and val.lower() != "navigate":
            return parse_sheet_metadata(ws), False

    # Check for link sheet (entity: in column A)
    for r in range(min(5, ws.max_row or 1)):
        val = _cell_value(ws, r, 0)
        if val.lower().startswith("entity:") or val.lower().startswith("entity "):
            return parse_link_sheet_metadata(ws), False

    # Check for transposed sheet (column B has keywords like "alternative", "parameter")
    for r in range(min(10, ws.max_row or 1)):
        val = _cell_value(ws, r, 1)
        if val and val.lower() in ("alternative", "parameter"):
            return parse_transposed_sheet_metadata(ws), False

    return None, False


# ---------------------------------------------------------------------------
# Data extraction
# ---------------------------------------------------------------------------


@dataclass
class SheetData:
    """Extracted data from one sheet, ready for DB import."""

    sheet_name: str
    metadata: SheetMetadata

    # For each entity class mapping:
    #   list of dicts with keys: alternative, entity_byname, param_name, value,
    #                            entity_existence (bool or None)
    records: list[dict[str, Any]] = field(default_factory=list)

    # Entity relationships (for link sheets): list of entity_byname tuples
    link_entities: list[tuple[str, ...]] = field(default_factory=list)


def extract_sheet_data(ws: Worksheet, meta: SheetMetadata) -> SheetData:
    """Extract data records from a parsed sheet."""
    if meta.is_stochastic:
        return _extract_stochastic(ws, meta)
    elif meta.is_transposed:
        return _extract_transposed(ws, meta)
    elif not meta.param_cols and meta.entity_classes:
        return _extract_link(ws, meta)
    else:
        return _extract_standard(ws, meta)


def _extract_standard(ws: Worksheet, meta: SheetMetadata) -> SheetData:
    """Extract data from a standard (constant/periodic) sheet."""
    data = SheetData(sheet_name=meta.sheet_name, metadata=meta)
    max_row = _actual_max_row(ws)

    # Find entity element columns (columns in definition row between entity def and first
    # non-structural column).  These are the dimension name columns.
    for c in sorted(meta.col_defs.keys()):
        defn = meta.col_defs[c].lower()
        if not (defn == "alternative" or defn.startswith("entity") or
                defn.startswith("filter") or defn.startswith("index")):
            continue
        if defn == "alternative" or defn.startswith("filter") or defn.startswith("index"):
            continue
        # This is an entity-related column but not the entity def itself
        # Actually, the entity def column has the "entity:" prefix.
        # The dimension columns follow it and have plain dimension names.

    # Simpler approach: entity def column + subsequent dimension columns
    # are identified from the entity_classes definitions.
    # The entity column index is the one with "entity:" in col_defs.
    entity_col: int | None = None
    for c, defn in meta.col_defs.items():
        if defn.lower().startswith("entity"):
            entity_col = c
            break

    # Dimension element columns follow the entity column
    # Their count = max dimensions across entity classes
    if meta.entity_classes:
        max_dims = max(len(ec.dimensions) for ec in meta.entity_classes)
    else:
        max_dims = 0

    if entity_col is not None and max_dims > 1:
        # For multi-dim entities, dimension element columns follow the entity column.
        # There are exactly max_dims dimension columns after the entity column.
        dim_cols = list(range(entity_col + 1, entity_col + 1 + max_dims))
    elif entity_col is not None:
        dim_cols = []  # single-dim: entity name is in entity_col itself
    else:
        dim_cols = []

    for r in range(meta.data_start_row, max_row):
        # Read alternative.  Normalise empty cells to ``None`` (rather than
        # ``""``) so downstream entity-only records propagate a real
        # absence — the writer's alt-creation step skips falsy names and
        # discards entity-only records before parameter writes, so a
        # synthesised ``""`` would create an empty-string alternative
        # in SpineDB.
        if meta.alt_col is not None:
            alt = _cell_value(ws, r, meta.alt_col) or None
        else:
            alt = None

        # Read entity byname
        if entity_col is not None:
            if max_dims <= 1:
                ent_name = _cell_value(ws, r, entity_col)
                entity_byname = (ent_name,) if ent_name else None
            else:
                elements = []
                for dc in dim_cols:
                    elements.append(_cell_value(ws, r, dc))
                if any(elements):
                    entity_byname = tuple(elements)
                else:
                    entity_byname = None
        else:
            entity_byname = None

        if entity_byname is None:
            continue

        # Determine entity class for this row (using filter if present)
        if meta.filter_col is not None and meta.filter_map and len(meta.entity_classes) > 1:
            filter_val = _cell_value(ws, r, meta.filter_col)
            entity_class = None
            for ec in meta.entity_classes:
                pattern = meta.filter_map.get(ec.class_name, "")
                if pattern and re.match(pattern, filter_val):
                    entity_class = ec
                    break
            if entity_class is None:
                logger.warning(
                    "Sheet '%s' row %d: filter value '%s' didn't match any entity class",
                    meta.sheet_name, r + 1, filter_val,
                )
                continue
        elif meta.entity_classes:
            entity_class = meta.entity_classes[0]
        else:
            continue

        # Read index value (for periodic/map sheets)
        index_val = None
        if meta.index_col is not None:
            index_val = _cell_value(ws, r, meta.index_col) or None

        # Read extra index values (e.g. ladder sheets carry index: period
        # + index: tier).  We keep them in the same column order as the
        # sheet so the DB-write side can reconstruct the nested Map.
        extra_index_values: list[tuple[str, str | None]] = []
        for c_extra, name_extra in meta.extra_index_cols:
            v = _cell_value(ws, r, c_extra) or None
            extra_index_values.append((name_extra, v))

        # Read entity existence (strict TRUE/FALSE — exporter only ever
        # writes those two literals; anything else is a data error).
        entity_existence = None
        if meta.entity_existence_col is not None:
            ee_val = _cell_value(ws, r, meta.entity_existence_col)
            if ee_val:
                entity_existence = _parse_strict_bool(
                    ee_val,
                    sheet_name=meta.sheet_name,
                    row_1based=r + 1,
                    col_1based=meta.entity_existence_col + 1,
                    field_name="entity existence",
                )

        # Read parameter values
        row_has_data = False
        for c, pname in meta.param_cols.items():
            if pname == ENTITY_EXISTENCE:
                continue  # already handled
            val = _cell_value(ws, r, c)
            if not val:
                continue

            # Convert based on data type.  Fallback order:
            #   1. explicit per-column data_type
            #   2. sheet default_data_type
            #   3. "string" — safe pass-through.  The exporter ALWAYS
            #      writes a data type for every parameter column, so
            #      reaching the fallback implies a hand-edited sheet;
            #      "string" preserves the cell verbatim instead of
            #      forcing a ``float()`` that may corrupt string-typed
            #      parameters (a previous "float" fallback turned cell
            #      "123" into 123.0 even for schema-string params).
            if c in meta.data_types:
                dtype = meta.data_types[c]
            elif meta.default_data_type:
                dtype = meta.default_data_type
            else:
                dtype = "string"
                logger.warning(
                    "Sheet '%s' col %d (parameter %r): no data type declared; "
                    "defaulting to 'string' (raw cell value).",
                    meta.sheet_name, c + 1, pname,
                )
            converted = _convert_value(val, dtype)

            record = {
                "alternative": alt,
                "entity_class": entity_class.class_name,
                "entity_byname": entity_byname,
                "param_name": pname,
                "value": converted,
                "index_value": index_val,
                "index_name": meta.index_name,
                "data_type": dtype,
            }
            if extra_index_values:
                record["extra_index_values"] = list(extra_index_values)
            data.records.append(record)
            row_has_data = True

        # Add entity existence record
        if entity_existence is not None:
            data.records.append({
                "alternative": alt,
                "entity_class": entity_class.class_name,
                "entity_byname": entity_byname,
                "param_name": ENTITY_EXISTENCE,
                "value": entity_existence,
                "index_value": None,
                "index_name": None,
            })
            row_has_data = True

        # Create entity-only record if no data was found for this row
        # (ensures data-less entities are still imported).  Propagate the
        # real ``alt`` even when None — the downstream writer
        # (``write_self_describing_to_db._import_sheet``) skips
        # alternative-creation for falsy alt names and discards
        # entity-only records before parameter writes, so a synthesised
        # ``""`` alt name would create an empty-string alternative in
        # SpineDB while a faithful ``None`` is dropped cleanly.
        if not row_has_data and entity_byname:
            data.records.append({
                "alternative": alt,
                "entity_class": entity_class.class_name,
                "entity_byname": entity_byname,
                "param_name": "",
                "value": None,
                "index_value": None,
                "index_name": None,
            })

    return data


def _extract_link(ws: Worksheet, meta: SheetMetadata) -> SheetData:
    """Extract entity relationships from a link sheet."""
    data = SheetData(sheet_name=meta.sheet_name, metadata=meta)
    max_row = _actual_max_row(ws)
    max_col = ws.max_column or 1

    for r in range(meta.data_start_row, max_row):
        elements = []
        for c in range(max_col):
            val = _cell_value(ws, r, c)
            if val:
                elements.append(val)
        if elements:
            data.link_entities.append(tuple(elements))

    return data


def _extract_transposed(ws: Worksheet, meta: SheetMetadata) -> SheetData:
    """Extract data from a transposed (timeseries) sheet."""
    data = SheetData(sheet_name=meta.sheet_name, metadata=meta)
    max_row = _actual_max_row(ws)
    max_col = ws.max_column or 1

    # Each data column (from col 2 onwards) is one data series
    data_col_start = 2  # columns C+ (0-based index 2)

    # On single-param transposed sheets (timeseries, array-transposed)
    # the parameter triplet in row 1 / col B carries the canonical data
    # type — e.g. ``"parameter: solves | data type: boolean (array)"``.
    # We extract it once and attach it to every emitted record so the
    # writer can decide whether to reconstruct an ``Array`` or a ``Map``.
    triplet_dtype: str | None = None
    if meta.param_row is not None:
        triplet_text = _cell_value(ws, meta.param_row, 1)  # col B
        triplet_dtype = _extract_triplet_field(triplet_text, "data type")
    if triplet_dtype is None and meta.default_data_type:
        triplet_dtype = meta.default_data_type

    for c in range(data_col_start, max_col):
        # Read column metadata from the header rows
        alt = _cell_value(ws, meta.alt_row, c) if meta.alt_row is not None else None
        pname = _cell_value(ws, meta.param_row, c) if meta.param_row is not None else None
        ent_name = _cell_value(ws, meta.entity_row, c) if meta.entity_row is not None else None

        if not pname and meta.default_parameter:
            pname = meta.default_parameter
        if not alt or not ent_name:
            continue

        entity_class = meta.entity_classes[0] if meta.entity_classes else None
        if entity_class is None:
            continue

        entity_byname = (ent_name,)

        # Per-column data type (multi-param _t sheets) takes precedence over
        # the single-param triplet; fall back to the triplet then "string".
        col_dtype = meta.data_types.get(c) or triplet_dtype

        # Read time-indexed data
        for r in range(meta.data_start_row, max_row):
            index_val = _cell_value(ws, r, 0) if meta.index_col_transposed == 0 else None
            val = _cell_value(ws, r, c)
            if not val or not index_val:
                continue

            # Use the column/triplet data type for value conversion.
            # The pre-fix code force-cast every cell to float, which only
            # worked by accident for ``boolean (array)`` (period names
            # like ``"Y2025"`` raise on ``float()`` and ``_convert_value``
            # falls back to returning the raw string).  A schema-string
            # series with numeric-looking entries (e.g. ``"01"``) would
            # silently get coerced to ``1.0`` and round-trip wrong.
            value_dtype = col_dtype or "string"
            converted = _convert_value(val, value_dtype)
            data.records.append({
                "alternative": alt,
                "entity_class": entity_class.class_name,
                "entity_byname": entity_byname,
                "param_name": pname,
                "value": converted,
                "index_value": index_val,
                "index_name": meta.index_name,
                "data_type": col_dtype,
            })

    return data


def _parse_strict_bool(
    val: str,
    *,
    sheet_name: str,
    row_1based: int,
    col_1based: int,
    field_name: str,
) -> bool:
    """Strict TRUE / FALSE parse, case-insensitive, mirroring the exporter.

    The exporter only ever writes the literal strings ``"TRUE"`` / ``"FALSE"``
    for entity-existence and boolean-array cells (excel_writer.py:2272 and
    :2773 — the dropdown validation also enforces those exact tokens).  A
    cell holding anything else (``"yes"``, ``"1"``, ``"Y"``, a typo) is a
    data error: raise so the user sees the offending sheet / cell rather
    than silently importing ``False``.
    """
    stripped = val.strip().lower()
    if stripped == _BOOL_TRUE_LITERAL:
        return True
    if stripped == _BOOL_FALSE_LITERAL:
        return False
    raise ValueError(
        f"Sheet '{sheet_name}' row {row_1based} col {col_1based}: "
        f"{field_name} cell has value {val!r}; expected 'TRUE' or 'FALSE' "
        f"(case-insensitive).  The exporter only emits those literals — "
        f"fix the cell or restore the exporter-written value."
    )


def _convert_value(val: str, dtype: str) -> Any:
    """Convert a string cell value to the appropriate Python type.

    ``float`` cells accept the IEEE non-finite sentinels ``inf`` /
    ``-inf`` / ``nan`` (case-insensitive) — the export writer emits
    them as strings because openpyxl drops actual non-finite floats.

    The leaf scalar type is taken from the part of *dtype* BEFORE any
    ``(...)`` container suffix: ``"float (1d-map)"``, ``"float (array)"``,
    ``"float (3d-map)"`` all carry a ``float`` leaf and must convert their
    per-cell values to ``float`` — otherwise a Map/Array would round-trip
    with string leaves (``["10", "20"]`` instead of ``[10.0, 20.0]``).
    Compound leaf types joined with ``/`` (e.g. ``"string/float (1d-map)"``
    — a param that accepts either) attempt the ``float`` conversion when
    ``float`` is one of the tokens, falling back to the raw string for a
    genuinely non-numeric cell.  Only the exact scalar ``"boolean"``
    strict-parses TRUE/FALSE: a ``"boolean (array)"`` cell holds a period
    token (the round-trip preservation form, see _extract_standard) and
    must stay raw.
    """
    base = dtype.split("(", 1)[0].strip().lower() if dtype else ""
    base_tokens = base.replace("/", " ").split()
    if "float" in base_tokens:
        if isinstance(val, str):
            lo = val.strip().lower()
            if lo in ("inf", "+inf", "infinity", "+infinity"):
                return float("inf")
            if lo in ("-inf", "-infinity"):
                return float("-inf")
            if lo == "nan":
                return float("nan")
        try:
            return float(val)
        except (ValueError, TypeError):
            return val  # keep as string if conversion fails
    elif dtype == "boolean":
        # Strict match — the exporter only writes ``"TRUE"`` / ``"FALSE"``
        # for boolean cells (see excel_writer.py:2272 / :2773).  Without
        # call-site context we can't name the offending sheet/row, so the
        # error message asks the user to inspect their workbook; the
        # per-cell entity-existence path uses ``_parse_strict_bool`` and
        # gives a much sharper message.
        stripped = val.strip().lower()
        if stripped == _BOOL_TRUE_LITERAL:
            return True
        if stripped == _BOOL_FALSE_LITERAL:
            return False
        raise ValueError(
            f"Boolean cell has value {val!r}; expected 'TRUE' or 'FALSE' "
            f"(case-insensitive)."
        )
    else:
        return val


# ---------------------------------------------------------------------------
# Public API: read an entire workbook
# ---------------------------------------------------------------------------


def read_self_describing_excel(
    file_path: str,
    skip_sheets: set[str] | None = None,
) -> list[SheetData]:
    """Read all self-describing sheets from an Excel workbook.

    Args:
        file_path: Path to the .xlsx file.
        skip_sheets: Sheet names to skip (e.g. navigate, version).

    Returns:
        List of SheetData objects, one per successfully parsed sheet.
    """
    if skip_sheets is None:
        skip_sheets = {"navigate", "version"}

    # Note: do NOT use read_only=True — it reports max_row=1048576 for
    # sheets without explicit dimensions, which makes iteration hang.
    wb = openpyxl.load_workbook(file_path, data_only=True)
    results: list[SheetData] = []

    for sheet_name in wb.sheetnames:
        if sheet_name in skip_sheets:
            continue

        ws = wb[sheet_name]
        meta, is_scenario = detect_and_parse_sheet(ws)

        if is_scenario:
            sheet_data = parse_scenario_sheet(ws)
            results.append(sheet_data)
            logger.info(
                "Sheet '%s' (scenario): %d records",
                sheet_name, len(sheet_data.records),
            )
            continue

        if meta is None:
            logger.debug("Skipping sheet '%s': no metadata found", sheet_name)
            continue

        sheet_data = extract_sheet_data(ws, meta)
        results.append(sheet_data)
        logger.info(
            "Sheet '%s': %d records, %d links",
            sheet_name, len(sheet_data.records), len(sheet_data.link_entities),
        )

    wb.close()
    return results
