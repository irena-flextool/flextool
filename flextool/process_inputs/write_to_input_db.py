import os
from spinedb_api import DatabaseMapping, from_database, to_database, SpineDBAPIError, Asterisk
from spinedb_api.exception import NothingToCommit
from flextool.update_flextool import FLEXTOOL_DB_VERSION

flextool_db_version = FLEXTOOL_DB_VERSION


def _import_legacy_timeset_timeline(db, input_path: str) -> None:
    """Set the ``timeset.timeline`` parameter from a legacy Excel sheet.

    Pre-v25 specification-format Excels store the timeset->timeline link as a
    two-column ``timeblockSet_timeline`` sheet (column 0 = timeset name,
    column 1 = timeline name; a header row, no alternative column).  The
    modern ``timeset_timeline`` sheet has a different layout, so this link is
    not picked up by the generic spec mappings.

    The v25 migration (``update_timestructure``) normally collapses the old
    ``timeblockSet__timeline`` relationship into the ``timeset.timeline``
    parameter, attaching each value to the alternative of that timeset's
    duration.  Mirror that here so a pre-v25 Excel imports directly into the
    v25+ schema: for every ``(timeset, alternative)`` that has a
    ``timeset_duration`` value, set ``timeline`` in the same alternative.
    Without it the solve fails with "Failed to map timeset to timeline".
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    except Exception:
        return
    try:
        if "timeblockSet_timeline" not in wb.sheetnames:
            return
        timeset_to_timeline: dict[str, str] = {}
        for i, row in enumerate(wb["timeblockSet_timeline"].iter_rows(values_only=True)):
            if i == 0:  # header row: (timeblockSet, timeline)
                continue
            if not row or row[0] is None or row[1] is None:
                continue
            timeset_to_timeline[str(row[0])] = str(row[1])
    finally:
        wb.close()

    if not timeset_to_timeline:
        return

    # Already-present (entity, alternative) pairs — stay idempotent.
    existing = {
        (pv["entity_byname"], pv["alternative_name"])
        for pv in db.find_parameter_values(
            entity_class_name="timeset", parameter_definition_name="timeline"
        )
    }
    added = 0
    for pv in db.find_parameter_values(
        entity_class_name="timeset", parameter_definition_name="timeset_duration"
    ):
        byname = pv["entity_byname"]
        timeline_name = timeset_to_timeline.get(byname[0])
        if timeline_name is None:
            continue
        key = (byname, pv["alternative_name"])
        if key in existing:
            continue
        value, type_ = to_database(timeline_name)
        try:
            db.add_parameter_value(
                entity_class_name="timeset",
                parameter_definition_name="timeline",
                entity_byname=byname,
                alternative_name=pv["alternative_name"],
                value=value,
                type=type_,
            )
            existing.add(key)
            added += 1
        except Exception as e:
            print(f"  Could not set timeset.timeline for {byname[0]}: {e}")
    if added:
        print(
            f"\nImported legacy timeset->timeline links "
            f"({added} (timeset, alternative) pairs) from 'timeblockSet_timeline'."
        )

def write_to_flextool_input_db(input_path, tabular_reader, target_db_url, input_type='excel',
                               migration_follows: bool = False):
    """Write tabular data to FlexTool input database, reading one sheet/file at a time.

    Args:
        input_path: Path to Excel/ODS file or directory containing CSV files
        tabular_reader: TabularReader instance with loaded specification
        target_db_url: URL to target SQLite database
        input_type: Either 'excel' or 'csv' to determine reading method
        migration_follows: If True, accept a version mismatch because
            the caller will migrate the database after import

    This function reads sheets/files one at a time to minimize memory usage.
    """
    # Refuse specification-format Excels whose template layout predates the
    # importer.  Deferred import avoids a circular dependency with the package.
    if input_type == 'excel':
        from flextool.process_inputs import (
            MIN_SUPPORTED_SPECIFICATION_VERSION,
            read_specification_changelog_version,
            unsupported_specification_message,
        )
        changelog_version = read_specification_changelog_version(input_path)
        if (
            changelog_version is not None
            and changelog_version < MIN_SUPPORTED_SPECIFICATION_VERSION
        ):
            print(unsupported_specification_message(changelog_version))
            exit(1)

    if target_db_url.startswith('sqlite://') or target_db_url.startswith('http://'):
        db_url = target_db_url
    elif os.path.exists(target_db_url) and target_db_url.endswith(".sqlite"):
        db_url = 'sqlite:///' + target_db_url
    else:
        print("No sqlite file at " + target_db_url)
        exit(-1)

    with DatabaseMapping(db_url, create = False, upgrade = True) as db:
        sq = db.object_parameter_definition_sq
        version_parameter = db.query(sq).filter(sq.c.object_class_name == "model").filter(sq.c.parameter_name == "version").one_or_none()
        if version_parameter is None:
            #if no version assume version 0
            print(f"No FlexTool database version found. Needs to be proper FlexTool database at version: {flextool_db_version}")
            exit(-1)
        else:
            version = from_database(version_parameter.default_value, version_parameter.default_type)
            if version == flextool_db_version:
                print(f"Valid FlexTool input database with correct version: {flextool_db_version}")
            elif migration_follows:
                print(f"Importing into a fresh FlexTool schema-v{int(version)} database "
                      "(a migration to the current version follows this import).")
            else:
                print(f"Wrong FlexTool input database version: {version}. Should be: {flextool_db_version}")
                exit(-1)

        # Get list of tables to process
        selected_tables = tabular_reader.get_selected_table_names()
        print(f"\nProcessing {len(selected_tables)} selected tables")

        # Purge the database empty of data times
        try:
            db.remove_entity(id=Asterisk)
            db.remove_alternative(id=Asterisk)
            db.remove_scenario(id=Asterisk)
            db.remove_parameter_value(id=Asterisk)
            db.commit_session("Purged the db")
        except Exception as e:
            raise RuntimeError(f"Failed to purge the input database before writing. {e}")

        # Process each table one at a time
        for table_name in selected_tables:
            print(f"\nProcessing table {table_name + ': ':<30}", end="")

            # Read the sheet/file that contains the data
            try:
                if input_type == 'excel':
                    raw_df, column_types = tabular_reader.read_excel_sheet(input_path, table_name)
                elif input_type == 'csv':
                    csv_file = os.path.join(input_path, f"{table_name}.csv")
                    if not os.path.exists(csv_file):
                        print(f"  Warning: CSV file not found: {csv_file}")
                        continue
                    raw_df, column_types = tabular_reader.read_csv_file(csv_file)
                else:
                    raise ValueError(f"Unknown input_type: {input_type}")

                # Skip if sheet not found in the file
                if raw_df is None:
                    print(f"Skipping {table_name} - no sheet found")
                    continue

                # Skip if no data (e.g., sheet not selected)
                if raw_df.empty:
                    print(f"  Skipping {table_name} - no data")
                    continue

            except Exception as e:
                print(f"  Error reading table '{table_name}': {e}")
                continue  # Continue to next table on read error

            # Get mappings for this sheet
            mappings = tabular_reader.get_table_mappings(table_name)
            if not mappings:
                raise ValueError(f"No specification found for sheet {table_name}")

            table_options = tabular_reader.get_table_options(table_name)
            if 'row' in table_options:
                raw_df = raw_df.iloc[table_options['row']:]
            if 'column' in table_options:
                raw_df = raw_df.iloc[table_options['column']:]

            # Process each mapping in the table
            for mapping_name in mappings.keys():
                print(f"{mapping_name:>25}", end="")

                mapping_info = tabular_reader._parse_mapping(mappings[mapping_name])
                if not mapping_info:
                    continue

                # Read data and process the dataframe
                try:
                    (data_df, ent_zip_list, ent_act_zip, scen_array, scen_alt_df) = tabular_reader._extract_data(raw_df, mapping_info, table_options, column_types, table_name, mapping_name)
                except Exception as e:
                    print(
                        f"\n  Error processing sheet '{table_name}', mapping "
                        f"'{mapping_name}': {type(e).__name__}: {e}"
                    )
                    continue  # Continue to next mapping on processing error

                if ent_zip_list:
                    tabular_reader._add_entities(ent_zip_list, db, table_name, mapping_name)

                # Add alternatives to the database
                if data_df is not None:
                    alt_array = data_df.columns.get_level_values('Alternative').unique().to_list()
                    tabular_reader._add_alternatives(alt_array, db, table_name, mapping_name)

                # Write re-organised dataframe to database
                if data_df is not None:
                #try:
                    # check_type:
                    value_type = mapping_info['rules'].get('ParameterValueType')
                    if value_type:
                        value_type = value_type['value']
                    else:
                        value_type = 'constant'
                    tabular_reader._add_parameters(data_df, db, table_name, mapping_name, value_type)

                # Ensure alternatives referenced by entity_alternatives exist
                if ent_act_zip is not None:
                    ent_act_list = list(ent_act_zip)
                    ea_alts = list({alt for _cls, _ent, alt, _act in ent_act_list})
                    if ea_alts:
                        tabular_reader._add_alternatives(ea_alts, db, table_name, mapping_name)
                    tabular_reader._add_entity_alternatives(ent_act_list, db, table_name, mapping_name)

                # Add scenarios to the database
                if scen_array is not None:
                    for scen in scen_array:
                        try:
                            db.add_scenario(name=scen)
                        except Exception as e:
                            raise SpineDBAPIError(f'Could not add scenario {scen} to the database: {e}')

                # Add scenario_alternatives to the database
                if scen_alt_df is not None:
                    tabular_reader._add_scenario_alternatives(scen_alt_df, db, table_name, mapping_name)

        # Backfill the timeset->timeline link from the legacy two-column sheet
        # if present (pre-v25 Excels); modern files set it via the spec.
        if input_type == 'excel':
            _import_legacy_timeset_timeline(db, input_path)

        # Commit the changes
        try:
            db.commit_session(f"Imported data from mapping '{mapping_name}'")
            tabular_reader.logger.info(f"Successfully committed data and items from mapping {mapping_name}")
        except NothingToCommit:
            pass
        except Exception as e:
            raise SpineDBAPIError(f"Could not commit data and items based on {mapping_name}: {e}")

    print("\nAll data processed successfully!    (* = No data found for the mapping)")
