"""Input data reading: converts CSV/Excel/ODS files to Spine database format."""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from enum import Enum
from pathlib import Path

from flextool.process_inputs.read_tabular_with_specification import TabularReader
from flextool.process_inputs.write_to_input_db import write_to_flextool_input_db
from flextool.update_flextool import FLEXTOOL_DB_VERSION as CURRENT_FLEXTOOL_DB_VERSION

__all__ = [
    'TabularReader', 'write_to_flextool_input_db',
    'ExcelFormat', 'ExcelFormatInfo', 'detect_excel_format',
    'CURRENT_FLEXTOOL_DB_VERSION',
    'MIN_SUPPORTED_SPECIFICATION_VERSION',
    'read_specification_changelog_version',
    'unsupported_specification_message',
]

logger = logging.getLogger(__name__)

# Oldest specification-format Excel *template changelog* version the importer
# can read.  Below this the sheet/parameter layout diverges too much (the
# changelog-v11 source/sink -> input/output rename and earlier) to bridge with
# the parameter/sheet remaps in read_tabular_with_specification.py.
MIN_SUPPORTED_SPECIFICATION_VERSION = 12


class ExcelFormat(Enum):
    """Excel input format variants recognised by FlexTool."""
    SELF_DESCRIBING = "self_describing"   # New format — metadata embedded in sheets
    SPECIFICATION = "specification"       # FlexTool 3.x — needs import_excel_input.json
    OLD_V2 = "old_v2"                     # FlexTool 2.x .xlsm — completely different layout
    UNKNOWN = "unknown"


@dataclass
class ExcelFormatInfo:
    """Result of :func:`detect_excel_format`."""
    format: ExcelFormat
    version: int | None = None  # DB schema version found in the file (None = unknown)


def detect_excel_format(file_path: str | Path) -> ExcelFormatInfo:
    """Detect which FlexTool Excel format *file_path* uses.

    Returns an :class:`ExcelFormatInfo` with the detected format and,
    where possible, the database schema version embedded in the file.

    Detection logic:

    * **SELF_DESCRIBING** — ``version`` sheet, cell A1 starts with
      ``"Generated from FlexTool sqlite version"``.  The version number
      is parsed from that same cell.
    * **SPECIFICATION** — ``version`` sheet, row 3 contains column
      headers ``Version number, Date, Author, Description``
      (FlexTool 3.x format that needs ``import_excel_input.json``).
      The version is always 25 (the schema these files target).
    * **OLD_V2** — file has a ``master`` sheet (FlexTool 2.x).
    * **UNKNOWN** — none of the above.
    """
    import openpyxl

    file_path = Path(file_path)
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception:
        logger.warning("Cannot open '%s' as an Excel file.", file_path)
        return ExcelFormatInfo(ExcelFormat.UNKNOWN)

    try:
        sheet_names_lower = {s.lower(): s for s in wb.sheetnames}

        # --- Check version sheet ---
        if "version" in sheet_names_lower:
            ws = wb[sheet_names_lower["version"]]
            rows = list(ws.iter_rows(max_row=4, values_only=True))

            # New self-describing: first cell starts with
            # "Generated from FlexTool sqlite version: <N>"
            if rows and rows[0] and rows[0][0]:
                first_cell = str(rows[0][0]).strip()
                if first_cell.lower().startswith("generated from flextool sqlite"):
                    version: int | None = None
                    m = re.search(r"version:\s*(\d+)", first_cell)
                    if m:
                        version = int(m.group(1))
                    logger.info(
                        "Detected SELF_DESCRIBING format (version %s): %s",
                        version, file_path.name,
                    )
                    return ExcelFormatInfo(ExcelFormat.SELF_DESCRIBING, version)

            # Old 3.x: row 2 (0-indexed) has "Version number", "Date", …
            if len(rows) >= 3 and rows[2]:
                header = [str(c).strip().lower() if c else "" for c in rows[2]]
                if "version number" in header and "date" in header:
                    # Specification-format files always target the v25 schema
                    logger.info(
                        "Detected SPECIFICATION (3.x) format: %s", file_path.name,
                    )
                    return ExcelFormatInfo(ExcelFormat.SPECIFICATION, 25)

        # --- Check navigate sheet for v2 self-describing format ---
        # When exported in v2 format, the version sheet is removed and
        # the version is written to the navigate sheet (row 15, col F-G).
        if "navigate" in sheet_names_lower:
            ws_nav = wb[sheet_names_lower["navigate"]]
            nav_version_label = ws_nav.cell(row=15, column=6).value
            if nav_version_label and "flextool db version" in str(nav_version_label).lower():
                version = None
                nav_version_val = ws_nav.cell(row=15, column=7).value
                if nav_version_val is not None:
                    try:
                        version = int(nav_version_val)
                    except (ValueError, TypeError):
                        pass
                logger.info(
                    "Detected SELF_DESCRIBING v2 format (version %s): %s",
                    version, file_path.name,
                )
                return ExcelFormatInfo(ExcelFormat.SELF_DESCRIBING, version)

        # --- Check for old 2.x .xlsm (has 'master' sheet) ---
        if "master" in sheet_names_lower:
            logger.info("Detected OLD_V2 format: %s", file_path.name)
            return ExcelFormatInfo(ExcelFormat.OLD_V2)

        logger.warning("Could not detect Excel format for '%s'.", file_path.name)
        return ExcelFormatInfo(ExcelFormat.UNKNOWN)

    finally:
        wb.close()


def read_specification_changelog_version(file_path: str | Path) -> int | None:
    """Return the template version of a specification-format Excel.

    Specification (3.x) Excels carry a ``version`` sheet whose changelog table
    lists template versions in the first column (``Version number, Date,
    Author, Description`` header on row 3).  The file's version is the highest
    integer in that column.  Used to gate import support and to message the
    user when a file's layout predates the importer.

    Returns ``None`` when the file has no such changelog (e.g. a
    self-describing Excel, a CSV directory, or an unreadable file), so callers
    can safely skip the check for non-specification inputs.
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception:
        return None
    try:
        sheet_names_lower = {s.lower(): s for s in wb.sheetnames}
        if "version" not in sheet_names_lower:
            return None
        ws = wb[sheet_names_lower["version"]]
        best: int | None = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 3 or not row:  # skip the intro text + the header row (idx 2)
                continue
            cell = row[0]
            if cell is None:
                continue
            m = re.match(r"\s*(\d+)", str(cell))
            if m:
                value = int(m.group(1))
                best = value if best is None else max(best, value)
        return best
    finally:
        wb.close()


def unsupported_specification_message(version: int | None) -> str:
    """User-facing message for specification Excels older than support."""
    shown = f"v{version}" if version is not None else "an unsupported version"
    return (
        f"The layout of the input file corresponds to {shown}. This is not "
        f"supported by the current importer, which supports imports from "
        f"v{MIN_SUPPORTED_SPECIFICATION_VERSION} onward. You can get it "
        f"imported by using legacy FlexTool (checkout branch 'master') with "
        f"Spine Toolbox. Alternatively, convert it manually."
    )
