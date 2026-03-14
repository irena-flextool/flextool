from __future__ import annotations

import logging
import sys
from pathlib import Path

from flextool.gui.data_models import InputSourceInfo, ProjectSettings, ScenarioInfo
from flextool.gui.settings_io import save_project_settings

logger = logging.getLogger(__name__)

# File extensions recognised as input sources
SUPPORTED_EXTENSIONS = {".xlsx", ".ods", ".sqlite"}


class InputSourceManager:
    """Manages input source files and their scenario discovery.

    Scans the ``input_sources/`` subdirectory of a project, reads scenario
    names from xlsx and sqlite files, and keeps track of persistent source
    numbering via project settings.
    """

    def __init__(self, project_path: Path, settings: ProjectSettings) -> None:
        self.project_path = project_path
        self.settings = settings
        self.input_dir = project_path / "input_sources"
        self._sources: list[InputSourceInfo] = []
        # Track mtimes for files opened via Edit button (Linux xlsx fallback)
        self._last_known_mtimes: dict[str, float] = {}
        # Track which sources were explicitly opened for editing
        self._editing_sources: set[str] = set()

    # ── Public API ────────────────────────────────────────────────────

    def scan_input_sources(self) -> list[InputSourceInfo]:
        """List files in the input_sources/ directory.

        Returns an :class:`InputSourceInfo` for each ``.xlsx``, ``.ods``, and
        ``.sqlite`` file found.  No scenario reading is performed here.
        """
        sources: list[InputSourceInfo] = []
        if not self.input_dir.is_dir():
            return sources

        for filepath in sorted(self.input_dir.iterdir()):
            if filepath.suffix.lower() not in SUPPORTED_EXTENSIONS:
                continue
            if not filepath.is_file():
                continue
            file_type = filepath.suffix.lstrip(".").lower()
            sources.append(
                InputSourceInfo(
                    name=filepath.name,
                    file_type=file_type,
                    number=0,  # assigned later in refresh()
                    status="error",  # updated later
                )
            )
        return sources

    def read_scenarios_xlsx(self, filepath: Path) -> list[str] | None:
        """Read scenario names from the 'scenario' sheet in an xlsx file.

        Reads the 2nd row, skips the first column, and reads cell values
        until an empty cell is encountered.  Returns ``None`` if the sheet
        does not exist or openpyxl is not available.
        """
        try:
            import openpyxl
        except ImportError:
            logger.warning("openpyxl is not installed -- cannot read xlsx scenarios")
            return None

        try:
            wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
        except Exception:
            logger.warning("Failed to open workbook: %s", filepath, exc_info=True)
            return None

        try:
            if "scenario" not in wb.sheetnames:
                wb.close()
                return None

            ws = wb["scenario"]
            scenarios: list[str] = []
            # Row 2 (1-indexed), skip first column (start from column 2)
            for cell in ws[2][1:]:  # type: ignore[index]
                value = cell.value
                if value is None or (isinstance(value, str) and value.strip() == ""):
                    break
                scenarios.append(str(value).strip())
            wb.close()
            return scenarios
        except Exception:
            logger.warning("Error reading scenarios from %s", filepath, exc_info=True)
            try:
                wb.close()
            except Exception:
                pass
            return None

    def read_scenarios_sqlite(self, filepath: Path) -> list[str] | None:
        """Read scenario names from a Spine database sqlite file.

        Uses ``spinedb_api.DatabaseMapping`` to extract scenario names (or
        alternative names as a fallback).  Returns ``None`` if the file is
        not a valid Spine database or spinedb_api is not installed.
        """
        try:
            from spinedb_api import DatabaseMapping
        except ImportError:
            logger.warning("spinedb_api is not installed -- cannot read sqlite scenarios")
            return None

        db_url = f"sqlite:///{filepath}"
        try:
            with DatabaseMapping(db_url) as db_map:
                # Try scenarios first
                try:
                    items = db_map.get_scenario_items()
                    scenarios = [x["name"] for x in items]
                    if scenarios:
                        return scenarios
                except (AttributeError, Exception):
                    pass

                # Fallback to alternatives
                try:
                    items = db_map.get_alternative_items()
                    alternatives = [x["name"] for x in items]
                    if alternatives:
                        return alternatives
                except (AttributeError, Exception):
                    pass

                return []
        except Exception:
            logger.warning("Failed to read spine db: %s", filepath, exc_info=True)
            return None

    def refresh(self) -> list[InputSourceInfo]:
        """Re-scan directory, re-read all scenarios, and assign persistent numbers.

        Numbers are auto-incremented and persisted in ``settings.yaml``.
        Returns the updated list of :class:`InputSourceInfo`.
        """
        sources = self.scan_input_sources()

        # Assign persistent numbers
        numbers = dict(self.settings.input_source_numbers)
        next_number = max(numbers.values(), default=0) + 1

        for source in sources:
            if source.name in numbers:
                source.number = numbers[source.name]
            else:
                source.number = next_number
                numbers[source.name] = next_number
                next_number += 1

        # Read scenarios for each source and determine status
        for source in sources:
            filepath = self.input_dir / source.name
            if not filepath.exists():
                source.status = "error"
                continue

            # Check for lock file (editing indicator)
            if self._check_lock(filepath):
                source.status = "editing"
                continue

            scenarios = self._read_scenarios(source, filepath)
            if scenarios is None:
                source.status = "error"
            elif len(scenarios) == 0:
                source.status = "error"
            else:
                source.status = "ok"
                source.scenarios = scenarios

        # Persist updated numbers
        self.settings.input_source_numbers = numbers
        save_project_settings(self.project_path, self.settings)

        self._sources = sources
        return sources

    def get_all_scenarios(
        self, selected_sources: list[str] | None = None
    ) -> list[ScenarioInfo]:
        """Get all scenarios, optionally filtered by source names.

        If *selected_sources* is ``None`` or empty, returns scenarios from
        all sources.
        """
        result: list[ScenarioInfo] = []
        for source in self._sources:
            if selected_sources and source.name not in selected_sources:
                continue
            if source.status != "ok":
                continue
            for scenario_name in source.scenarios:
                result.append(
                    ScenarioInfo(
                        name=scenario_name,
                        source_number=source.number,
                        source_name=source.name,
                    )
                )
        return result

    def check_db_versions(self) -> list[str]:
        """Check and upgrade all sqlite input sources.

        Delegates to :func:`~flextool.gui.db_version_check.check_and_upgrade_database`
        for each ``.sqlite`` file in the input sources directory.

        Returns a list of human-readable upgrade messages (empty if nothing
        was upgraded).
        """
        all_messages: list[str] = []
        if not self.input_dir.is_dir():
            return all_messages

        for filepath in sorted(self.input_dir.iterdir()):
            if filepath.suffix.lower() != ".sqlite":
                continue
            if not filepath.is_file():
                continue
            try:
                from flextool.gui.db_version_check import check_and_upgrade_database

                _upgraded, messages = check_and_upgrade_database(filepath)
                all_messages.extend(messages)
            except Exception as exc:
                all_messages.append(f"{filepath.name}: version check error: {exc}")
                logger.warning(
                    "Version check failed for %s: %s", filepath, exc, exc_info=True
                )

        return all_messages

    # ── Private helpers ───────────────────────────────────────────────

    def _read_scenarios(
        self, source: InputSourceInfo, filepath: Path
    ) -> list[str] | None:
        """Dispatch scenario reading based on file type."""
        ext = filepath.suffix.lower()
        if ext in (".xlsx", ".ods"):
            return self.read_scenarios_xlsx(filepath)
        elif ext == ".sqlite":
            return self.read_scenarios_sqlite(filepath)
        return None

    def mark_as_editing(self, source_name: str) -> None:
        """Mark a source as being edited (store mtime for comparison).

        Called when the Edit button opens a file, so that subsequent
        refreshes can detect ongoing editing via mtime comparison
        (Linux xlsx fallback).
        """
        filepath = self.input_dir / source_name
        self._editing_sources.add(source_name)
        try:
            self._last_known_mtimes[source_name] = filepath.stat().st_mtime
        except OSError:
            pass

    def _check_lock(self, filepath: Path) -> bool:
        """Detect if the file is currently being edited.

        For .xlsx / .ods files:
        - Windows/macOS: check for ``~$filename`` lock file in same directory.
        - Linux: check for ``.~lock.filename#`` (LibreOffice lock file).
        - Linux fallback: if the file was opened via the Edit button,
          compare ``st_mtime`` against the stored value. If mtime
          changed, the file is still being edited.

        For .sqlite files:
        - Check for ``<filepath>-journal`` or ``<filepath>-wal`` files.

        Returns ``True`` if the file appears to be in an editing state.
        """
        ext = filepath.suffix.lower()

        if ext in (".xlsx", ".ods"):
            parent = filepath.parent
            name = filepath.name

            # Windows / macOS Excel lock file: ~$<filename>
            excel_lock = parent / f"~${name}"
            if excel_lock.exists():
                return True

            # LibreOffice lock file (Linux): .~lock.<filename>#
            libreoffice_lock = parent / f".~lock.{name}#"
            if libreoffice_lock.exists():
                return True

            # Linux mtime-based fallback for files opened via Edit button
            if (
                sys.platform not in ("win32", "darwin")
                and name in self._editing_sources
            ):
                try:
                    current_mtime = filepath.stat().st_mtime
                except OSError:
                    return False
                last_mtime = self._last_known_mtimes.get(name)
                if last_mtime is not None and current_mtime != last_mtime:
                    # mtime changed -- file was modified externally
                    return True

            return False

        if ext == ".sqlite":
            journal = filepath.parent / (filepath.name + "-journal")
            wal = filepath.parent / (filepath.name + "-wal")
            if journal.exists() or wal.exists():
                return True
            return False

        return False
