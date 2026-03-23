"""Cell formatting constants and helper functions for the Excel export.

Provides fill colours, fonts, and per-layout formatting functions that match
the existing FlexTool Excel template.  All theme colours reference the Office
default colour theme shipped with openpyxl workbooks.
"""

from __future__ import annotations

from openpyxl.styles import Font, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Colour / font constants
# ---------------------------------------------------------------------------

# Header row (row 2 in constant/periodic sheets) fills
FILL_ALT_HEADER = PatternFill(
    patternType="solid", fgColor=Color(theme=9, tint=0.3999755851924192)
)
FILL_ENTITY_HEADER = PatternFill(
    patternType="solid", fgColor=Color(theme=7, tint=0.5999938962981048)
)
FILL_PARAM_HEADER = PatternFill(
    patternType="solid", fgColor=Color(theme=4, tint=0.5999938962981048)
)

# Data row fills (rows 3+ in constant/periodic sheets)
FILL_ALT_DATA = PatternFill(
    patternType="solid", fgColor=Color(theme=9, tint=0.5999938962981048)
)
FILL_ENTITY_DATA = PatternFill(
    patternType="solid", fgColor=Color(theme=7, tint=0.7999816888943144)
)

# Timeseries-specific fills
FILL_TIME_HEADER = PatternFill(
    patternType="solid", fgColor=Color(theme=0, tint=-0.249977111117893)
)
FILL_TIME_DATA = PatternFill(
    patternType="solid", fgColor=Color(theme=0, tint=-0.1499984740745262)
)

# Description/data-type label cells (row 1-2, definition column) – medium grey, black font
FILL_DESC_ROW = PatternFill(
    patternType="solid", fgColor=Color(theme=0, tint=-0.3499862666707358)
)
FONT_DESC_ROW = Font(color=Color(theme=1, tint=0.0))

# Description/data-type data cells (row 1-2, parameter columns) – lighter grey, black font
FILL_DESC_DATA = PatternFill(
    patternType="solid", fgColor=Color(theme=0, tint=-0.249977111117893)
)
FONT_DESC_DATA = Font(color=Color(theme=1, tint=0.0))

# Definition column background (darker grey) — conveys "do not fill here"
FILL_DEF_COL = PatternFill(
    patternType="solid", fgColor=Color(theme=0, tint=-0.3499862666707358)
)

# 'navigate' label in row 1 of each sheet (blue link colour)
FONT_NAVIGATE_LINK = Font(color=Color(theme=10, tint=0.0))

# Parameter *label* cell (e.g. the word "parameter" in timeseries sheets)
FILL_PARAM_LABEL = PatternFill(
    patternType="solid", fgColor=Color(theme=4, tint=0.3999755851924192)
)

# Parameter data columns — slightly darker grey than default white
FILL_PARAM_DATA = PatternFill(
    patternType="solid", fgColor=Color(theme=0, tint=-0.0499893185216834)
)

# Index dimension columns (period, time, constraint) — light green
FILL_INDEX_HEADER = PatternFill(
    patternType="solid", fgColor="C6EFCE"
)
FILL_INDEX_DATA = PatternFill(
    patternType="solid", fgColor="E2EFDA"
)


# ---------------------------------------------------------------------------
# Formatting helpers
# ---------------------------------------------------------------------------


def format_constant_sheet(
    ws: Worksheet,
    n_entity_cols: int,
    n_extra_cols: int = 0,
    has_entity_alt: bool = True,
) -> None:
    """Apply formatting to a constant-layout sheet.

    Column layout (1-based):
        1                       alternative
        2 .. 1+n_entity_cols    entity dimension columns
        next n_extra_cols       extra columns (e.g. left_node, right_node)
        remaining               parameter columns (+ optional Entity Alternative)

    Row layout:
        1   description / navigate
        2   headers
        3+  data
    """
    max_row = ws.max_row
    max_col = ws.max_column

    first_param_col = 1 + n_entity_cols + n_extra_cols + 1  # 1-based

    for col in range(1, max_col + 1):
        # --- Row 1: description row ---
        cell_r1 = ws.cell(row=1, column=col)
        if col == 1:
            # 'navigate' cell gets link font, no dark fill
            pass
        elif cell_r1.value is not None:
            cell_r1.fill = FILL_DESC_ROW
            cell_r1.font = FONT_DESC_ROW

        # --- Row 2: header row ---
        cell_r2 = ws.cell(row=2, column=col)
        if col == 1:
            cell_r2.fill = FILL_ALT_HEADER
        elif col <= 1 + n_entity_cols + n_extra_cols:
            cell_r2.fill = FILL_ENTITY_HEADER
        else:
            cell_r2.fill = FILL_PARAM_HEADER

        # --- Rows 3+: data rows ---
        for row in range(3, max_row + 1):
            cell = ws.cell(row=row, column=col)
            if col == 1:
                cell.fill = FILL_ALT_DATA
            elif col <= 1 + n_entity_cols + n_extra_cols:
                cell.fill = FILL_ENTITY_DATA
            # parameter data cells: no fill (default white)


def format_periodic_sheet(
    ws: Worksheet,
    n_entity_cols: int,
    n_extra_cols: int = 0,
) -> None:
    """Apply formatting to a periodic-layout sheet.

    Same as constant but:
    - The period / index column after entity columns gets entity colouring.
    - No Entity Alternative column.

    Column layout (1-based):
        1                       alternative
        2 .. 1+n_entity_cols    entity dimension columns
        next n_extra_cols       extra columns
        remaining               parameter columns
    """
    max_row = ws.max_row
    max_col = ws.max_column

    entity_end_col = 1 + n_entity_cols + n_extra_cols  # last entity-ish column

    for col in range(1, max_col + 1):
        # --- Row 1: description row ---
        cell_r1 = ws.cell(row=1, column=col)
        if col == 1:
            pass  # navigate link
        elif cell_r1.value is not None:
            cell_r1.fill = FILL_DESC_ROW
            cell_r1.font = FONT_DESC_ROW

        # --- Row 2: header row ---
        cell_r2 = ws.cell(row=2, column=col)
        if col == 1:
            cell_r2.fill = FILL_ALT_HEADER
        elif col <= entity_end_col:
            cell_r2.fill = FILL_ENTITY_HEADER
        else:
            cell_r2.fill = FILL_PARAM_HEADER

        # --- Rows 3+: data rows ---
        for row in range(3, max_row + 1):
            cell = ws.cell(row=row, column=col)
            if col == 1:
                cell.fill = FILL_ALT_DATA
            elif col <= entity_end_col:
                cell.fill = FILL_ENTITY_DATA
            # parameter data cells: no fill


def format_timeseries_sheet(
    ws: Worksheet,
    n_header_rows: int,
) -> None:
    """Apply formatting to a transposed timeseries-layout sheet.

    Row layout (1-based):
        1                   alternative row   (B='alternative', C+= alt values)
        2                   parameter row     (B='parameter',   C+= param names)
        3 .. n_header_rows-1  entity dim rows (B=dim name,      C+= entity names)
        n_header_rows       last entity/time  (A='time', B=dim, C+= entity names)
        n_header_rows+1 ..  data rows         (A=time values,   C+= numeric data)

    Column layout:
        A   navigate / empty / 'time' label / time values
        B   row labels ('alternative', 'parameter', entity dim names)
        C+  data columns
    """
    max_row = ws.max_row
    max_col = ws.max_column

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)

            if col == 1:
                # Column A
                if row == 1:
                    # navigate cell – handled by add_navigate_link
                    pass
                elif row < n_header_rows:
                    # empty cells above the time row
                    pass
                elif row == n_header_rows:
                    cell.fill = FILL_TIME_HEADER
                else:
                    # time data values
                    cell.fill = FILL_TIME_DATA

            elif col == 2:
                # Column B – row labels
                if row == 1:
                    cell.fill = FILL_ALT_HEADER
                elif row == 2:
                    cell.fill = FILL_PARAM_LABEL
                elif row <= n_header_rows:
                    cell.fill = FILL_ENTITY_HEADER
                # data rows in col B: no fill

            else:
                # Columns C+ – data columns
                if row == 1:
                    cell.fill = FILL_ALT_DATA
                elif row == 2:
                    cell.fill = FILL_PARAM_HEADER
                elif row <= n_header_rows:
                    cell.fill = FILL_ENTITY_DATA
                # data rows: no fill


def format_link_sheet(ws: Worksheet) -> None:
    """Apply formatting to a link-only (relationship) sheet.

    Row 1: entity headers.  Rows 2+: entity data.
    """
    max_row = ws.max_row
    max_col = ws.max_column

    for col in range(1, max_col + 1):
        ws.cell(row=1, column=col).fill = FILL_ENTITY_HEADER
        for row in range(2, max_row + 1):
            ws.cell(row=row, column=col).fill = FILL_ENTITY_DATA


def add_navigate_link(ws: Worksheet) -> None:
    """Write a 'navigate' hyperlink in cell A1."""
    cell = ws.cell(row=1, column=1)
    cell.value = "navigate"
    cell.hyperlink = "#navigate!A1"
    cell.font = FONT_NAVIGATE_LINK


# ---------------------------------------------------------------------------
# v2 format formatting helpers
# ---------------------------------------------------------------------------


def format_constant_sheet_v2(
    ws: Worksheet,
    n_entity_cols: int,
    n_extra_cols: int = 0,
    def_col: int = 3,
    index_cols: set[int] | None = None,
    last_data_col: int = 0,
) -> None:
    """Apply formatting to a v2 constant-layout sheet.

    Row layout:
        1   description / navigate (from def_col onward)
        2   data type row (from def_col onward)
        3   definition row (alternative, entity: X, parameter, param names)
        4+  data

    Column layout (1-based):
        1                       alternative
        2 .. 1+n_entity_cols    entity dimension columns
        next n_extra_cols       extra columns (direction, index, etc.)
        def_col                 definition column (description/data type/parameter)
        def_col+1 ..            parameter columns

    Args:
        index_cols: 1-based column numbers that are index dimensions
            (period, time, constraint) — these get a light green fill
            instead of entity fill.
    """
    if index_cols is None:
        index_cols = set()

    max_row = ws.max_row
    max_col = last_data_col if last_data_col > 0 else ws.max_column

    for col in range(1, max_col + 1):
        is_index = col in index_cols

        # --- Row 1: description row (from def_col onward) ---
        cell_r1 = ws.cell(row=1, column=col)
        if col == 1:
            pass  # navigate link
        elif col == def_col and cell_r1.value is not None:
            # Label cell ("description") — dark
            cell_r1.fill = FILL_DESC_ROW
            cell_r1.font = FONT_DESC_ROW
        elif col > def_col and cell_r1.value is not None:
            # Data cells (actual descriptions) — lighter
            cell_r1.fill = FILL_DESC_DATA
            cell_r1.font = FONT_DESC_DATA

        # --- Row 2: data type row (from def_col onward) ---
        cell_r2 = ws.cell(row=2, column=col)
        if col == def_col and cell_r2.value is not None:
            # Label cell ("data type") — dark
            cell_r2.fill = FILL_DESC_ROW
            cell_r2.font = FONT_DESC_ROW
        elif col > def_col and cell_r2.value is not None:
            # Data cells (actual types) — lighter
            cell_r2.fill = FILL_DESC_DATA
            cell_r2.font = FONT_DESC_DATA

        # --- Row 3: definition row ---
        cell_r3 = ws.cell(row=3, column=col)
        if col == 1:
            cell_r3.fill = FILL_ALT_HEADER
        elif is_index:
            cell_r3.fill = FILL_INDEX_HEADER
        elif col < def_col:
            cell_r3.fill = FILL_ENTITY_HEADER
        elif col == def_col:
            cell_r3.fill = FILL_PARAM_LABEL
        else:
            cell_r3.fill = FILL_PARAM_HEADER

    # --- Rows 4+: data row fills via conditional formatting (range-based, fast) ---
    # Extend 100 rows beyond data so new user entries get the right colors
    if max_row >= 4:
        from openpyxl.formatting.rule import CellIsRule

        _always = ['"§§§NEVER§§§"']
        data_range_end = max_row + 100

        for col in range(1, (last_data_col if last_data_col > 0 else max_col) + 1):
            is_index = col in index_cols
            cl = get_column_letter(col)
            rng = f"{cl}4:{cl}{data_range_end}"

            if col == 1:
                fill = FILL_ALT_DATA
            elif is_index:
                fill = FILL_INDEX_DATA
            elif col == def_col:
                fill = FILL_DEF_COL
            elif col < def_col:
                fill = FILL_ENTITY_DATA
            elif col > def_col:
                fill = FILL_PARAM_DATA
            else:
                continue

            ws.conditional_formatting.add(
                rng, CellIsRule(operator="notEqual", formula=_always, fill=fill)
            )


def format_periodic_sheet_v2(
    ws: Worksheet,
    n_entity_cols: int,
    n_extra_cols: int = 0,
    def_col: int = 3,
    index_cols: set[int] | None = None,
    last_data_col: int = 0,
) -> None:
    """Apply formatting to a v2 periodic-layout sheet.

    Same structure as format_constant_sheet_v2 but for periodic sheets.
    """
    format_constant_sheet_v2(ws, n_entity_cols, n_extra_cols, def_col, index_cols, last_data_col)


def format_timeseries_sheet_v2(
    ws: Worksheet,
    n_header_rows: int,
    single_param: bool = False,
    row_types: dict[int, str] | None = None,
    last_data_col: int = 0,
) -> None:
    """Apply formatting to a v2 transposed timeseries-layout sheet.

    Formatting is based on the *row_types* dict which maps 1-based row
    numbers to their semantic type: 'description', 'data_type', 'alternative',
    'parameter', 'entity', 'param_info' (grey single-param row).

    If *row_types* is not provided, it is inferred from *single_param*.
    """
    max_row = ws.max_row
    max_col = ws.max_column

    if row_types is None:
        row_types = {}
        if single_param:
            row_types = {1: "param_info", 2: "alternative", 3: "entity"}
        else:
            row_types = {1: "alternative", 2: "parameter", 3: "entity"}

    # Map row type to (label_fill, data_fill) for col 2 and cols 3+
    type_fills = {
        "description": (FILL_DESC_ROW, FILL_DESC_DATA, FONT_DESC_ROW, FONT_DESC_DATA),
        "data_type":   (FILL_DESC_ROW, FILL_DESC_DATA, FONT_DESC_ROW, FONT_DESC_DATA),
        "alternative": (FILL_ALT_HEADER, FILL_ALT_DATA, None, None),
        "parameter":   (FILL_PARAM_LABEL, FILL_PARAM_HEADER, None, None),
        "entity":      (FILL_ENTITY_HEADER, FILL_ENTITY_DATA, None, None),
        "param_info":  (FILL_DEF_COL, FILL_DEF_COL, None, None),
    }

    # Format header rows (per-cell — only a few rows, only data columns)
    header_col_end = last_data_col if last_data_col > 0 else max_col
    for row in range(1, n_header_rows + 1):
        rtype = row_types.get(row)

        for col in range(1, header_col_end + 1):
            cell = ws.cell(row=row, column=col)

            if col == 1:
                if row == n_header_rows:
                    cell.fill = FILL_INDEX_HEADER
                # else: navigate link or empty — no fill

            elif rtype in type_fills:
                label_fill, data_fill, label_font, data_font = type_fills[rtype]
                if col == 2:
                    cell.fill = label_fill
                    if label_font:
                        cell.font = label_font
                else:
                    cell.fill = data_fill
                    if data_font:
                        cell.font = data_font

    # Note: Column A data fill is NOT applied here — each writer adds its own
    # (FILL_TIME_DATA for timeseries, FILL_INDEX_DATA for array sheets).


def format_link_sheet_v2(ws: Worksheet) -> None:
    """Apply formatting to a v2 link-only (relationship) sheet.

    Row 1: navigate
    Row 2: entity definition row (entity: X, Y | dim1 | dim2)
    Rows 3+: entity data
    """
    max_row = ws.max_row
    max_col = ws.max_column

    for col in range(1, max_col + 1):
        # Row 2: entity header row
        ws.cell(row=2, column=col).fill = FILL_ENTITY_HEADER
        # Rows 3+: entity data
        for row in range(3, max_row + 1):
            ws.cell(row=row, column=col).fill = FILL_ENTITY_DATA


def auto_column_width(
    ws: Worksheet,
    min_param_width: float = 10,
    non_param_width: float = 22,
    def_col_width: float = 11,
    index_col_width: float = 12,
    max_width: float = 40,
    header_row: int = 0,
    def_col: int = 0,
    index_cols: set[int] | None = None,
    last_data_col: int = 0,
) -> None:
    """Size columns based on their role and header content.

    Args:
        min_param_width: Minimum width for parameter columns (right of def_col).
        non_param_width: Width for structural columns (alternative, entity).
        def_col_width: Width for the definition column.
        index_col_width: Width for index columns (period, time, constraint).
        max_width: Maximum column width cap.
        header_row: Row containing parameter names for sizing.  When 0,
            scan all rows (for navigate, scenario, etc.).
        def_col: 1-based column of the definition column (the boundary).
            When 0, all columns use content-based sizing.
        index_cols: 1-based column numbers that are index dimensions.
        last_data_col: Last column of the data area (1-based).  Columns
            beyond this are not sized (e.g. convenience/reference sections
            set their own widths).  When 0, size all columns.
    """
    if index_cols is None:
        index_cols = set()

    end_col = last_data_col if last_data_col > 0 else (ws.max_column or 0)
    for col_idx in range(1, end_col + 1):
        if def_col > 0 and col_idx in index_cols:
            width = index_col_width
        elif def_col > 0 and col_idx < def_col:
            width = non_param_width
        elif def_col > 0 and col_idx == def_col:
            width = def_col_width
        elif def_col > 0 and header_row > 0:
            # Parameter column — size to header row text
            header_val = ws.cell(row=header_row, column=col_idx).value
            if header_val is not None:
                header_len = len(str(header_val)) + 1
            else:
                header_len = 0
            width = max(min_param_width, min(header_len, max_width))
        else:
            # No def_col (navigate, scenario, etc.) — scan all rows
            longest = non_param_width
            for row_idx in range(1, min((ws.max_row or 0) + 1, 200)):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is not None:
                    length = len(str(value)) + 1
                    if length > longest:
                        longest = length
            width = min(longest, max_width)

        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width
