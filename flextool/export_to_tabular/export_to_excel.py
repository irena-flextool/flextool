"""Top-level orchestrator for exporting a FlexTool Spine DB to an Excel file."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook

from flextool.export_to_tabular.db_reader import DatabaseContents, read_database
from flextool.export_to_tabular.sheet_config import SheetSpec, build_sheet_specs, load_settings
from flextool.export_to_tabular.excel_writer import (
    IndexClassifier,
    write_array_transposed_sheet_v2,
    write_constant_sheet,
    write_constant_sheet_v2,
    write_link_sheet,
    write_link_sheet_v2,
    write_navigate_sheet,
    write_nested_periodic_sheet,
    write_nested_periodic_sheet_v2,
    write_periodic_sheet,
    write_periodic_sheet_v2,
    write_scenario_sheet,
    write_timeseries_sheet,
    write_timeseries_sheet_v2,
    write_version_sheet,
)
import flextool.export_to_tabular.excel_writer as _ew


def _build_tab_color_map(navigate_groups: list[dict[str, Any]]) -> dict[str, str]:
    """Build a mapping from sheet name to tab color (hex RGB string) from navigate groups."""
    color_map: dict[str, str] = {}
    for group in navigate_groups:
        color = group["color"]
        for row in group.get("rows", []):
            for sheet_name in row:
                color_map[sheet_name] = color
    return color_map


def export_to_excel(
    db_url: str,
    output_path: str,
    include_advanced: bool = False,
    use_new_format: bool = True,
    include_stochastics: bool = False,  # deprecated alias for include_advanced
) -> None:
    """Export a FlexTool Spine DB to an Excel (.xlsx) file.

    Args:
        db_url: Spine DB URL (e.g., 'sqlite:///path/to/db.sqlite')
        output_path: Output .xlsx file path
        include_advanced: Whether to include advanced sheets (solve sequences,
            periods_available, stochastic data).
        use_new_format: If True, use the v2 self-describing format with embedded
            metadata.  If False, use the original v1 format.
        include_stochastics: Deprecated alias for include_advanced.
    """
    if include_stochastics:
        include_advanced = True

    # 1. Read all data from the database
    print(f"Reading database: {db_url}")
    db_contents: DatabaseContents = read_database(db_url)

    # 1b. Build index classifier and entity alternatives index
    _ew._index_classifier = IndexClassifier(db_contents)
    _ew._build_entity_alts_index(db_contents)

    # 2. Load settings and build ordered sheet specifications
    print("Building sheet specifications...")
    settings = load_settings()

    # 2b. Load data type overrides and width settings
    _ew._data_type_overrides = settings.get("data_type_overrides", {})
    _ew._min_param_width = settings.get("minimum_parameter_column_width", 10)
    _ew._non_param_width = settings.get("width_non_parameter_columns", 22)
    _ew._def_col_width = settings.get("definition_column_width", 11)
    _ew._index_col_width = settings.get("index_column_width", 12)
    _ew._period_only_params = settings.get("period_only_params", {})
    _ew._time_structure_classes = set(settings.get("time_structure_classes", []))
    specs: list[SheetSpec] = build_sheet_specs(db_contents, settings)
    navigate_groups: list[dict[str, Any]] = settings.get("navigate_groups", [])
    tab_color_map = _build_tab_color_map(navigate_groups) if navigate_groups else {}

    # Filter out advanced sheets unless requested OR they have data
    advanced_sheets: set[str] = set(settings.get("advanced_sheets", []))
    if not include_advanced:
        # Pre-build a set of (class, param) pairs that have data for fast lookup
        _data_keys: set[tuple[str, str]] = set()
        for (cls, _bn, pn, _alt) in db_contents.parameter_values:
            _data_keys.add((cls, pn))

        def _has_data(spec: SheetSpec) -> bool:
            for cls in spec.entity_classes:
                for pn in spec.parameter_names:
                    if (cls, pn) in _data_keys:
                        return True
            return False

        specs = [s for s in specs
                 if s.sheet_name not in advanced_sheets or _has_data(s)]

    # Load info rows setting
    _ew._info_rows = settings.get("info_rows", {})

    # 3. Create the workbook
    wb = Workbook()

    # 4. Delete the default 'Sheet' worksheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Track counts by layout for the summary
    layout_counts: dict[str, int] = {}

    # 5. Write each sheet
    for spec in specs:
        ws = wb.create_sheet(spec.sheet_name)
        layout = spec.layout

        if layout == "constant":
            if use_new_format:
                # Check if this is a pure-array sheet (model_solve_sequence etc.)
                _is_array_sheet = (
                    len(spec.parameter_names) == 1
                    and _ew._get_param_data_type(
                        spec.parameter_names[0], spec.entity_classes,
                        db_contents, layout="constant",
                    ) == "boolean (array)"
                )
                if _is_array_sheet:
                    write_array_transposed_sheet_v2(ws, spec, db_contents)
                else:
                    write_constant_sheet_v2(ws, spec, db_contents)
            else:
                write_constant_sheet(ws, spec, db_contents)
        elif layout == "periodic":
            if use_new_format:
                write_periodic_sheet_v2(ws, spec, db_contents)
            else:
                write_periodic_sheet(ws, spec, db_contents)
        elif layout == "nested_periodic":
            if use_new_format:
                write_nested_periodic_sheet_v2(ws, spec, db_contents)
            else:
                write_nested_periodic_sheet(ws, spec, db_contents)
        elif layout == "timeseries":
            if use_new_format:
                write_timeseries_sheet_v2(ws, spec, db_contents)
            else:
                write_timeseries_sheet(ws, spec, db_contents)
        elif layout == "link":
            if use_new_format:
                write_link_sheet_v2(ws, spec, db_contents)
            else:
                write_link_sheet(ws, spec, db_contents)
        elif layout == "scenario":
            write_scenario_sheet(ws, db_contents, include_stochastics=include_advanced)
        elif layout == "version":
            if use_new_format:
                # In v2, version info is on the navigate sheet — skip separate sheet
                del wb[spec.sheet_name]
                continue
            write_version_sheet(ws, db_contents.version)
        elif layout == "navigate":
            write_navigate_sheet(
                ws, specs, navigate_groups=navigate_groups,
                version=db_contents.version if use_new_format else None,
            )
        else:
            print(f"  Warning: unknown layout '{layout}' for sheet '{spec.sheet_name}', skipping.")
            continue

        # Set tab color from navigate group config
        if spec.sheet_name in tab_color_map:
            ws.sheet_properties.tabColor = tab_color_map[spec.sheet_name]

        layout_counts[layout] = layout_counts.get(layout, 0) + 1

    # 6. Ensure the output directory exists and save
    output_dir = Path(output_path).parent
    output_dir.mkdir(parents=True, exist_ok=True)

    wb.save(output_path)

    # 7. Print summary
    format_label = "v2 (self-describing)" if use_new_format else "v1 (original)"
    total = len(specs)
    print(f"\nExport complete: {output_path} [{format_label}]")
    print(f"  Total sheets: {total}")
    for layout_type, count in sorted(layout_counts.items()):
        print(f"    {layout_type}: {count}")
    if db_contents.version is not None:
        version_display = int(db_contents.version) if db_contents.version == int(db_contents.version) else db_contents.version
        print(f"  DB version: {version_display}")
