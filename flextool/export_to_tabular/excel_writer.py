"""Write FlexTool data to Excel sheets based on SheetSpec configurations.

Each ``write_*`` function populates one openpyxl worksheet according to
its layout type (constant, periodic, timeseries, link, scenario, version,
navigate).
"""

from __future__ import annotations

from typing import Any

import numpy as np
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.worksheet.worksheet import Worksheet

from spinedb_api import Map, Array

from flextool.export_to_tabular.db_reader import DatabaseContents
from openpyxl.styles import Protection

from flextool.export_to_tabular.formatting import (
    add_navigate_link,
    auto_column_width,
    format_constant_sheet,
    format_constant_sheet_v2,
    format_link_sheet,
    format_link_sheet_v2,
    format_periodic_sheet,
    format_periodic_sheet_v2,
    format_timeseries_sheet,
    format_timeseries_sheet_v2,
    FILL_DEF_COL,
    FILL_ENTITY_HEADER,
    FILL_INDEX_DATA,
    FILL_INDEX_HEADER,
    FILL_DESC_ROW,
    FILL_DESC_DATA,
    FILL_PARAM_HEADER,
    FONT_DESC_ROW,
    FONT_DESC_DATA,
    FONT_NAVIGATE_LINK,
)
from flextool.export_to_tabular.sheet_config import SheetSpec


def _to_native(value: Any) -> Any:
    """Convert numpy scalar types to native Python types for openpyxl."""
    if isinstance(value, np.floating):
        return float(value)
    if isinstance(value, np.integer):
        return int(value)
    if isinstance(value, (np.str_, np.bytes_)):
        return str(value)
    if isinstance(value, np.bool_):
        return bool(value)
    return value


def _array_to_str(arr: Array) -> str:
    """Convert an Array value to a comma-separated string."""
    return ", ".join(str(_to_native(v)) for v in arr.values)


def _is_map(value: Any) -> bool:
    """Check if a value is a Map."""
    return isinstance(value, Map)


def _is_array(value: Any) -> bool:
    """Check if a value is an Array."""
    return isinstance(value, Array)


def _is_scalar(value: Any) -> bool:
    """Check if a value is a scalar (float, int, str) — not Map or Array."""
    return isinstance(value, (int, float, str, np.floating, np.integer, np.str_))


class IndexClassifier:
    """Classifies Map index values as time-indexed or period-indexed.

    Built once from DatabaseContents by collecting known timestep names
    (from timeline entities) and period names (from solve entities).
    """

    def __init__(self, db_contents: DatabaseContents) -> None:
        self.time_indexes: set[str] = set()
        self.period_indexes: set[str] = set()

        for (cls, _byname, pname, _alt), val in db_contents.parameter_values.items():
            if not isinstance(val, Map):
                continue
            # Timeline timestep_duration maps contain all timestep names as indexes
            if cls == "timeline" and pname == "timestep_duration":
                self.time_indexes.update(str(_to_native(idx)) for idx in val.indexes)
            # Solve parameters with Map values use period names as indexes
            if cls == "solve" and pname != "solver_arguments":
                self.period_indexes.update(str(_to_native(idx)) for idx in val.indexes)

    def is_time_indexed(self, value: Any) -> bool:
        """Check if a Map value contains time-series data (should go on _t sheet).

        Classification order:
        1. Explicit index_name == "time" → True
        2. Explicit index_name == "period" → False
        3. Cross-reference indexes with known timesteps/periods from the DB
        4. Fallback: large index count (>100) suggests time-series
        """
        if not isinstance(value, Map) or len(value.indexes) == 0:
            return False
        iname = getattr(value, "index_name", "")
        if iname == "time":
            return True
        if iname == "period":
            return False

        # Cross-reference with known indexes
        map_idxs = {str(_to_native(idx)) for idx in value.indexes}
        has_time_overlap = bool(map_idxs & self.time_indexes)
        has_period_overlap = bool(map_idxs & self.period_indexes)

        if has_time_overlap and not has_period_overlap:
            return True
        if has_period_overlap and not has_time_overlap:
            return False

        # Both or neither overlap — use count as last resort
        return len(map_idxs) > 100


# Module-level classifier instance, set by export_to_excel before writing sheets
_index_classifier: IndexClassifier | None = None


def _is_time_indexed_map(value: Any) -> bool:
    """Check if a Map value contains time-series data (should go on _t sheet, not _p)."""
    if _index_classifier is not None:
        return _index_classifier.is_time_indexed(value)
    # Fallback if classifier not initialized
    if not isinstance(value, Map) or len(value.indexes) == 0:
        return False
    return getattr(value, "index_name", "") == "time"


def _get_extra_columns_for_entity(
    entity_byname: tuple,
    entity_class_name: str,
    spec: SheetSpec,
    db_contents: DatabaseContents,
) -> list[str | None]:
    """Look up extra entity column values for a given entity.

    For example, for connection_c with extra_entity_class='connection__node__node',
    finds the matching multi-dim entity and extracts the extra dimension values.

    Returns a list of values for each extra_entity_column, or Nones if not found.
    """
    if not spec.extra_entity_columns or not spec.extra_entity_class:
        return []

    extra_entities = db_contents.entities.get(spec.extra_entity_class, [])
    # The main entity name is the first element (e.g. 'ConnBat' for connection)
    main_entity_name = entity_byname[0] if entity_byname else None

    # Find the extra entity that has this main entity as its first element
    for extra_ent in extra_entities:
        extra_byname = extra_ent["entity_byname"]
        if extra_byname[0] == main_entity_name:
            # Return the extra dimensions (skip the first which is the main entity)
            extra_vals = [str(_to_native(v)) for v in extra_byname[1:]]
            # Pad or truncate to match extra_entity_columns length
            while len(extra_vals) < len(spec.extra_entity_columns):
                extra_vals.append(None)
            return extra_vals[: len(spec.extra_entity_columns)]

    return [None] * len(spec.extra_entity_columns)


# Pre-built index for fast alternative lookups, set by export_to_excel.
_entity_alts_index: dict[tuple[str, tuple], set[str]] = {}


def _build_entity_alts_index(db_contents: DatabaseContents) -> None:
    """Build a lookup from (class, byname) → set of alternatives."""
    global _entity_alts_index
    _entity_alts_index = {}
    for (cls, byname, _param, alt) in db_contents.parameter_values:
        key = (cls, byname)
        if key not in _entity_alts_index:
            _entity_alts_index[key] = set()
        _entity_alts_index[key].add(alt)
    for (cls, byname, alt) in db_contents.entity_alternatives:
        key = (cls, byname)
        if key not in _entity_alts_index:
            _entity_alts_index[key] = set()
        _entity_alts_index[key].add(alt)


def _find_alternatives_for_entity(
    entity_class: str,
    entity_byname: tuple,
    spec: SheetSpec,
    db_contents: DatabaseContents,
) -> list[str]:
    """Find all alternatives that have parameter values or entity_alternatives for an entity.

    If the entity has no data at all, returns ``[""]`` so that it still
    gets written as an empty row (preserving the entity in the export).
    """
    key = (entity_class, entity_byname)
    alts = _entity_alts_index.get(key, set())

    if not alts:
        return [""]  # empty alt → entity row with no data

    return sorted(alts)


def _get_direction(entity_class: str, spec: SheetSpec) -> str | None:
    """Get the direction value for an entity class from the spec's direction_map."""
    if spec.direction_column and spec.direction_map:
        return spec.direction_map.get(entity_class)
    return None


# ---------------------------------------------------------------------------
# Constant sheet
# ---------------------------------------------------------------------------


def write_constant_sheet(
    ws: Worksheet,
    spec: SheetSpec,
    db_contents: DatabaseContents,
) -> None:
    """Write a constant-layout sheet."""
    # Build header columns
    headers: list[str] = ["alternative"]
    headers.extend(spec.entity_columns)
    headers.extend(spec.extra_entity_columns)
    # Pre-EA params (e.g. has_balance for node_c) go before Entity Alternative
    if spec.pre_ea_params:
        headers.extend(spec.pre_ea_params)
    if spec.has_entity_alternative:
        headers.append("Entity Alternative")
    if spec.direction_column:
        headers.append(spec.direction_column)
    if spec.unpack_index_column:
        headers.append(spec.unpack_index_column)

    # Parameter columns (pre_ea_params already in headers, not repeated here)
    param_headers: list[str] = list(spec.parameter_names)

    all_headers = headers + param_headers

    # --- Row 1: descriptions ---
    ws.cell(row=1, column=1, value="navigate")
    n_fixed = len(headers)
    for col_idx, hdr in enumerate(all_headers, start=1):
        if hdr == "Entity Alternative":
            ws.cell(row=1, column=col_idx, value="Whether the entity is active in this alternative.")
        elif hdr in spec.descriptions:
            ws.cell(row=1, column=col_idx, value=spec.descriptions[hdr])

    # --- Row 2: headers ---
    for col_idx, hdr in enumerate(all_headers, start=1):
        ws.cell(row=2, column=col_idx, value=hdr)

    # --- Collect data rows ---
    data_rows: list[list[Any]] = []

    if spec.unpack_index_column:
        _collect_unpack_rows(data_rows, spec, db_contents, headers, param_headers)
    else:
        _collect_constant_rows(data_rows, spec, db_contents, headers, param_headers)

    # --- Sort rows ---
    data_rows.sort(key=lambda r: tuple(str(v) if v is not None else "" for v in r))

    # --- Write data rows ---
    for row_idx, row_data in enumerate(data_rows, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            if value is not None:
                ws.cell(row=row_idx, column=col_idx, value=value)

    # --- Formatting ---
    n_entity_cols = len(spec.entity_columns)
    n_extra = len(spec.extra_entity_columns) + len(spec.pre_ea_params)
    if spec.has_entity_alternative:
        n_extra += 1
    if spec.direction_column:
        n_extra += 1
    if spec.unpack_index_column:
        n_extra += 1

    format_constant_sheet(ws, n_entity_cols, n_extra, has_entity_alt=False)
    add_navigate_link(ws)
    auto_column_width(ws)


def _collect_constant_rows(
    data_rows: list[list[Any]],
    spec: SheetSpec,
    db_contents: DatabaseContents,
    headers: list[str],
    param_headers: list[str],
) -> None:
    """Collect rows for a regular (non-unpack) constant sheet."""
    for entity_class in spec.entity_classes:
        entities = db_contents.entities.get(entity_class, [])
        direction = _get_direction(entity_class, spec)

        for entity in entities:
            entity_byname = entity["entity_byname"]
            alts = _find_alternatives_for_entity(
                entity_class, entity_byname, spec, db_contents
            )

            for alt in alts:
                row: list[Any] = [alt]

                # Entity elements
                for elem in entity_byname:
                    row.append(str(_to_native(elem)))

                # Extra entity columns
                if spec.extra_entity_columns:
                    extras = _get_extra_columns_for_entity(
                        entity_byname, entity_class, spec, db_contents
                    )
                    row.extend(extras)

                # Pre-EA params (e.g. has_balance for node_c)
                for pname in spec.pre_ea_params:
                    key = (entity_class, entity_byname, pname, alt)
                    value = db_contents.parameter_values.get(key)
                    if value is not None and _is_scalar(value):
                        row.append(_to_native(value))
                    else:
                        row.append(None)

                # Entity Alternative (after entity cols + pre-EA, before direction)
                if spec.has_entity_alternative:
                    ea_key = (entity_class, entity_byname, alt)
                    ea_val = db_contents.entity_alternatives.get(ea_key)
                    row.append(ea_val if ea_val is not None else None)

                # Direction
                if spec.direction_column:
                    row.append(direction)

                # Parameter values — check for Arrays that need expansion
                param_values: list[Any] = []
                has_array = False
                array_param_idx: int | None = None
                array_elements: list[Any] = []

                for pi, pname in enumerate(spec.parameter_names):
                    key = (entity_class, entity_byname, pname, alt)
                    value = db_contents.parameter_values.get(key)

                    if value is None:
                        param_values.append(None)
                    elif _is_map(value):
                        # Maps go on _p or _t sheets, not constant
                        param_values.append(None)
                    elif _is_array(value):
                        has_array = True
                        array_param_idx = pi
                        array_elements = [_to_native(v) for v in value.values]
                        param_values.append(None)  # placeholder
                    elif _is_scalar(value):
                        param_values.append(_to_native(value))
                    else:
                        param_values.append(None)

                if has_array and array_elements:
                    # Expand: one row per array element
                    for elem in array_elements:
                        expanded_row = list(row)
                        expanded_params = list(param_values)
                        expanded_params[array_param_idx] = elem
                        expanded_row.extend(expanded_params)
                        data_rows.append(expanded_row)
                else:
                    row.extend(param_values)
                    data_rows.append(row)


def _collect_unpack_rows(
    data_rows: list[list[Any]],
    spec: SheetSpec,
    db_contents: DatabaseContents,
    headers: list[str],
    param_headers: list[str],
) -> None:
    """Collect rows for an unpacked-Map constant sheet (e.g. constraint coefficients)."""
    for entity_class in spec.entity_classes:
        entities = db_contents.entities.get(entity_class, [])
        direction = _get_direction(entity_class, spec)

        for entity in entities:
            entity_byname = entity["entity_byname"]
            alts = _find_alternatives_for_entity(
                entity_class, entity_byname, spec, db_contents
            )

            for alt in alts:
                # Collect all map indexes across all unpack params
                all_indexes: set[str] = set()
                param_maps: dict[str, Map] = {}

                for pname in spec.parameter_names:
                    key = (entity_class, entity_byname, pname, alt)
                    value = db_contents.parameter_values.get(key)
                    if value is not None and _is_map(value):
                        param_maps[pname] = value
                        for idx in value.indexes:
                            all_indexes.add(str(_to_native(idx)))

                if not all_indexes:
                    continue

                for idx_val in sorted(all_indexes):
                    row: list[Any] = [alt]

                    # Entity elements
                    for elem in entity_byname:
                        row.append(str(_to_native(elem)))

                    # Extra entity columns
                    if spec.extra_entity_columns:
                        extras = _get_extra_columns_for_entity(
                            entity_byname, entity_class, spec, db_contents
                        )
                        row.extend(extras)

                    # Pre-EA params
                    for pre_pname in spec.pre_ea_params:
                        pre_key = (entity_class, entity_byname, pre_pname, alt)
                        pre_val = db_contents.parameter_values.get(pre_key)
                        if pre_val is not None and _is_scalar(pre_val):
                            row.append(_to_native(pre_val))
                        else:
                            row.append(None)

                    # Entity Alternative (after entity cols + pre-EA, before direction)
                    if spec.has_entity_alternative:
                        ea_key = (entity_class, entity_byname, alt)
                        ea_val = db_contents.entity_alternatives.get(ea_key)
                        row.append(ea_val if ea_val is not None else None)

                    # Direction
                    if spec.direction_column:
                        row.append(direction)

                    # Unpack index column
                    row.append(idx_val)

                    # Parameter values
                    for pname in spec.parameter_names:
                        m = param_maps.get(pname)
                        if m is not None:
                            # Find the value at this index
                            found = False
                            for mi, mv in zip(m.indexes, m.values):
                                if str(_to_native(mi)) == idx_val:
                                    row.append(_to_native(mv))
                                    found = True
                                    break
                            if not found:
                                row.append(None)
                        else:
                            row.append(None)

                    data_rows.append(row)


# ---------------------------------------------------------------------------
# Periodic sheet
# ---------------------------------------------------------------------------


def write_periodic_sheet(
    ws: Worksheet,
    spec: SheetSpec,
    db_contents: DatabaseContents,
) -> None:
    """Write a periodic-layout sheet."""
    # Determine index column name from Map values
    index_col_name = _find_index_column_name(spec, db_contents, default="period")

    # Build header columns
    headers: list[str] = ["alternative"]
    headers.extend(spec.entity_columns)
    if spec.direction_column:
        headers.append(spec.direction_column)
    headers.append(index_col_name)

    # Parameter columns (no Entity Alternative on periodic sheets)
    all_headers = headers + list(spec.parameter_names)

    # --- Row 1: descriptions ---
    ws.cell(row=1, column=1, value="navigate")
    n_fixed = len(headers)
    for i, pname in enumerate(spec.parameter_names):
        col_idx = n_fixed + 1 + i
        if pname in spec.descriptions:
            ws.cell(row=1, column=col_idx, value=spec.descriptions[pname])

    # --- Row 2: headers ---
    for col_idx, hdr in enumerate(all_headers, start=1):
        ws.cell(row=2, column=col_idx, value=hdr)

    # --- Collect data rows ---
    data_rows: list[list[Any]] = []
    index_col_pos = len(headers) - 1  # 0-based position of index column

    for entity_class in spec.entity_classes:
        entities = db_contents.entities.get(entity_class, [])
        direction = _get_direction(entity_class, spec)

        for entity in entities:
            entity_byname = entity["entity_byname"]
            alts = _find_alternatives_for_entity(
                entity_class, entity_byname, spec, db_contents
            )

            for alt in alts:
                # Collect all index values and map/array data across params
                all_indexes: set[str] = set()
                param_maps: dict[str, Map] = {}
                param_arrays: dict[str, list[str]] = {}

                for pname in spec.parameter_names:
                    key = (entity_class, entity_byname, pname, alt)
                    value = db_contents.parameter_values.get(key)
                    if value is not None and _is_map(value):
                        # Check for nested Map (Map whose values are Maps)
                        is_nested = any(_is_map(mv) for mv in value.values)
                        if is_nested:
                            # Skip nested Maps — they go to solve_period_period
                            continue
                        # Skip time-indexed Maps — they belong on _t sheets
                        if _is_time_indexed_map(value):
                            continue
                        param_maps[pname] = value
                        for idx in value.indexes:
                            all_indexes.add(str(_to_native(idx)))
                    elif value is not None and _is_array(value):
                        # Array = boolean indicator (period names)
                        period_names = [str(_to_native(v)) for v in value.values]
                        param_arrays[pname] = period_names

                if not all_indexes:
                    continue

                for idx_val in sorted(all_indexes):
                    row: list[Any] = [alt]

                    # Entity elements
                    for elem in entity_byname:
                        row.append(str(_to_native(elem)))

                    # Direction
                    if spec.direction_column:
                        row.append(direction)

                    # Index column value
                    row.append(idx_val)

                    # Parameter values
                    for pname in spec.parameter_names:
                        m = param_maps.get(pname)
                        if m is not None:
                            found = False
                            for mi, mv in zip(m.indexes, m.values):
                                if str(_to_native(mi)) == idx_val:
                                    row.append(_to_native(mv))
                                    found = True
                                    break
                            if not found:
                                row.append(None)
                        elif pname in param_arrays:
                            # Boolean: check if current period is in the array
                            if idx_val in param_arrays[pname]:
                                row.append(True)
                            else:
                                row.append(None)
                        else:
                            row.append(None)

                    data_rows.append(row)

    # --- Sort rows ---
    data_rows.sort(key=lambda r: tuple(str(v) if v is not None else "" for v in r))

    # --- Write data rows ---
    for row_idx, row_data in enumerate(data_rows, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            if value is not None:
                ws.cell(row=row_idx, column=col_idx, value=value)

    # --- Formatting ---
    n_entity_cols = len(spec.entity_columns)
    n_extra = 0
    if spec.direction_column:
        n_extra += 1
    n_extra += 1  # index column counts as extra

    format_periodic_sheet(ws, n_entity_cols, n_extra)
    add_navigate_link(ws)
    auto_column_width(ws)


def _find_index_column_name(
    spec: SheetSpec,
    db_contents: DatabaseContents,
    default: str = "period",
) -> str:
    """Determine the index column name by scanning Map values for index_name.

    Ignores the placeholder 'x' from the DB and falls back to the spec's
    ``index_name_default`` (from export_settings.yaml) or *default*.
    """
    fallback = spec.index_name_default or default
    name_counts: dict[str, int] = {}

    for entity_class in spec.entity_classes:
        for (cls, _byname, pname, _alt), value in db_contents.parameter_values.items():
            if cls == entity_class and pname in spec.parameter_names and _is_map(value):
                iname = value.index_name
                if iname and iname != "x":
                    name_counts[iname] = name_counts.get(iname, 0) + 1

    if name_counts:
        return max(name_counts, key=name_counts.get)  # type: ignore[arg-type]
    return fallback


# ---------------------------------------------------------------------------
# Nested periodic sheet (solve_period_period)
# ---------------------------------------------------------------------------


def write_nested_periodic_sheet(
    ws: Worksheet,
    spec: SheetSpec,
    db_contents: DatabaseContents,
) -> None:
    """Write a nested-periodic-layout sheet for nested Map parameters.

    Handles parameters whose values are Map-of-Map (outer index =
    current_solve_period, inner Map = periods_included -> 'yes').
    """
    # Build header columns
    headers: list[str] = ["alternative"]
    headers.extend(spec.entity_columns)
    headers.append("current_solve_period")
    headers.append("periods_included")

    # Only include params that actually have nested Map values
    nested_params: list[str] = []
    for pname in spec.parameter_names:
        for (cls, _byname, p, _alt), value in db_contents.parameter_values.items():
            if cls in spec.entity_classes and p == pname and _is_map(value):
                if any(_is_map(mv) for mv in value.values):
                    nested_params.append(pname)
                    break

    if not nested_params:
        # No nested data — write minimal headers
        ws.cell(row=1, column=1, value="navigate")
        all_headers = headers + list(spec.parameter_names)
        for col_idx, hdr in enumerate(all_headers, start=1):
            ws.cell(row=2, column=col_idx, value=hdr)
        add_navigate_link(ws)
        auto_column_width(ws)
        return

    all_headers = headers + nested_params

    # --- Row 1: descriptions ---
    ws.cell(row=1, column=1, value="navigate")
    n_fixed = len(headers)
    for i, pname in enumerate(nested_params):
        col_idx = n_fixed + 1 + i
        if pname in spec.descriptions:
            ws.cell(row=1, column=col_idx, value=spec.descriptions[pname])

    # --- Row 2: headers ---
    for col_idx, hdr in enumerate(all_headers, start=1):
        ws.cell(row=2, column=col_idx, value=hdr)

    # --- Collect data rows ---
    data_rows: list[list[Any]] = []

    for entity_class in spec.entity_classes:
        entities = db_contents.entities.get(entity_class, [])

        for entity in entities:
            entity_byname = entity["entity_byname"]
            alts = _find_alternatives_for_entity(
                entity_class, entity_byname, spec, db_contents
            )

            for alt in alts:
                # Collect nested Map data for this entity+alt
                nested_data: dict[str, Map] = {}
                all_outer_indexes: set[str] = set()
                all_inner_indexes: dict[str, set[str]] = {}

                for pname in nested_params:
                    key = (entity_class, entity_byname, pname, alt)
                    value = db_contents.parameter_values.get(key)
                    if value is not None and _is_map(value):
                        if any(_is_map(mv) for mv in value.values):
                            nested_data[pname] = value
                            for oi in value.indexes:
                                outer_str = str(_to_native(oi))
                                all_outer_indexes.add(outer_str)

                if not nested_data:
                    continue

                # Collect all (outer, inner) combos across all params
                row_combos: set[tuple[str, str]] = set()
                for pname, outer_map in nested_data.items():
                    for oi, inner_val in zip(outer_map.indexes, outer_map.values):
                        outer_str = str(_to_native(oi))
                        if _is_map(inner_val):
                            for ii in inner_val.indexes:
                                inner_str = str(_to_native(ii))
                                row_combos.add((outer_str, inner_str))

                for outer_idx, inner_idx in sorted(row_combos):
                    row: list[Any] = [alt]

                    # Entity elements
                    for elem in entity_byname:
                        row.append(str(_to_native(elem)))

                    # current_solve_period and periods_included
                    row.append(outer_idx)
                    row.append(inner_idx)

                    # Parameter values
                    for pname in nested_params:
                        outer_map = nested_data.get(pname)
                        if outer_map is not None:
                            # Find the inner map at this outer index
                            found = False
                            for oi, inner_val in zip(outer_map.indexes, outer_map.values):
                                if str(_to_native(oi)) == outer_idx and _is_map(inner_val):
                                    # Find value at inner index
                                    for ii, iv in zip(inner_val.indexes, inner_val.values):
                                        if str(_to_native(ii)) == inner_idx:
                                            row.append(_to_native(iv))
                                            found = True
                                            break
                                    break
                            if not found:
                                row.append(None)
                        else:
                            row.append(None)

                    data_rows.append(row)

    # --- Sort rows ---
    data_rows.sort(key=lambda r: tuple(str(v) if v is not None else "" for v in r))

    # --- Write data rows ---
    for row_idx, row_data in enumerate(data_rows, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            if value is not None:
                ws.cell(row=row_idx, column=col_idx, value=value)

    # --- Formatting ---
    n_entity_cols = len(spec.entity_columns)
    n_extra = 2  # current_solve_period + periods_included

    format_periodic_sheet(ws, n_entity_cols, n_extra)
    add_navigate_link(ws)
    auto_column_width(ws)


def write_nested_periodic_sheet_v2(
    ws: Worksheet,
    spec: SheetSpec,
    db_contents: DatabaseContents,
) -> None:
    """Write a nested-periodic-layout sheet in v2 self-describing format.

    Layout (solve_period_period):
        Row 1: navigate | | | | description | data type | ...
        Row 2: | | | | data type | string/float (2d-map) | ...
        Row 3: alternative | entity: solve | index: current_solve_period | index: periods_included | parameter | param1 | ...
        Row 4+: data...
    """
    # Only include params that actually have nested Map values
    nested_params: list[str] = []
    for pname in spec.parameter_names:
        for (cls, _byname, p, _alt), value in db_contents.parameter_values.items():
            if cls in spec.entity_classes and p == pname and _is_map(value):
                if any(_is_map(mv) for mv in value.values):
                    nested_params.append(pname)
                    break

    if not nested_params:
        nested_params = list(spec.parameter_names)

    # Build left-side columns
    left_cols: list[str] = ["alternative"]
    entity_label = _build_entity_def_label(spec)
    left_cols.append(entity_label)
    left_cols.append("index: current_solve_period")
    left_cols.append("index: periods_included")

    def_col = len(left_cols) + 1  # 1-based

    # Right-side columns
    right_cols: list[str] = ["parameter"]
    right_cols.extend(nested_params)

    # --- Row 1: descriptions ---
    ws.cell(row=1, column=def_col, value="description")
    ws.cell(row=1, column=def_col).fill = FILL_DESC_ROW
    ws.cell(row=1, column=def_col).font = FONT_DESC_ROW
    for i, pname in enumerate(nested_params):
        col = def_col + 1 + i
        desc = _get_param_description(pname, spec.entity_classes, db_contents)
        if desc:
            ws.cell(row=1, column=col, value=desc)
            ws.cell(row=1, column=col).fill = FILL_DESC_DATA
            ws.cell(row=1, column=col).font = FONT_DESC_DATA

    # --- Row 2: data types ---
    ws.cell(row=2, column=def_col, value="data type")
    ws.cell(row=2, column=def_col).fill = FILL_DESC_ROW
    ws.cell(row=2, column=def_col).font = FONT_DESC_ROW
    for i, pname in enumerate(nested_params):
        col = def_col + 1 + i
        dtype = _get_param_data_type(pname, spec.entity_classes, db_contents, layout="nested_periodic")
        ws.cell(row=2, column=col, value=dtype)
        ws.cell(row=2, column=col).fill = FILL_DESC_DATA
        ws.cell(row=2, column=col).font = FONT_DESC_DATA

    # --- Row 3: definition row ---
    for col_idx, label in enumerate(left_cols, start=1):
        ws.cell(row=3, column=col_idx, value=label)
    for col_idx, label in enumerate(right_cols, start=def_col):
        ws.cell(row=3, column=col_idx, value=label)

    # --- Collect data rows ---
    data_rows: list[list[Any]] = []

    for entity_class in spec.entity_classes:
        entities = db_contents.entities.get(entity_class, [])

        for entity in entities:
            entity_byname = entity["entity_byname"]
            alts = _find_alternatives_for_entity(
                entity_class, entity_byname, spec, db_contents
            )

            for alt in alts:
                nested_data: dict[str, Map] = {}
                for pname in nested_params:
                    key = (entity_class, entity_byname, pname, alt)
                    value = db_contents.parameter_values.get(key)
                    if value is not None and _is_map(value):
                        if any(_is_map(mv) for mv in value.values):
                            nested_data[pname] = value

                if not nested_data:
                    continue

                row_combos: set[tuple[str, str]] = set()
                for pname, outer_map in nested_data.items():
                    for oi, inner_val in zip(outer_map.indexes, outer_map.values):
                        outer_str = str(_to_native(oi))
                        if _is_map(inner_val):
                            for ii in inner_val.indexes:
                                inner_str = str(_to_native(ii))
                                row_combos.add((outer_str, inner_str))

                for outer_idx, inner_idx in sorted(row_combos):
                    row: list[Any] = [alt]
                    for elem in entity_byname:
                        row.append(str(_to_native(elem)))
                    row.append(outer_idx)
                    row.append(inner_idx)
                    row.append(None)  # def column

                    for pname in nested_params:
                        outer_map = nested_data.get(pname)
                        if outer_map is not None:
                            found = False
                            for oi, inner_val in zip(outer_map.indexes, outer_map.values):
                                if str(_to_native(oi)) == outer_idx and _is_map(inner_val):
                                    for ii, iv in zip(inner_val.indexes, inner_val.values):
                                        if str(_to_native(ii)) == inner_idx:
                                            row.append(_to_native(iv))
                                            found = True
                                            break
                                    break
                            if not found:
                                row.append(None)
                        else:
                            row.append(None)

                    data_rows.append(row)

    data_rows.sort(key=lambda r: tuple(str(v) if v is not None else "" for v in r))

    # --- Write data rows ---
    for row_idx, row_data in enumerate(data_rows, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            if value is not None:
                ws.cell(row=row_idx, column=col_idx, value=value)

    # --- Reference section ---
    last_data_col = def_col + len(right_cols) - 1
    _write_param_reference(
        ws, last_data_col + 1, spec, db_contents, header_row=3,
        n_data_rows=len(data_rows), layout="nested_periodic",
        shown_params=nested_params,
    )

    # --- Formatting ---
    n_entity_cols = len(spec.entity_columns)
    index_col_positions = {3, 4}  # current_solve_period and periods_included
    last_data_col_nested = def_col + len(right_cols) - 1
    format_constant_sheet_v2(ws, n_entity_cols, 2, def_col, index_col_positions, last_data_col_nested)
    add_navigate_link(ws)
    auto_column_width(ws, min_param_width=_min_param_width,
                      non_param_width=_non_param_width,
                      def_col_width=_def_col_width,
                      index_col_width=_index_col_width,
                      header_row=3, def_col=def_col,
                      index_cols=index_col_positions)


# ---------------------------------------------------------------------------
# Timeseries sheet
# ---------------------------------------------------------------------------


def write_timeseries_sheet(
    ws: Worksheet,
    spec: SheetSpec,
    db_contents: DatabaseContents,
) -> None:
    """Write a transposed timeseries-layout sheet."""
    # Collect all data columns: each is (entity_class, entity_byname, param, alt, Map)
    columns: list[tuple[str, tuple, str, str, Map]] = []
    all_time_indexes: set[str] = set()

    for entity_class in spec.entity_classes:
        entities = db_contents.entities.get(entity_class, [])

        for entity in entities:
            entity_byname = entity["entity_byname"]

            for pname in spec.parameter_names:
                for alt in db_contents.alternatives:
                    key = (entity_class, entity_byname, pname, alt)
                    value = db_contents.parameter_values.get(key)
                    if value is not None and _is_map(value):
                        # Only include time-indexed Maps on timeseries sheets
                        if not _is_time_indexed_map(value):
                            continue
                        columns.append(
                            (entity_class, entity_byname, pname, alt, value)
                        )
                        for idx in value.indexes:
                            all_time_indexes.add(str(_to_native(idx)))

    if not columns:
        # Write minimal headers even if no data
        ws.cell(row=1, column=1, value="navigate")
        ws.cell(row=1, column=2, value="alternative")
        ws.cell(row=2, column=2, value="parameter")
        add_navigate_link(ws)
        auto_column_width(ws)
        return

    # Sort columns by (alt, entity_byname, param)
    columns.sort(key=lambda c: (c[3], c[1], c[2]))

    # Sort time indexes
    sorted_times = sorted(all_time_indexes)

    # Determine header structure
    n_entity_dims = len(spec.entity_columns)
    has_direction = spec.direction_column is not None

    # n_header_rows = 2 (alt, param) + n_entity_dims + (1 if direction)
    n_header_rows = 2 + n_entity_dims
    if has_direction:
        n_header_rows += 1

    # --- Write header rows ---
    # Row 1: 'navigate' | 'alternative' | alt values...
    ws.cell(row=1, column=1, value="navigate")
    ws.cell(row=1, column=2, value="alternative")
    for col_idx, (ec, byname, pname, alt, _m) in enumerate(columns, start=3):
        ws.cell(row=1, column=col_idx, value=alt)

    # Row 2: '' | 'parameter' | param names...
    ws.cell(row=2, column=2, value="parameter")
    for col_idx, (ec, byname, pname, alt, _m) in enumerate(columns, start=3):
        ws.cell(row=2, column=col_idx, value=pname)

    # Entity dimension rows
    for dim_idx in range(n_entity_dims):
        row = 3 + dim_idx
        dim_label = spec.entity_columns[dim_idx]
        ws.cell(row=row, column=2, value=dim_label)
        for col_idx, (ec, byname, pname, alt, _m) in enumerate(columns, start=3):
            if dim_idx < len(byname):
                ws.cell(row=row, column=col_idx, value=str(_to_native(byname[dim_idx])))

    # Direction row (if applicable)
    if has_direction:
        dir_row = 2 + n_entity_dims + 1
        ws.cell(row=dir_row, column=2, value=spec.direction_column)
        for col_idx, (ec, byname, pname, alt, _m) in enumerate(columns, start=3):
            direction = _get_direction(ec, spec)
            if direction:
                ws.cell(row=dir_row, column=col_idx, value=direction)

    # 'time' label on the last header row, column A
    ws.cell(row=n_header_rows, column=1, value="time")

    # --- Write time data rows ---
    # Build index lookup for each column's Map for fast access
    col_index_maps: list[dict[str, Any]] = []
    for ec, byname, pname, alt, m in columns:
        idx_map: dict[str, Any] = {}
        for mi, mv in zip(m.indexes, m.values):
            idx_map[str(_to_native(mi))] = _to_native(mv)
        col_index_maps.append(idx_map)

    for time_row_idx, time_val in enumerate(sorted_times):
        row = n_header_rows + 1 + time_row_idx
        ws.cell(row=row, column=1, value=time_val)
        for col_idx, idx_map in enumerate(col_index_maps):
            value = idx_map.get(time_val)
            if value is not None:
                ws.cell(row=row, column=col_idx + 3, value=value)

    # --- Formatting ---
    format_timeseries_sheet(ws, n_header_rows)
    add_navigate_link(ws)
    auto_column_width(ws)


# ---------------------------------------------------------------------------
# Link sheet
# ---------------------------------------------------------------------------


def write_link_sheet(
    ws: Worksheet,
    spec: SheetSpec,
    db_contents: DatabaseContents,
) -> None:
    """Write a link-only (relationship) sheet."""
    # Row 1: headers = entity dimension names
    for col_idx, col_name in enumerate(spec.entity_columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Collect and sort entity rows
    entity_class = spec.entity_classes[0] if spec.entity_classes else None
    if not entity_class:
        format_link_sheet(ws)
        auto_column_width(ws)
        return

    entities = db_contents.entities.get(entity_class, [])
    rows: list[tuple] = []
    for entity in entities:
        byname = entity["entity_byname"]
        rows.append(tuple(str(_to_native(v)) for v in byname))

    rows.sort()

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    format_link_sheet(ws)
    auto_column_width(ws)


# ---------------------------------------------------------------------------
# Scenario sheet
# ---------------------------------------------------------------------------


def write_scenario_sheet(
    ws: Worksheet,
    db_contents: DatabaseContents,
    include_stochastics: bool = False,
) -> None:
    """Write the scenario sheet with formatting.

    Layout:
        Row 1: navigate | Scenario names
        Row 2: index:   | scenario_1 | scenario_2 | ...
        Row 3: base_alternative | alt | alt | ...
        Row 4-33: alternative_1..30 | ...
    """
    scenarios = db_contents.scenarios
    font_dark = Font(color=Color(theme=1, tint=0.0))

    # Row 1: 'navigate' | 'Scenario names'
    cell_b1 = ws.cell(row=1, column=2, value="Scenario names")
    cell_b1.font = font_dark

    # Row 2: 'index:' | scenario names
    ws.cell(row=2, column=1, value="index:")
    for col_idx, sc in enumerate(scenarios, start=2):
        cell = ws.cell(row=2, column=col_idx, value=sc["name"])
        cell.fill = FILL_ENTITY_HEADER
        cell.font = font_dark

    # Row 3: 'base_alternative' | alt names (rank 0)
    cell_a3 = ws.cell(row=3, column=1, value="base_alternative")
    cell_a3.fill = FILL_PARAM_HEADER
    cell_a3.font = font_dark
    for col_idx, sc in enumerate(scenarios, start=2):
        alts = sc.get("alternatives", [])
        if alts:
            cell = ws.cell(row=3, column=col_idx, value=alts[0][0])
            cell.font = font_dark

    # Rows 4-33: alternative_1..30 — autofill to 30
    max_alts = max((len(sc.get("alternatives", [])) for sc in scenarios), default=0)
    total_alt_rows = max(30, max_alts)

    for alt_idx in range(1, total_alt_rows + 1):
        row = 3 + alt_idx
        cell_label = ws.cell(row=row, column=1, value=f"alternative_{alt_idx}")
        cell_label.fill = FILL_PARAM_HEADER
        cell_label.font = font_dark
        for col_idx, sc in enumerate(scenarios, start=2):
            alts = sc.get("alternatives", [])
            if alt_idx < len(alts):
                cell = ws.cell(row=row, column=col_idx, value=alts[alt_idx][0])
                cell.font = font_dark

    add_navigate_link(ws)
    auto_column_width(ws)


# ---------------------------------------------------------------------------
# Version sheet
# ---------------------------------------------------------------------------


def write_version_sheet(
    ws: Worksheet,
    version: float | None,
) -> None:
    """Write the version sheet."""
    if version is not None:
        version_int = int(version) if version == int(version) else version
        ws.cell(row=1, column=1, value=f"Generated from FlexTool sqlite version: {version_int}")
    else:
        ws.cell(row=1, column=1, value="Generated from FlexTool sqlite")


# ---------------------------------------------------------------------------
# Array-transposed sheet (model_solve_sequence, model_periods_available)
# ---------------------------------------------------------------------------


def write_array_transposed_sheet_v2(
    ws: Worksheet,
    spec: SheetSpec,
    db_contents: DatabaseContents,
) -> None:
    """Write an array-parameter sheet in transposed layout.

    Like profile_t but for array parameters: columns are (entity, alt)
    combinations, rows are array index positions.

    Layout:
        Row 1: navigate | parameter: solves | data type: boolean (array) | description: ...
        Row 2: INFO: ... (optional, from _info_rows setting)
        Row N:          | entity: model | flexTool  | flexTool  | ...
        Row N+1:        | alternative   | init      | 5weeks    | ...
        Row N+2: index: | 0             | solve_a   | solve_b   | ...
        ...autofill to 30
    """
    if not spec.parameter_names:
        add_navigate_link(ws)
        return

    pname = spec.parameter_names[0]
    entity_label = _build_entity_def_label(spec)
    dtype = _get_param_data_type(pname, spec.entity_classes, db_contents, layout="constant")
    desc = _get_param_description(pname, spec.entity_classes, db_contents)

    # Collect data: (entity_byname, alt) → array values
    columns: list[tuple[str, tuple, list[Any]]] = []  # (alt, entity_byname, values)
    max_array_len = 0

    for entity_class in spec.entity_classes:
        for entity in db_contents.entities.get(entity_class, []):
            entity_byname = entity["entity_byname"]
            for alt in db_contents.alternatives:
                key = (entity_class, entity_byname, pname, alt)
                value = db_contents.parameter_values.get(key)
                if value is not None and _is_array(value):
                    arr = [_to_native(v) for v in value.values]
                    columns.append((alt, entity_byname, arr))
                    if len(arr) > max_array_len:
                        max_array_len = len(arr)
                elif value is not None and _is_scalar(value):
                    # String value = single-element array
                    columns.append((alt, entity_byname, [_to_native(value)]))
                    if 1 > max_array_len:
                        max_array_len = 1

    columns.sort(key=lambda c: (c[0], c[1]))

    cur_row = 1

    # INFO row (starts from definition column B, not A)
    info_text = _info_rows.get(spec.sheet_name)
    if not info_text:
        info_text = "INFO: Add new data columns as needed by right clicking on the column name and then selecting 'Insert columns before'."
    if info_text:
        ws.cell(row=cur_row, column=2, value=info_text)
        cur_row += 1

    # Triplet row (parameter info)
    triplet = f"parameter: {pname} | data type: {dtype}"
    if desc:
        triplet += f" | description: {desc}"
    ws.cell(row=cur_row, column=2, value=triplet)
    cur_row += 1

    # Entity row
    ws.cell(row=cur_row, column=2, value=entity_label)
    for col_idx, (alt, byname, _arr) in enumerate(columns, start=3):
        ws.cell(row=cur_row, column=col_idx, value=str(_to_native(byname[0])))
    cur_row += 1

    # Alternative row — this is the last header row; put index: label in A
    ws.cell(row=cur_row, column=1, value="index:")
    ws.cell(row=cur_row, column=2, value="alternative")
    for col_idx, (alt, byname, _arr) in enumerate(columns, start=3):
        ws.cell(row=cur_row, column=col_idx, value=alt)
    cur_row += 1

    # Data rows — index numbers in column A, values in columns C+
    n_index_rows = max(30, max_array_len)
    index_start_row = cur_row

    for idx in range(n_index_rows):
        row = index_start_row + idx
        ws.cell(row=row, column=1, value=idx)
        for col_idx, (alt, byname, arr) in enumerate(columns, start=3):
            if idx < len(arr):
                ws.cell(row=row, column=col_idx, value=arr[idx])

    # Formatting — build row_types based on actual row positions
    n_header_rows = index_start_row - 1
    row_types: dict[int, str] = {}
    r = 1
    if info_text:
        # INFO row has no special formatting type (plain text)
        r += 1
    row_types[r] = "param_info"  # triplet row
    r += 1
    row_types[r] = "entity"
    r += 1
    row_types[r] = "alternative"

    arr_last_data_col = max(len(columns) + 2, 3)
    format_timeseries_sheet_v2(ws, n_header_rows, single_param=True,
                               row_types=row_types, last_data_col=arr_last_data_col)

    # Column A: green index fill (header + data + 100 extra rows)
    from openpyxl.formatting.rule import CellIsRule
    extend_end = index_start_row + n_index_rows + 100
    ws.cell(row=n_header_rows, column=1).fill = FILL_INDEX_HEADER
    ws.conditional_formatting.add(
        f"A{index_start_row}:A{extend_end}",
        CellIsRule(operator="notEqual", formula=['"§§§NEVER§§§"'],
                   fill=FILL_INDEX_DATA),
    )
    # Column B: grey definition column fill for data rows
    ws.conditional_formatting.add(
        f"B{index_start_row}:B{extend_end}",
        CellIsRule(operator="notEqual", formula=['"§§§NEVER§§§"'],
                   fill=FILL_DEF_COL),
    )

    add_navigate_link(ws)  # always A1
    auto_column_width(ws)


# ---------------------------------------------------------------------------
# Stochastic sheet (_s)
# ---------------------------------------------------------------------------


def _flatten_nested_map(
    value: Any,
    depth: int,
    prefix: tuple = (),
) -> list[tuple[tuple[str, ...], Any]]:
    """Flatten a nested Map into (index_tuple, scalar_value) pairs.

    For a 3d Map (Map of Map of Map), returns tuples of length 3.
    For a 4d Map, length 4, etc.
    """
    if depth <= 1 or not _is_map(value):
        if _is_map(value):
            # Deepest level — expand to individual entries
            return [(prefix + (str(_to_native(idx)),), _to_native(val))
                    for idx, val in zip(value.indexes, value.values)]
        else:
            return [(prefix, _to_native(value))]

    result: list[tuple[tuple[str, ...], Any]] = []
    for idx, inner in zip(value.indexes, value.values):
        idx_str = str(_to_native(idx))
        result.extend(_flatten_nested_map(inner, depth - 1, prefix + (idx_str,)))
    return result


def _get_nested_index_names(value: Any, depth: int) -> list[str]:
    """Extract index names from each level of a nested Map.

    Falls back to stochastic_index_names defaults from settings when
    the DB stores 'x' or empty index names.
    """
    names: list[str] = []
    current = value
    for level in range(depth):
        if _is_map(current) and hasattr(current, "index_name"):
            iname = current.index_name
            if iname and iname != "x":
                names.append(iname)
            elif level < len(_stochastic_index_names):
                names.append(_stochastic_index_names[level])
            else:
                names.append(f"index_{level + 1}")
            if current.values:
                current = current.values[0]
            else:
                break
        else:
            break
    return names


def write_stochastic_sheet_v2(
    ws: Worksheet,
    spec: SheetSpec,
    db_contents: DatabaseContents,
) -> None:
    """Write a stochastic (_s) sheet in the v2 self-describing format.

    Layout (e.g. profile_s_profile with 3d map: branch × analysis_time × time):
        Row 1: navigate    |                    |             | parameter: profile | data type: ... | ...
        Row 2:             |                    |             | entity: profile    | wind1          | wind1
        Row 3: index: branch | index: analysis_time | index: time | alternative  | 1week_rolling  | 2day
        Row 4+: realized   | t0001              | t0001       |                    | 0.5            | 0.3
    """
    if not spec.parameter_names:
        add_navigate_link(ws)
        return

    pname = spec.parameter_names[0]
    entity_label = _build_entity_def_label(spec)
    dtype = _get_param_data_type(pname, spec.entity_classes, db_contents, layout="timeseries")
    desc = _get_param_description(pname, spec.entity_classes, db_contents)

    # Collect data: flatten nested maps and build columns
    columns: list[tuple[str, tuple, str, list[tuple[tuple[str, ...], Any]]]] = []
    # (alt, entity_byname, entity_class, flat_data)
    max_depth = 0
    index_names: list[str] = []

    for entity_class in spec.entity_classes:
        for entity in db_contents.entities.get(entity_class, []):
            entity_byname = entity["entity_byname"]
            for alt in db_contents.alternatives:
                key = (entity_class, entity_byname, pname, alt)
                value = db_contents.parameter_values.get(key)
                if value is None or not _is_map(value):
                    continue
                # Determine nesting depth
                depth = 1
                current = value
                while _is_map(current) and current.values and _is_map(current.values[0]):
                    depth += 1
                    current = current.values[0]
                if depth < 2:
                    continue  # not stochastic (< 3d)

                if not index_names:
                    index_names = _get_nested_index_names(value, depth)
                max_depth = max(max_depth, depth)

                flat = _flatten_nested_map(value, depth)
                columns.append((alt, entity_byname, entity_class, flat))

    columns.sort(key=lambda c: (c[0], c[1]))

    # Determine expected depth from parameter type_list even if no data
    if not max_depth:
        for entity_class in spec.entity_classes:
            for pdef in db_contents.parameter_definitions.get(entity_class, []):
                if pdef["name"] == pname:
                    tl = pdef.get("parameter_type_list") or ()
                    if "4d_map" in tl:
                        max_depth = 4
                    elif "3d_map" in tl:
                        max_depth = 3
                    break
        if not max_depth:
            max_depth = 3  # default for stochastic

    n_index_cols = max_depth
    def_col = n_index_cols + 1  # 1-based

    # Pad index_names to match depth using stochastic defaults
    while len(index_names) < n_index_cols:
        level = len(index_names)
        if level < len(_stochastic_index_names):
            index_names.append(_stochastic_index_names[level])
        else:
            index_names.append(f"index_{level + 1}")

    # Collect unique index tuples preserving original Map order
    seen_tuples: set[tuple[str, ...]] = set()
    ordered_tuples: list[tuple[str, ...]] = []
    for _alt, _bn, _ec, flat in columns:
        for idx_tuple, _val in flat:
            if idx_tuple not in seen_tuples:
                seen_tuples.add(idx_tuple)
                ordered_tuples.append(idx_tuple)

    # Build column data lookups (index_tuple → value)
    col_lookups: list[dict[tuple[str, ...], Any]] = []
    for _alt, _bn, _ec, flat in columns:
        lookup = {idx_tuple: val for idx_tuple, val in flat}
        col_lookups.append(lookup)

    # ── Write header rows ────────────────────────────────────────
    cur_row = 1

    # Row 1: navigate + triplet in def_col
    triplet = f"parameter: {pname} | data type: {dtype}"
    if desc:
        triplet += f" | description: {desc}"
    ws.cell(row=cur_row, column=def_col, value=triplet)
    row_types: dict[int, str] = {cur_row: "param_info"}
    cur_row += 1

    # Row 2: entity
    ws.cell(row=cur_row, column=def_col, value=entity_label)
    for col_idx, (alt, byname, ec, _flat) in enumerate(columns, start=def_col + 1):
        if byname:
            ws.cell(row=cur_row, column=col_idx, value=str(_to_native(byname[0])))
    row_types[cur_row] = "entity"
    cur_row += 1

    # Row 3: index labels + alternative
    for i in range(n_index_cols):
        iname = index_names[i] if i < len(index_names) and index_names[i] else f"index_{i+1}"
        ws.cell(row=cur_row, column=i + 1, value=f"index: {iname}")
    ws.cell(row=cur_row, column=def_col, value="alternative")
    for col_idx, (alt, byname, ec, _flat) in enumerate(columns, start=def_col + 1):
        ws.cell(row=cur_row, column=col_idx, value=alt)
    row_types[cur_row] = "alternative"
    n_header_rows = cur_row
    cur_row += 1

    # ── Write data rows ──────────────────────────────────────────
    for idx_tuple in ordered_tuples:
        for i, idx_val in enumerate(idx_tuple):
            ws.cell(row=cur_row, column=i + 1, value=idx_val)
        for col_idx, lookup in enumerate(col_lookups, start=def_col + 1):
            val = lookup.get(idx_tuple)
            if val is not None:
                ws.cell(row=cur_row, column=col_idx, value=val)
        cur_row += 1

    # ── Convenience section (two-column gap) ────────────────────
    last_data_col = def_col + len(columns)
    color_end_col = last_data_col + 2  # extend coloring 2 cols beyond data
    _write_param_reference(
        ws, color_end_col, spec, db_contents, header_row=n_header_rows,
        n_data_rows=len(ordered_tuples), layout="stochastic",
        shown_params=list(spec.parameter_names),
    )

    # ── Formatting (extend to color_end_col) ─────────────────────
    format_timeseries_sheet_v2(ws, n_header_rows, single_param=True,
                               row_types=row_types, last_data_col=color_end_col)

    # Index columns: green fill
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.utils import get_column_letter
    data_end = cur_row + 100
    for i in range(n_index_cols):
        cl = get_column_letter(i + 1)
        ws.cell(row=n_header_rows, column=i + 1).fill = FILL_INDEX_HEADER
        ws.conditional_formatting.add(
            f"{cl}{n_header_rows + 1}:{cl}{data_end}",
            CellIsRule(operator="notEqual", formula=['"§§§NEVER§§§"'],
                       fill=FILL_INDEX_DATA),
        )

    # Definition column: dark grey
    def_cl = get_column_letter(def_col)
    ws.conditional_formatting.add(
        f"{def_cl}{n_header_rows + 1}:{def_cl}{data_end}",
        CellIsRule(operator="notEqual", formula=['"§§§NEVER§§§"'],
                   fill=FILL_DEF_COL),
    )

    add_navigate_link(ws)
    auto_column_width(ws, min_param_width=_min_param_width,
                      non_param_width=_non_param_width,
                      def_col_width=_def_col_width,
                      index_col_width=_index_col_width,
                      header_row=n_header_rows, def_col=def_col,
                      last_data_col=last_data_col)


# ---------------------------------------------------------------------------
# Navigate sheet
# ---------------------------------------------------------------------------


def write_navigate_sheet(
    ws: Worksheet,
    all_specs: list[SheetSpec],
    navigate_groups: list[dict[str, Any]] | None = None,
    version: float | None = None,
) -> None:
    """Write the navigate sheet with grouped, colour-coded hyperlinks.

    Args:
        ws: Target worksheet.
        all_specs: All sheet specifications (used for fallback if no groups).
        navigate_groups: Navigate group configuration from export_settings.yaml.
        version: FlexTool DB version to display.
    """
    # Build set of sheet names that actually exist in the workbook specs
    existing_sheets: set[str] = {spec.sheet_name for spec in all_specs if spec.layout != "navigate"}

    if not navigate_groups:
        # Fallback: flat 3-column list (original behaviour)
        _write_navigate_flat(ws, all_specs, existing_sheets)
        return

    # --- Row 1: column headers ---
    ws.cell(row=1, column=1, value="Constants")
    ws.cell(row=1, column=2, value="Periodic")
    ws.cell(row=1, column=3, value="Timeseries")
    ws.cell(row=1, column=4, value="Stochastic")

    # --- Column F: help text ---
    help_lines = [
        "FlexTool can take parameter data in four ways:",
        "- constants (sheet ends with '_c') - Almost any value can be a constant or can be set as a time series.",
        "- period series (sheets ends with '_p') - These parameters have a period index (e.g. year).",
        "- time series (sheets ends with '_t') - These parameters can have values for each timestep.",
        "- stochastic (sheets ends with '_s') - These parameters have forecast branches with time series.",
        "",
        "Some sheets establish only relationships between entities without additional parameters.",
        "",
        "Parameters with valid types (drop-down lists) show allowed values when you select the cell.",
    ]
    for i, text in enumerate(help_lines):
        if text:
            ws.cell(row=1 + i, column=6, value=text)

    # --- Version info at E15 ---
    if version is not None:
        version_int = int(version) if version == int(version) else version
        ws.cell(row=15, column=6, value="FlexTool DB version:")
        ws.cell(row=15, column=7, value=version_int)

    # --- Write grouped rows ---
    current_row = 2  # start after header row

    for group in navigate_groups:
        group_color = group["color"]
        group_fill = PatternFill(patternType="solid", fgColor=group_color)
        rows_in_group = group.get("rows", [])

        for sheet_row in rows_in_group:
            # Place each sheet name into columns A, B, C in order
            for col_offset, sheet_name in enumerate(sheet_row):
                if sheet_name in existing_sheets:
                    col = col_offset + 1  # 1-based
                    cell = ws.cell(row=current_row, column=col, value=sheet_name)
                    cell.hyperlink = f"#{sheet_name}!A1"
                    cell.font = FONT_NAVIGATE_LINK

            # Apply group fill to columns A, B, C for this row
            for col in range(1, 5):
                ws.cell(row=current_row, column=col).fill = group_fill

            current_row += 1

        # Blank separator row after each group
        current_row += 1

    # --- Sheet mappings section (documents multi-mapping entity classes) ---
    mapping_specs = [s for s in all_specs
                     if s.direction_map and len(s.entity_classes) > 1
                     and s.sheet_name in existing_sheets]
    if mapping_specs:
        current_row += 1
        header_fill = PatternFill(patternType="solid", fgColor="BDD7EE")
        ws.cell(row=current_row, column=1, value="Sheet mappings")
        ws.cell(row=current_row, column=1).fill = header_fill
        ws.cell(row=current_row, column=2, value="Entity classes")
        ws.cell(row=current_row, column=2).fill = header_fill
        ws.cell(row=current_row, column=3, value="Filter column")
        ws.cell(row=current_row, column=3).fill = header_fill
        ws.cell(row=current_row, column=4, value="Filter value")
        ws.cell(row=current_row, column=4).fill = header_fill
        ws.cell(row=current_row, column=5, value="Dimensions")
        ws.cell(row=current_row, column=5).fill = header_fill
        ws.cell(row=current_row, column=6, value="Orientation")
        ws.cell(row=current_row, column=6).fill = header_fill
        current_row += 1

        seen_sheets: set[str] = set()
        for spec in mapping_specs:
            if spec.sheet_name in seen_sheets:
                continue
            seen_sheets.add(spec.sheet_name)

            orientation = "rows" if spec.layout == "timeseries" else "columns"
            for cls_name, direction_val in spec.direction_map.items():
                ws.cell(row=current_row, column=1, value=spec.sheet_name)
                ws.cell(row=current_row, column=2, value=cls_name)
                ws.cell(row=current_row, column=3, value=spec.direction_column)
                ws.cell(row=current_row, column=4, value=direction_val)
                dims = ", ".join(spec.entity_columns)
                ws.cell(row=current_row, column=5, value=dims)
                ws.cell(row=current_row, column=6, value=orientation)
                current_row += 1

    auto_column_width(ws)


def _write_navigate_flat(
    ws: Worksheet,
    all_specs: list[SheetSpec],
    existing_sheets: set[str],
) -> None:
    """Fallback: write a flat 3-column navigate sheet (original behaviour)."""
    constants: list[str] = []
    periodic: list[str] = []
    timeseries: list[str] = []

    for spec in all_specs:
        if spec.layout == "navigate":
            continue
        elif spec.layout in ("periodic", "nested_periodic"):
            periodic.append(spec.sheet_name)
        elif spec.layout == "timeseries":
            timeseries.append(spec.sheet_name)
        else:
            constants.append(spec.sheet_name)

    ws.cell(row=1, column=1, value="Constants")
    ws.cell(row=1, column=2, value="Periodic")
    ws.cell(row=1, column=3, value="Timeseries")

    for row_idx, name in enumerate(constants, start=2):
        cell = ws.cell(row=row_idx, column=1, value=name)
        cell.hyperlink = f"#{name}!A1"
        cell.font = FONT_NAVIGATE_LINK

    for row_idx, name in enumerate(periodic, start=2):
        cell = ws.cell(row=row_idx, column=2, value=name)
        cell.hyperlink = f"#{name}!A1"
        cell.font = FONT_NAVIGATE_LINK

    for row_idx, name in enumerate(timeseries, start=2):
        cell = ws.cell(row=row_idx, column=3, value=name)
        cell.hyperlink = f"#{name}!A1"
        cell.font = FONT_NAVIGATE_LINK

    auto_column_width(ws)


# ===========================================================================
# v2 format writers — self-describing sheets with embedded metadata
# ===========================================================================


# Data type overrides from settings, set by export_to_excel before writing.
# Maps (layout, param_name) -> override string.
_data_type_overrides: dict[str, dict[str, str]] = {}
_min_param_width: float = 10
_non_param_width: float = 22
_def_col_width: float = 11
_index_col_width: float = 12


def _get_param_data_type(
    param_name: str,
    entity_classes: list[str],
    db_contents: DatabaseContents,
    layout: str = "constant",
) -> str:
    """Determine the data type label for a parameter on a specific sheet layout.

    The label tells the user what to enter in each cell and helps the reader
    interpret values correctly.

    Returns labels like:
        "float", "string", "boolean (array)",
        "float (1d-map)", "string (1d-map)", "string/float (1d-map)"
    """
    # Check overrides first (e.g. nested_periodic convention for 2d-maps)
    overrides = _data_type_overrides.get(layout, {})
    if param_name in overrides:
        return overrides[param_name]

    for entity_class in entity_classes:
        for pdef in db_contents.parameter_definitions.get(entity_class, []):
            if pdef["name"] == param_name:
                type_list = pdef.get("parameter_type_list")
                if not type_list:
                    return "float"
                types = set(type_list)

                # Pure array parameters — entered as boolean values
                if "array" in types and not (types & {"float", "1d_map", "2d_map", "3d_map", "4d_map"}):
                    return "boolean (array)"

                # On constant sheets: report the scalar type
                if layout == "constant":
                    if "str" in types:
                        return "string"
                    return "float"

                # On periodic sheets: array params are boolean (array index = period column)
                if layout == "periodic" and "array" in types:
                    return "boolean (array)"

                # On periodic/timeseries sheets: report the map dimension
                # relevant to THIS layout, not the highest available.
                # Periodic/timeseries use 1d_map; nested_periodic uses 2d_map;
                # 3d_map/4d_map are for stochastic sheets.
                if layout == "nested_periodic":
                    map_suffix = "2d-map" if "2d_map" in types else "1d-map"
                elif layout in ("periodic", "timeseries"):
                    map_suffix = "1d-map" if "1d_map" in types else ""
                elif "4d_map" in types:
                    map_suffix = "4d-map"
                elif "3d_map" in types:
                    map_suffix = "3d-map"
                elif "2d_map" in types:
                    map_suffix = "2d-map"
                elif "1d_map" in types:
                    map_suffix = "1d-map"
                else:
                    map_suffix = ""

                # Determine inner value type
                has_str = "str" in types
                has_float = "float" in types
                if has_str and not has_float:
                    base = "string"
                elif has_float and not has_str:
                    base = "float"
                elif has_str and has_float:
                    base = "string/float"
                else:
                    # No str or float info (pure map) — allow either
                    base = "string/float"

                if map_suffix:
                    return f"{base} ({map_suffix})"
                return base

    return "float"


def _get_param_value_list(
    param_name: str,
    entity_classes: list[str],
    db_contents: DatabaseContents,
) -> list[str] | None:
    """Get the allowed value list for a parameter, if any."""
    for entity_class in entity_classes:
        for pdef in db_contents.parameter_definitions.get(entity_class, []):
            if pdef["name"] == param_name:
                vl_name = pdef.get("parameter_value_list_name")
                if vl_name and vl_name in db_contents.list_values:
                    return [str(v) for v in db_contents.list_values[vl_name]]
    return None


def _get_param_description(
    param_name: str,
    entity_classes: list[str],
    db_contents: DatabaseContents,
) -> str | None:
    """Get the description for a parameter from DB definitions."""
    for entity_class in entity_classes:
        for pdef in db_contents.parameter_definitions.get(entity_class, []):
            if pdef["name"] == param_name:
                return pdef.get("description") or None
    return None


def _build_entity_def_label(spec: SheetSpec) -> str:
    """Build the 'entity: ...' definition label for the definition row.

    Always uses simplified form with just dimension names.
    Full multi-mapping details are on the navigate sheet.
    """
    return f"entity: {', '.join(spec.entity_columns)}"


def _build_filter_label(spec: SheetSpec) -> str | None:
    """Build the filter column label.

    Returns the direction_column name (e.g. 'input_output') or None.
    Full mapping details are on the navigate sheet.
    """
    if not spec.direction_column or not spec.direction_map:
        return None
    return spec.direction_column


def _add_data_validation(
    ws: Worksheet,
    param_names: list[str],
    def_col: int,
    ea_offset: int,
    spec: SheetSpec,
    db_contents: DatabaseContents,
    max_data_rows: int = 500,
) -> None:
    """Add dropdown validation lists for parameters that have value lists.

    Applies validation from row 4 down to a reasonable limit (not the entire
    column) to avoid bloating the file.
    """
    from openpyxl.worksheet.datavalidation import DataValidation

    last_row = min(4 + max_data_rows, 504)  # cap at ~500 data rows

    from openpyxl.utils import get_column_letter

    # Entity existence column gets TRUE/FALSE dropdown
    if ea_offset > 0:
        ee_col = def_col + 1
        dv_ee = DataValidation(
            type="list",
            formula1='",TRUE,FALSE"',
            allow_blank=True,
            showDropDown=False,
        )
        dv_ee.error = "Select TRUE or FALSE"
        dv_ee.errorTitle = "Entity existence"
        ee_letter = get_column_letter(ee_col)
        dv_ee.sqref = f"{ee_letter}4:{ee_letter}{last_row}"
        ws.add_data_validation(dv_ee)

    for i, pname in enumerate(param_names):
        col = def_col + 1 + ea_offset + i
        values = _get_param_value_list(pname, spec.entity_classes, db_contents)

        # Boolean (array) params get TRUE/FALSE dropdown
        dtype = _get_param_data_type(pname, spec.entity_classes, db_contents, layout="constant")
        if dtype == "boolean (array)" and not values:
            values = ["TRUE", "FALSE"]

        if not values:
            continue
        # Add empty choice so user can clear a selection
        all_values = [""] + values
        formula = ",".join(all_values)
        dv = DataValidation(
            type="list",
            formula1=f'"{formula}"',
            allow_blank=True,
            showDropDown=False,
        )
        dv.error = f"Value must be one of: {', '.join(values)}"
        dv.errorTitle = f"Invalid {pname}"
        col_letter = get_column_letter(col)
        dv.sqref = f"{col_letter}4:{col_letter}{last_row}"
        ws.add_data_validation(dv)


# Period-only params from settings, set by export_to_excel.
_period_only_params: dict[str, list[str]] = {}
_time_structure_classes: set[str] = set()
_info_rows: dict[str, str] = {}
_stochastic_index_names: list[str] = ["forecast", "branch_time", "time", "is_realized"]


def _is_param_valid_for_layout(
    pdef: dict, layout: str, entity_class: str = "",
) -> bool:
    """Check if a parameter is valid for a given sheet layout."""
    type_list = pdef.get("parameter_type_list")
    if not type_list:
        return layout == "constant"
    types = set(type_list)
    pname = pdef["name"]

    if layout == "constant":
        return bool(types & {"str", "float", "array", "2d_map"})
    elif layout == "periodic":
        return bool(types & {"1d_map", "2d_map", "array"})
    elif layout == "timeseries":
        if not (types & {"1d_map", "3d_map", "4d_map"}):
            return False
        # Filter out period-only params
        period_only = _period_only_params.get(entity_class, [])
        if pname in period_only:
            return False
        return True
    return True


def _get_all_param_defs_for_class(
    entity_classes: list[str],
    db_contents: DatabaseContents,
    layout: str = "constant",
) -> list[dict]:
    """Get parameter definitions valid for the given layout, deduplicated by name."""
    seen: set[str] = set()
    result: list[dict] = []
    for entity_class in entity_classes:
        for pdef in db_contents.parameter_definitions.get(entity_class, []):
            if pdef["name"] not in seen and _is_param_valid_for_layout(pdef, layout, entity_class):
                seen.add(pdef["name"])
                result.append(pdef)
    result.sort(key=lambda p: p["name"])
    return result


def _write_param_reference(
    ws: Worksheet,
    start_col: int,
    spec: SheetSpec,
    db_contents: DatabaseContents,
    header_row: int = 3,
    n_data_rows: int = 0,
    layout: str = "constant",
    shown_params: list[str] | None = None,
    row_types: dict[int, str] | None = None,
) -> None:
    """Write reference sections to the right of the data area.

    For transposed sheets with row_types (multi-param timeseries):
      Writes a "quadruplet" section mirroring description/data_type/
      alternative(empty)/parameter rows, plus a convenience section below.

    For standard sheets:
      Writes only the convenience section (parameter name | description).
    """
    if not shown_params:
        return

    # Get the full definitions for the shown params (preserving order)
    all_pdefs_map: dict[str, dict] = {}
    for entity_class in spec.entity_classes:
        for pdef in db_contents.parameter_definitions.get(entity_class, []):
            all_pdefs_map[pdef["name"]] = pdef

    # Get ALL valid params for this layout (for the quadruplet — extra params not shown)
    all_valid = _get_all_param_defs_for_class(spec.entity_classes, db_contents, layout=layout)
    shown_set = set(shown_params)
    extra_pdefs = [p for p in all_valid if p["name"] not in shown_set]

    pdefs = [all_pdefs_map[p] for p in shown_params if p in all_pdefs_map]
    if not pdefs:
        return

    ref_col = start_col + 1  # after empty separator column

    # ------------------------------------------------------------------
    # Quadruplet section for transposed sheets (description, data_type,
    # alternative[empty], parameter) — shows ALL valid params so user
    # can copy columns when adding new entities.
    # ------------------------------------------------------------------
    if row_types and all_valid:
        for i, pdef in enumerate(all_valid):
            col = ref_col + i
            pname = pdef["name"]
            for row_num, rtype in row_types.items():
                if rtype == "description":
                    desc = pdef.get("description", "")
                    if desc:
                        ws.cell(row=row_num, column=col, value=desc)
                elif rtype == "data_type":
                    dtype = _get_param_data_type(pname, spec.entity_classes, db_contents, layout=layout)
                    ws.cell(row=row_num, column=col, value=dtype)
                elif rtype == "alternative":
                    pass  # leave empty — user fills this when copying
                elif rtype == "parameter":
                    ws.cell(row=row_num, column=col, value=pname)
                # entity row: leave empty

    # ------------------------------------------------------------------
    # Convenience section (parameter | data type | description)
    # ------------------------------------------------------------------
    conv_row = header_row + 2  # one empty row after header

    # Header
    ws.cell(row=conv_row, column=ref_col, value="Parameter")
    ws.cell(row=conv_row, column=ref_col).fill = FILL_PARAM_HEADER
    ws.cell(row=conv_row, column=ref_col + 1, value="Data type")
    ws.cell(row=conv_row, column=ref_col + 1).fill = FILL_PARAM_HEADER
    ws.cell(row=conv_row, column=ref_col + 2, value="Description")
    ws.cell(row=conv_row, column=ref_col + 2).fill = FILL_PARAM_HEADER

    max_name_len = len("Parameter")
    for i, pdef in enumerate(pdefs):
        row = conv_row + 1 + i
        name = pdef["name"]
        ws.cell(row=row, column=ref_col, value=name)
        if len(name) > max_name_len:
            max_name_len = len(name)
        dtype = _get_param_data_type(name, spec.entity_classes, db_contents, layout=layout)
        ws.cell(row=row, column=ref_col + 1, value=dtype)
        desc = pdef.get("description")
        if desc:
            ws.cell(row=row, column=ref_col + 2, value=desc)

    # Size the parameter name column and data type column
    from openpyxl.utils import get_column_letter
    col_letter = get_column_letter(ref_col)
    ws.column_dimensions[col_letter].width = max_name_len + 2
    # Data type column: 50% wider than param name column
    dt_col_letter = get_column_letter(ref_col + 1)
    ws.column_dimensions[dt_col_letter].width = int((max_name_len + 2) * 1.5)


def _lock_metadata_cells(
    ws: Worksheet,
    meta_rows: int,
    meta_cols: int,
    data_start_row: int,
    data_start_col: int,
    max_data_rows: int = 500,
) -> None:
    """Apply sheet protection, locking metadata cells and unlocking data cells.

    Uses column-level unlocking for efficiency: sets entire data columns as
    unlocked rather than individual cells.  For the metadata columns in the
    data region, only the metadata rows stay locked (the default).

    Skips protection entirely when the data region exceeds *max_data_rows*
    to avoid bloating the file with per-cell styles.

    Args:
        ws: Target worksheet.
        meta_rows: Number of metadata rows (rows 1..meta_rows are locked).
        meta_cols: Number of metadata columns (columns 1..meta_cols are locked in data rows).
        data_start_row: First row of the data region.
        data_start_col: First column of the data region.
        max_data_rows: Skip protection if data exceeds this many rows.
    """
    # TODO: Re-enable sheet protection once LibreOffice compatibility is verified.
    # The current openpyxl set_password("") creates a hash that can crash
    # LibreOffice on some systems.  Disabled for now.
    return


# ---------------------------------------------------------------------------
# v2 Constant sheet
# ---------------------------------------------------------------------------


def write_constant_sheet_v2(
    ws: Worksheet,
    spec: SheetSpec,
    db_contents: DatabaseContents,
) -> None:
    """Write a constant-layout sheet in the v2 self-describing format.

    Layout:
        Row 1: navigate | (empty) | description | desc/param descriptions...
        Row 2: (empty)  | (empty) | data type   | string/float...
        Row 3: alternative | entity: X | [filter:] [index:] parameter | [entity existence] | param names...
        Row 4+: data
    """
    # Determine the definition column position
    n_entity_cols = len(spec.entity_columns)
    n_extra_entity = len(spec.extra_entity_columns)
    n_dims = n_entity_cols

    # Build left-side column labels for row 3
    left_cols: list[str] = ["alternative"]
    entity_label = _build_entity_def_label(spec)
    left_cols.append(entity_label)

    # For multi-dim entities, add ALL individual dim name columns after the
    # entity label (the entity label column has no data in data rows)
    if n_dims > 1:
        for dim_name in spec.entity_columns:
            left_cols.append(dim_name)

    # Extra entity columns (e.g. left_node, right_node for connection)
    for extra_col in spec.extra_entity_columns:
        left_cols.append(extra_col)

    # Filter column for direction
    filter_label = _build_filter_label(spec)
    filter_col_pos: int | None = None
    if filter_label:
        left_cols.append(filter_label)
        filter_col_pos = len(left_cols)  # 1-based

    # Unpack index column
    index_col_positions: set[int] = set()
    if spec.unpack_index_column:
        left_cols.append(f"index: {spec.unpack_index_column}")
        index_col_positions.add(len(left_cols))  # 1-based

    def_col = len(left_cols) + 1  # 1-based position of definition column

    # Build right-side columns (from def_col onward)
    right_cols: list[str] = ["parameter"]
    if spec.has_entity_alternative:
        right_cols.append("entity existence")
    # All parameters (pre_ea_params + parameter_names)
    all_params = list(spec.pre_ea_params) + list(spec.parameter_names)
    right_cols.extend(all_params)

    # --- Row 1: navigate + descriptions (from def_col onward) ---
    ws.cell(row=1, column=1, value="navigate")
    ws.cell(row=1, column=def_col, value="description")
    ea_offset = 1 if spec.has_entity_alternative else 0
    if spec.has_entity_alternative:
        ws.cell(row=1, column=def_col + 1, value="Entity existence")
    for i, pname in enumerate(all_params):
        col = def_col + 1 + ea_offset + i
        desc = spec.descriptions.get(pname)
        if not desc:
            desc = _get_param_description(pname, spec.entity_classes, db_contents)
        if desc:
            ws.cell(row=1, column=col, value=desc)

    # --- Row 2: data types (from def_col onward) ---
    ws.cell(row=2, column=def_col, value="data type")
    if spec.has_entity_alternative:
        ws.cell(row=2, column=def_col + 1, value="string")
    for i, pname in enumerate(all_params):
        col = def_col + 1 + ea_offset + i
        dtype = _get_param_data_type(pname, spec.entity_classes, db_contents, layout="constant")
        ws.cell(row=2, column=col, value=dtype)

    # --- Row 3: definition row ---
    for col_idx, label in enumerate(left_cols, start=1):
        ws.cell(row=3, column=col_idx, value=label)
    for col_idx, label in enumerate(right_cols, start=def_col):
        ws.cell(row=3, column=col_idx, value=label)

    # --- Map column positions for data writing ---
    col_alt = 1

    if n_dims <= 1:
        entity_data_cols = [2]
        next_col_after_entity = 3
    else:
        # Multi-dim: entity label in col 2 (no data in data rows),
        # dims in cols 3..2+n_dims
        entity_data_cols = list(range(3, 3 + n_dims))
        next_col_after_entity = 3 + n_dims

    extra_entity_data_cols = list(
        range(next_col_after_entity, next_col_after_entity + n_extra_entity)
    )
    next_col = next_col_after_entity + n_extra_entity

    filter_data_col: int | None = None
    if filter_label:
        filter_data_col = next_col
        next_col += 1

    unpack_data_col: int | None = None
    if spec.unpack_index_column:
        unpack_data_col = next_col
        next_col += 1

    # def_col should match
    assert next_col == def_col, f"Column mismatch: expected def_col={def_col}, got {next_col}"

    ea_data_col: int | None = None
    param_data_start = def_col + 1
    if spec.has_entity_alternative:
        ea_data_col = def_col + 1
        param_data_start = def_col + 2

    param_col_map: dict[str, int] = {}
    for i, pname in enumerate(all_params):
        param_col_map[pname] = param_data_start + i

    # --- Collect data rows ---
    data_rows: list[list[tuple[int, Any]]] = []

    if spec.unpack_index_column:
        _collect_unpack_rows_v2(
            data_rows, spec, db_contents,
            col_alt=col_alt,
            entity_data_cols=entity_data_cols,
            extra_entity_data_cols=extra_entity_data_cols,
            filter_data_col=filter_data_col,
            unpack_data_col=unpack_data_col,
            ea_data_col=ea_data_col,
            param_col_map=param_col_map,
        )
    else:
        _collect_constant_rows_v2(
            data_rows, spec, db_contents,
            col_alt=col_alt,
            entity_data_cols=entity_data_cols,
            extra_entity_data_cols=extra_entity_data_cols,
            filter_data_col=filter_data_col,
            ea_data_col=ea_data_col,
            param_col_map=param_col_map,
        )

    # --- Sort rows ---
    data_rows.sort(key=lambda r: tuple(
        str(v) if v is not None else "" for _, v in r
    ))

    # --- Write data rows (starting at row 4) ---
    for row_idx, row_cells in enumerate(data_rows, start=4):
        for col, value in row_cells:
            if value is not None:
                ws.cell(row=row_idx, column=col, value=value)

    # --- Write parameter reference section ---
    last_data_col = def_col + len(right_cols) - 1
    _write_param_reference(
        ws, last_data_col + 1, spec, db_contents, header_row=3,
        n_data_rows=len(data_rows), layout="constant",
        shown_params=all_params,
    )

    # --- Formatting ---
    n_extra = (
        n_extra_entity
        + (1 if filter_label else 0)
        + (1 if spec.unpack_index_column else 0)
    )
    format_constant_sheet_v2(ws, n_entity_cols, n_extra, def_col, index_col_positions, last_data_col)
    add_navigate_link(ws)

    # --- Data validation (dropdown lists) for parameters with value lists ---
    _add_data_validation(ws, all_params, def_col, ea_offset, spec, db_contents,
                         max_data_rows=max(len(data_rows), 1) + 3)

    # --- Filter dropdown for merged classes ---
    if filter_col_pos is not None and spec.direction_map:
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
        filter_values = sorted(set(spec.direction_map.values()))
        filter_list = ",".join([""] + filter_values)
        dv_filter = DataValidation(
            type="list",
            formula1=f'"{filter_list}"',
            allow_blank=True,
            showDropDown=False,
        )
        dv_filter.error = f"Select: {', '.join(filter_values)}"
        dv_filter.errorTitle = spec.direction_column or "Filter"
        cl = get_column_letter(filter_col_pos)
        last_row = max(len(data_rows) + 3, 504)
        dv_filter.sqref = f"{cl}4:{cl}{last_row}"
        ws.add_data_validation(dv_filter)

    auto_column_width(ws, min_param_width=_min_param_width,
                      non_param_width=_non_param_width,
                      def_col_width=_def_col_width,
                      index_col_width=_index_col_width,
                      header_row=3, def_col=def_col,
                      index_cols=index_col_positions,
                      last_data_col=last_data_col)

    # --- Lock metadata ---
    _lock_metadata_cells(
        ws,
        meta_rows=3,
        meta_cols=def_col,
        data_start_row=4,
        data_start_col=1,
    )


def _collect_constant_rows_v2(
    data_rows: list[list[tuple[int, Any]]],
    spec: SheetSpec,
    db_contents: DatabaseContents,
    *,
    col_alt: int,
    entity_data_cols: list[int],
    extra_entity_data_cols: list[int],
    filter_data_col: int | None,
    ea_data_col: int | None,
    param_col_map: dict[str, int],
) -> None:
    """Collect rows for a v2 constant sheet (non-unpack)."""
    all_params = list(spec.pre_ea_params) + list(spec.parameter_names)

    for entity_class in spec.entity_classes:
        entities = db_contents.entities.get(entity_class, [])
        direction = _get_direction(entity_class, spec)

        for entity in entities:
            entity_byname = entity["entity_byname"]
            alts = _find_alternatives_for_entity(
                entity_class, entity_byname, spec, db_contents
            )

            for alt in alts:
                cells: list[tuple[int, Any]] = []

                # Alternative
                cells.append((col_alt, alt))

                # Entity dimension values
                for i, col in enumerate(entity_data_cols):
                    if i < len(entity_byname):
                        cells.append((col, str(_to_native(entity_byname[i]))))

                # Extra entity columns
                if spec.extra_entity_columns:
                    extras = _get_extra_columns_for_entity(
                        entity_byname, entity_class, spec, db_contents
                    )
                    for i, col in enumerate(extra_entity_data_cols):
                        if i < len(extras):
                            cells.append((col, extras[i]))

                # Direction/filter value
                if filter_data_col is not None and direction is not None:
                    cells.append((filter_data_col, direction))

                # Entity existence
                if ea_data_col is not None:
                    ea_key = (entity_class, entity_byname, alt)
                    ea_val = db_contents.entity_alternatives.get(ea_key)
                    if ea_val is not None:
                        cells.append((ea_data_col, "TRUE" if ea_val else "FALSE"))

                # Parameter values — check for Arrays
                param_values: dict[str, Any] = {}
                has_array = False
                array_param: str | None = None
                array_elements: list[Any] = []

                for pname in all_params:
                    key = (entity_class, entity_byname, pname, alt)
                    value = db_contents.parameter_values.get(key)

                    if value is None:
                        param_values[pname] = None
                    elif _is_map(value):
                        param_values[pname] = None
                    elif _is_array(value):
                        has_array = True
                        array_param = pname
                        array_elements = [_to_native(v) for v in value.values]
                        param_values[pname] = None
                    elif _is_scalar(value):
                        param_values[pname] = _to_native(value)
                    else:
                        param_values[pname] = None

                if has_array and array_elements and array_param is not None:
                    for elem in array_elements:
                        expanded_cells = list(cells)
                        for pname in all_params:
                            col = param_col_map[pname]
                            if pname == array_param:
                                expanded_cells.append((col, elem))
                            elif param_values[pname] is not None:
                                expanded_cells.append((col, param_values[pname]))
                        data_rows.append(expanded_cells)
                else:
                    for pname in all_params:
                        if param_values[pname] is not None:
                            cells.append((param_col_map[pname], param_values[pname]))
                    data_rows.append(cells)


def _collect_unpack_rows_v2(
    data_rows: list[list[tuple[int, Any]]],
    spec: SheetSpec,
    db_contents: DatabaseContents,
    *,
    col_alt: int,
    entity_data_cols: list[int],
    extra_entity_data_cols: list[int],
    filter_data_col: int | None,
    unpack_data_col: int | None,
    ea_data_col: int | None,
    param_col_map: dict[str, int],
) -> None:
    """Collect rows for a v2 unpacked-Map constant sheet."""
    for entity_class in spec.entity_classes:
        entities = db_contents.entities.get(entity_class, [])
        direction = _get_direction(entity_class, spec)

        for entity in entities:
            entity_byname = entity["entity_byname"]
            alts = _find_alternatives_for_entity(
                entity_class, entity_byname, spec, db_contents
            )

            for alt in alts:
                # Collect all map indexes across all unpack params
                all_indexes: set[str] = set()
                param_maps: dict[str, Map] = {}

                for pname in spec.parameter_names:
                    key = (entity_class, entity_byname, pname, alt)
                    value = db_contents.parameter_values.get(key)
                    if value is not None and _is_map(value):
                        param_maps[pname] = value
                        for idx in value.indexes:
                            all_indexes.add(str(_to_native(idx)))

                if not all_indexes:
                    continue

                for idx_val in sorted(all_indexes):
                    cells: list[tuple[int, Any]] = []

                    # Alternative
                    cells.append((col_alt, alt))

                    # Entity dimension values
                    for i, col in enumerate(entity_data_cols):
                        if i < len(entity_byname):
                            cells.append((col, str(_to_native(entity_byname[i]))))

                    # Extra entity columns
                    if spec.extra_entity_columns:
                        extras = _get_extra_columns_for_entity(
                            entity_byname, entity_class, spec, db_contents
                        )
                        for i, col in enumerate(extra_entity_data_cols):
                            if i < len(extras):
                                cells.append((col, extras[i]))

                    # Direction/filter value
                    if filter_data_col is not None and direction is not None:
                        cells.append((filter_data_col, direction))

                    # Unpack index value
                    if unpack_data_col is not None:
                        cells.append((unpack_data_col, idx_val))

                    # Entity existence
                    if ea_data_col is not None:
                        ea_key = (entity_class, entity_byname, alt)
                        ea_val = db_contents.entity_alternatives.get(ea_key)
                        if ea_val is not None:
                            cells.append((ea_data_col, ea_val))

                    # Parameter values from maps
                    for pname in spec.parameter_names:
                        m = param_maps.get(pname)
                        if m is not None:
                            for mi, mv in zip(m.indexes, m.values):
                                if str(_to_native(mi)) == idx_val:
                                    cells.append((param_col_map[pname], _to_native(mv)))
                                    break

                    # Pre-EA scalar params (these are not maps)
                    for pname in spec.pre_ea_params:
                        key = (entity_class, entity_byname, pname, alt)
                        value = db_contents.parameter_values.get(key)
                        if value is not None and _is_scalar(value):
                            cells.append((param_col_map[pname], _to_native(value)))

                    data_rows.append(cells)


# ---------------------------------------------------------------------------
# v2 Periodic sheet
# ---------------------------------------------------------------------------


def write_periodic_sheet_v2(
    ws: Worksheet,
    spec: SheetSpec,
    db_contents: DatabaseContents,
) -> None:
    """Write a periodic-layout sheet in the v2 self-describing format.

    Layout:
        Row 1: navigate | | | description | param descriptions...
        Row 2: | | | data type | float...
        Row 3: alternative | entity: X | [filter:] index: period | parameter | param names...
        Row 4+: data
    """
    # Determine index column name
    index_col_name = _find_index_column_name(spec, db_contents, default="period")

    n_entity_cols = len(spec.entity_columns)
    n_dims = n_entity_cols

    # Build left-side columns for row 3
    left_cols: list[str] = ["alternative"]
    entity_label = _build_entity_def_label(spec)
    left_cols.append(entity_label)

    # For multi-dim, add ALL individual dim name columns after entity label
    if n_dims > 1:
        for dim_name in spec.entity_columns:
            left_cols.append(dim_name)

    # Direction/filter column
    filter_label = _build_filter_label(spec)
    if filter_label:
        left_cols.append(filter_label)

    # Index column (period)
    left_cols.append(f"index: {index_col_name}")
    index_col_pos = len(left_cols)  # 1-based

    def_col = len(left_cols) + 1

    # Right-side columns
    right_cols: list[str] = ["parameter"]
    right_cols.extend(spec.parameter_names)

    # --- Row 1: navigate + descriptions ---
    ws.cell(row=1, column=1, value="navigate")
    ws.cell(row=1, column=def_col, value="description")
    for i, pname in enumerate(spec.parameter_names):
        col = def_col + 1 + i
        desc = spec.descriptions.get(pname)
        if not desc:
            desc = _get_param_description(pname, spec.entity_classes, db_contents)
        if desc:
            ws.cell(row=1, column=col, value=desc)

    # --- Row 2: data types ---
    ws.cell(row=2, column=def_col, value="data type")
    for i, pname in enumerate(spec.parameter_names):
        col = def_col + 1 + i
        dtype = _get_param_data_type(pname, spec.entity_classes, db_contents, layout="periodic")
        ws.cell(row=2, column=col, value=dtype)

    # --- Row 3: definition row ---
    for col_idx, label in enumerate(left_cols, start=1):
        ws.cell(row=3, column=col_idx, value=label)
    for col_idx, label in enumerate(right_cols, start=def_col):
        ws.cell(row=3, column=col_idx, value=label)

    # --- Map column positions ---
    col_alt = 1

    if n_dims <= 1:
        entity_data_cols = [2]
        next_col = 3
    else:
        entity_data_cols = list(range(3, 3 + n_dims))
        next_col = 3 + n_dims

    filter_data_col: int | None = None
    if filter_label:
        filter_data_col = next_col
        next_col += 1

    index_data_col = next_col
    next_col += 1

    assert next_col == def_col, f"Column mismatch: expected def_col={def_col}, got {next_col}"

    param_data_start = def_col + 1
    param_col_map: dict[str, int] = {}
    for i, pname in enumerate(spec.parameter_names):
        param_col_map[pname] = param_data_start + i

    # --- Collect data rows ---
    data_rows: list[list[tuple[int, Any]]] = []

    for entity_class in spec.entity_classes:
        entities = db_contents.entities.get(entity_class, [])
        direction = _get_direction(entity_class, spec)

        for entity in entities:
            entity_byname = entity["entity_byname"]
            alts = _find_alternatives_for_entity(
                entity_class, entity_byname, spec, db_contents
            )

            for alt in alts:
                all_indexes: set[str] = set()
                param_maps: dict[str, Map] = {}
                param_arrays: dict[str, list[str]] = {}

                for pname in spec.parameter_names:
                    key = (entity_class, entity_byname, pname, alt)
                    value = db_contents.parameter_values.get(key)
                    if value is not None and _is_map(value):
                        is_nested = any(_is_map(mv) for mv in value.values)
                        if is_nested:
                            continue
                        # Skip time-indexed maps on periodic sheets — unless
                        # the entity class defines time structure itself
                        if _is_time_indexed_map(value) and entity_class not in _time_structure_classes:
                            continue
                        param_maps[pname] = value
                        for idx in value.indexes:
                            all_indexes.add(str(_to_native(idx)))
                    elif value is not None and _is_array(value):
                        period_names = [str(_to_native(v)) for v in value.values]
                        param_arrays[pname] = period_names

                # If no Map indexes but we have Array values, use the array
                # values as indexes (they contain period names)
                if not all_indexes and param_arrays:
                    for arr_vals in param_arrays.values():
                        all_indexes.update(arr_vals)

                if not all_indexes:
                    continue

                # Build dict lookups from Maps for O(1) access (avoids O(n²))
                param_map_dicts: dict[str, dict[str, Any]] = {}
                for pname, m in param_maps.items():
                    param_map_dicts[pname] = {
                        str(_to_native(mi)): _to_native(mv)
                        for mi, mv in zip(m.indexes, m.values)
                    }

                for idx_val in sorted(all_indexes):
                    cells: list[tuple[int, Any]] = []

                    cells.append((col_alt, alt))

                    for i, col in enumerate(entity_data_cols):
                        if i < len(entity_byname):
                            cells.append((col, str(_to_native(entity_byname[i]))))

                    if filter_data_col is not None and direction is not None:
                        cells.append((filter_data_col, direction))

                    cells.append((index_data_col, idx_val))

                    for pname in spec.parameter_names:
                        md = param_map_dicts.get(pname)
                        if md is not None:
                            val = md.get(idx_val)
                            if val is not None:
                                cells.append((param_col_map[pname], val))
                        elif pname in param_arrays:
                            if idx_val in param_arrays[pname]:
                                cells.append((param_col_map[pname], "TRUE"))

                    data_rows.append(cells)

    # --- Sort rows ---
    data_rows.sort(key=lambda r: tuple(
        str(v) if v is not None else "" for _, v in r
    ))

    # --- Write data rows ---
    for row_idx, row_cells in enumerate(data_rows, start=4):
        for col, value in row_cells:
            if value is not None:
                ws.cell(row=row_idx, column=col, value=value)

    # --- Write parameter reference section ---
    last_data_col = def_col + len(right_cols) - 1
    _write_param_reference(
        ws, last_data_col + 1, spec, db_contents, header_row=3,
        n_data_rows=len(data_rows), layout="periodic",
        shown_params=list(spec.parameter_names),
    )

    # --- Formatting ---
    n_extra = (
        (1 if filter_label else 0)
        + 1  # index column
    )
    format_periodic_sheet_v2(ws, n_entity_cols, n_extra, def_col, {index_col_pos}, last_data_col)
    add_navigate_link(ws)
    auto_column_width(ws, min_param_width=_min_param_width,
                      non_param_width=_non_param_width,
                      def_col_width=_def_col_width,
                      index_col_width=_index_col_width,
                      header_row=3, def_col=def_col,
                      index_cols={index_col_pos},
                      last_data_col=last_data_col)

    # --- Lock metadata ---
    _lock_metadata_cells(
        ws,
        meta_rows=3,
        meta_cols=def_col,
        data_start_row=4,
        data_start_col=1,
    )


# ---------------------------------------------------------------------------
# v2 Timeseries sheet
# ---------------------------------------------------------------------------


def write_timeseries_sheet_v2(
    ws: Worksheet,
    spec: SheetSpec,
    db_contents: DatabaseContents,
) -> None:
    """Write a transposed timeseries-layout sheet in the v2 self-describing format.

    All _t sheets are transposed: time index in column A, data in columns C+.
    Header rows follow the canonical order, applied to whichever dimensions
    are needed:

        description → data_type → entity → entity_dims → filter →
        alternative → parameter

    For single-param sheets, description/data_type/parameter collapse into
    a single 'triplet row' in B1 using ``|`` separator.
    """
    # ── Collect data columns ──────────────────────────────────────
    columns: list[tuple[str, tuple, str, str, Map]] = []
    all_time_indexes: set[str] = set()

    for entity_class in spec.entity_classes:
        entities = db_contents.entities.get(entity_class, [])
        for entity in entities:
            entity_byname = entity["entity_byname"]
            for pname in spec.parameter_names:
                for alt in db_contents.alternatives:
                    key = (entity_class, entity_byname, pname, alt)
                    value = db_contents.parameter_values.get(key)
                    if value is not None and _is_map(value):
                        if not _is_time_indexed_map(value):
                            continue
                        columns.append(
                            (entity_class, entity_byname, pname, alt, value)
                        )
                        for idx in value.indexes:
                            all_time_indexes.add(str(_to_native(idx)))

    columns.sort(key=lambda c: (c[3], c[1], c[2]))
    sorted_times = sorted(all_time_indexes)

    single_param = len(spec.parameter_names) == 1
    n_entity_dims = len(spec.entity_columns)
    has_direction = spec.direction_column is not None
    entity_label = _build_entity_def_label(spec)

    # ── Build header rows following canonical order ───────────────
    row_types: dict[int, str] = {}
    cur_row = 1

    # INFO row for all transposed sheets
    info_text = _info_rows.get(spec.sheet_name)
    if not info_text:
        info_text = "INFO: Add new data columns as needed by right clicking on the column name and then selecting 'Insert columns before'."
    ws.cell(row=cur_row, column=2, value=info_text)
    cur_row += 1

    if single_param:
        # Triplet row: combined parameter | data_type | description
        default_param = spec.parameter_names[0]
        dtype = _get_param_data_type(default_param, spec.entity_classes, db_contents, layout="timeseries")
        desc = _get_param_description(default_param, spec.entity_classes, db_contents) or ""
        triplet = f"parameter: {default_param} | data type: {dtype}"
        if desc:
            triplet += f" | description: {desc}"
        ws.cell(row=cur_row, column=2, value=triplet)
        row_types[cur_row] = "param_info"
        cur_row += 1

    # entity row — simplified label (just dimension names, not full mapping)
    if n_entity_dims <= 1:
        simple_label = f"entity: {spec.entity_columns[0]}"
        ws.cell(row=cur_row, column=2, value=simple_label)
        for col_idx, (ec, byname, pname, alt, _m) in enumerate(columns, start=3):
            if byname:
                ws.cell(row=cur_row, column=col_idx, value=str(_to_native(byname[0])))
        row_types[cur_row] = "entity"
        cur_row += 1
    else:
        # Multi-dim: one row per dimension
        for dim_idx in range(n_entity_dims):
            if dim_idx == 0:
                ws.cell(row=cur_row, column=2, value=f"entity: {', '.join(spec.entity_columns)}")
            else:
                ws.cell(row=cur_row, column=2, value=spec.entity_columns[dim_idx])
            for col_idx, (ec, byname, pname, alt, _m) in enumerate(columns, start=3):
                if dim_idx < len(byname):
                    ws.cell(row=cur_row, column=col_idx, value=str(_to_native(byname[dim_idx])))
            row_types[cur_row] = "entity"
            cur_row += 1

    # filter row (for merged classes with direction) — simple label + dropdown
    filter_row_num = None
    if has_direction:
        filter_row_num = cur_row
        ws.cell(row=cur_row, column=2, value=spec.direction_column)
        filter_values = sorted(set(spec.direction_map.values()))
        for col_idx, (ec, byname, pname, alt, _m) in enumerate(columns, start=3):
            direction = _get_direction(ec, spec)
            if direction:
                ws.cell(row=cur_row, column=col_idx, value=direction)
        row_types[cur_row] = "entity"
        cur_row += 1

    # alternative row
    ws.cell(row=cur_row, column=2, value="alternative")
    for col_idx, (ec, byname, pname, alt, _m) in enumerate(columns, start=3):
        ws.cell(row=cur_row, column=col_idx, value=alt)
    row_types[cur_row] = "alternative"
    cur_row += 1

    # parameter row (multi-param only)
    param_row_num = None
    if not single_param:
        param_row_num = cur_row
        ws.cell(row=cur_row, column=2, value="parameter")
        for col_idx, (ec, byname, pname, alt, _m) in enumerate(columns, start=3):
            ws.cell(row=cur_row, column=col_idx, value=pname)
        row_types[cur_row] = "parameter"
        cur_row += 1

    n_header_rows = cur_row - 1

    # index: time label on column A of the last header row
    ws.cell(row=n_header_rows, column=1, value="index: time")

    # ── Write time data rows ─────────────────────────────────────
    col_index_maps: list[dict[str, Any]] = []
    for ec, byname, pname, alt, m in columns:
        idx_map: dict[str, Any] = {}
        for mi, mv in zip(m.indexes, m.values):
            idx_map[str(_to_native(mi))] = _to_native(mv)
        col_index_maps.append(idx_map)

    for time_row_idx, time_val in enumerate(sorted_times):
        row = n_header_rows + 1 + time_row_idx
        ws.cell(row=row, column=1, value=time_val)
        for col_idx, idx_map in enumerate(col_index_maps):
            value = idx_map.get(time_val)
            if value is not None:
                ws.cell(row=row, column=col_idx + 3, value=value)

    add_navigate_link(ws)

    # ── Reference section (two-column gap, no triplet) ──────────
    last_data_col = max(len(columns) + 2, 3)
    # Extend coloring to include the two-column gap before convenience
    color_end_col = last_data_col + 2
    _write_param_reference(
        ws, color_end_col, spec, db_contents, header_row=n_header_rows,
        n_data_rows=len(sorted_times), layout="timeseries",
        shown_params=list(spec.parameter_names),
    )

    # ── Dropdowns ─────────────────────────────────────────────────
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    # Parameter dropdown for multi-param transposed sheets
    if param_row_num is not None and len(spec.parameter_names) > 1:
        param_list = ",".join([""] + list(spec.parameter_names))
        dv = DataValidation(
            type="list",
            formula1=f'"{param_list}"',
            allow_blank=True,
            showDropDown=False,
        )
        dv.error = "Select a parameter"
        dv.errorTitle = "Parameter"
        first_col = get_column_letter(3)
        last_col = get_column_letter(color_end_col)
        dv.sqref = f"{first_col}{param_row_num}:{last_col}{param_row_num}"
        ws.add_data_validation(dv)

    # Filter dropdown for merged classes (input/output etc.)
    if filter_row_num is not None and has_direction:
        filter_values = sorted(set(spec.direction_map.values()))
        filter_list = ",".join([""] + filter_values)
        dv_filter = DataValidation(
            type="list",
            formula1=f'"{filter_list}"',
            allow_blank=True,
            showDropDown=False,
        )
        dv_filter.error = f"Select: {', '.join(filter_values)}"
        dv_filter.errorTitle = spec.direction_column
        first_col = get_column_letter(3)
        last_col = get_column_letter(color_end_col)
        dv_filter.sqref = f"{first_col}{filter_row_num}:{last_col}{filter_row_num}"
        ws.add_data_validation(dv_filter)

    # ── Formatting (extend to color_end_col to cover gap) ─────────
    format_timeseries_sheet_v2(ws, n_header_rows, single_param=single_param,
                               row_types=row_types, last_data_col=color_end_col)

    # Column A: index fill (green, consistent with other index columns)
    from openpyxl.formatting.rule import CellIsRule
    data_end = max(n_header_rows + len(sorted_times), n_header_rows + 1) + 100
    ws.cell(row=n_header_rows, column=1).fill = FILL_INDEX_HEADER
    ws.conditional_formatting.add(
        f"A{n_header_rows + 1}:A{data_end}",
        CellIsRule(operator="notEqual", formula=['"§§§NEVER§§§"'], fill=FILL_INDEX_DATA),
    )
    # Column B: dark grey definition column
    ws.conditional_formatting.add(
        f"B{n_header_rows + 1}:B{data_end}",
        CellIsRule(operator="notEqual", formula=['"§§§NEVER§§§"'], fill=FILL_DEF_COL),
    )

    auto_column_width(ws, min_param_width=_min_param_width,
                      non_param_width=_non_param_width,
                      def_col_width=_def_col_width,
                      index_col_width=_index_col_width,
                      header_row=n_header_rows, def_col=3,
                      last_data_col=last_data_col)

    _lock_metadata_cells(
        ws, meta_rows=n_header_rows, meta_cols=2,
        data_start_row=n_header_rows + 1, data_start_col=1,
    )


# ---------------------------------------------------------------------------
# v2 Link sheet
# ---------------------------------------------------------------------------


def write_link_sheet_v2(
    ws: Worksheet,
    spec: SheetSpec,
    db_contents: DatabaseContents,
) -> None:
    """Write a link-only (relationship) sheet in the v2 self-describing format.

    Layout:
        Row 1: navigate
        Row 2: 'entity: X, Y' | dim1_name | dim2_name
        Row 3+: (empty)       | dim1_val  | dim2_val
    """
    entity_class = spec.entity_classes[0] if spec.entity_classes else None

    # Row 1: navigate
    ws.cell(row=1, column=1, value="navigate")

    # Row 2: entity definition + dimension headers
    entity_label = _build_entity_def_label(spec)
    ws.cell(row=2, column=1, value=entity_label)
    for col_idx, col_name in enumerate(spec.entity_columns, start=2):
        ws.cell(row=2, column=col_idx, value=col_name)

    if not entity_class:
        format_link_sheet_v2(ws)
        add_navigate_link(ws)
        auto_column_width(ws, non_param_width=_non_param_width,
                      def_col_width=_def_col_width,
                      index_col_width=_index_col_width)
        return

    # Collect and sort entity rows
    entities = db_contents.entities.get(entity_class, [])
    rows: list[tuple] = []
    for entity in entities:
        byname = entity["entity_byname"]
        rows.append(tuple(str(_to_native(v)) for v in byname))

    rows.sort()

    for row_idx, row_data in enumerate(rows, start=3):
        for col_idx, value in enumerate(row_data, start=2):
            ws.cell(row=row_idx, column=col_idx, value=value)

    format_link_sheet_v2(ws)
    add_navigate_link(ws)
    auto_column_width(ws, non_param_width=_non_param_width,
                      def_col_width=_def_col_width,
                      index_col_width=_index_col_width)

    # --- Lock metadata ---
    _lock_metadata_cells(
        ws,
        meta_rows=2,
        meta_cols=1,
        data_start_row=3,
        data_start_col=2,
    )
